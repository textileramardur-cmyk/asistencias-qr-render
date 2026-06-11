import hashlib
import io
import json
import os
import re
import secrets
import unicodedata
import zipfile
from datetime import datetime, date, time, timedelta
from urllib.parse import quote
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import qrcode
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from PIL import Image, ImageDraw, ImageFont
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Connection

APP_NAME = "Control QR Asistencias"
TZ = ZoneInfo("America/Mexico_City")
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("DATA_DIR", BASE_DIR / "data"))
UPLOADS_DIR = DATA_DIR / "uploads"
CORRECTIONS_DIR = DATA_DIR / "correcciones"
EMPLOYEE_PHOTOS_DIR = UPLOADS_DIR / "empleados"
VEHICLE_PHOTOS_DIR = UPLOADS_DIR / "vehiculos"
DB_PATH = DATA_DIR / "asistencias.db"

AUTH_COOKIE_NAME = "asistencias_session"
SESSION_DAYS = int(os.getenv("SESSION_DAYS", "14"))
SESSION_COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "1") == "1"
LOGIN_MAX_ATTEMPTS = int(os.getenv("LOGIN_MAX_ATTEMPTS", "6"))
LOGIN_LOCK_MINUTES = int(os.getenv("LOGIN_LOCK_MINUTES", "15"))
MIN_EXIT_AFTER_ENTRY_MINUTES = int(os.getenv("MIN_EXIT_AFTER_ENTRY_MINUTES", "5"))
EXIT_WINDOW_BEFORE_MINUTES = int(os.getenv("EXIT_WINDOW_BEFORE_MINUTES", "180"))
EXIT_WINDOW_AFTER_MINUTES = int(os.getenv("EXIT_WINDOW_AFTER_MINUTES", "240"))
ACCIDENTAL_EXIT_ENTRY_BEFORE_MINUTES = int(os.getenv("ACCIDENTAL_EXIT_ENTRY_BEFORE_MINUTES", "120"))
# Usuarios semilla solicitados para beta. Se guardan con hash, no en texto plano en la base.
DEFAULT_SYSTEM_USERS = [
    {"username": "Adjm", "password": "Adjm4rdur", "role": "Supremo", "display_name": "Usuario Supremo"},
    {"username": "Admin4rd", "password": "Adm4rd", "role": "Supremo", "display_name": "Administrador Supremo"},
    {"username": "Altima", "password": "Altima", "role": "Vigilancia", "display_name": "Vigilancia Altima"},
    {"username": "Adhm4", "password": "4dhm", "role": "RH", "display_name": "Jefa de RH"},
]

DEFAULT_SYSTEM_PASSWORDS = {item["username"]: item["password"] for item in DEFAULT_SYSTEM_USERS}

# Vigilantes semilla para cambio de turno por QR.
DEFAULT_GUARDS = [
    {"code": "VIG-ALTIMA-1", "alias": "Altima 1", "nombre": "Reymundo Méndez", "active": 1, "qr_activo": 1},
    {"code": "VIG-ALTIMA-2", "alias": "Altima 2", "nombre": "David Martinez", "active": 1, "qr_activo": 1},
]

DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
CORRECTIONS_DIR.mkdir(parents=True, exist_ok=True)
EMPLOYEE_PHOTOS_DIR.mkdir(parents=True, exist_ok=True)
VEHICLE_PHOTOS_DIR.mkdir(parents=True, exist_ok=True)

raw_database_url = os.getenv("DATABASE_URL", "").strip()
if raw_database_url:
    # Render puede entregar postgresql://; SQLAlchemy también lo acepta en versiones nuevas,
    # pero normalizamos a psycopg para evitar dramas de dependencias, que ya tenemos suficientes.
    if raw_database_url.startswith("postgres://"):
        DATABASE_URL = raw_database_url.replace("postgres://", "postgresql+psycopg://", 1)
    elif raw_database_url.startswith("postgresql://"):
        DATABASE_URL = raw_database_url.replace("postgresql://", "postgresql+psycopg://", 1)
    else:
        DATABASE_URL = raw_database_url
else:
    DATABASE_URL = f"sqlite:///{DB_PATH}"

connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, future=True, pool_pre_ping=True, connect_args=connect_args)

app = FastAPI(title=APP_NAME)
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), geolocation=(), microphone=()")
    # CSP prudente para esta app. Permitimos inline por las plantillas actuales, pero cerramos fuentes externas.
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; img-src 'self' data: blob:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline' https://unpkg.com; connect-src 'self'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'"
    )
    return response


# -----------------------------
# Helpers
# -----------------------------

def now_mx() -> datetime:
    return datetime.now(TZ)


def clean_id(value: str) -> str:
    value = (value or "").strip().upper()
    value = re.sub(r"[^A-Z0-9_\-]", "", value)
    return value[:40]


def clean_employee_id(value) -> str:
    """Normaliza IDs de empleado. Regla vigente: ID únicamente numérico."""
    raw = as_text(value) if "as_text" in globals() else str(value or "").strip()
    raw = raw.strip()
    # Excel suele convertir 126 en 126.0 cuando juega a ser útil. Gracias, Excel.
    if re.fullmatch(r"\d+\.0+", raw):
        raw = raw.split(".", 1)[0]
    raw = re.sub(r"\s+", "", raw)
    return raw if re.fullmatch(r"\d+", raw) else ""


def employee_id_number(value: str) -> int:
    """Extrae el número usado para calcular el siguiente ID.
    Usa IDs numéricos y también el tramo final de IDs viejos tipo EMP-000125 para migrar sin drama.
    """
    raw = as_text(value) if "as_text" in globals() else str(value or "")
    raw = raw.strip()
    if re.fullmatch(r"\d+", raw):
        return int(raw)
    match = re.search(r"(\d+)$", raw)
    return int(match.group(1)) if match else 0


def normalize_catalog_value(value: str) -> str:
    """Normaliza textos de catálogo para reglas administrativas simples."""
    return re.sub(r"\s+", " ", as_text(value).strip().upper())


def is_gerencia_general_record(row: dict) -> bool:
    """Registros de Gerencia General no cuentan para calcular el consecutivo operativo."""
    area = normalize_catalog_value(row.get("area", ""))
    puesto = normalize_catalog_value(row.get("puesto", ""))
    return area == "GG" or puesto == "GERENCIA GENERAL"


def next_numeric_employee_id(conn: Connection, used_extra: Optional[set[str]] = None) -> str:
    """Devuelve el siguiente ID numérico operativo.

    Regla vigente:
    - El ID debe ser únicamente numérico.
    - Para calcular el consecutivo se toma el número más alto de empleados que NO sean área GG
      y que NO tengan puesto GERENCIA GENERAL.
    - Luego se suma 1.
    - Si ese número ya existe por algún registro excluido, se avanza al siguiente libre para no duplicar.
    """
    used_extra = {as_text(x) for x in (used_extra or set())}
    rows = fetch_all(conn, "SELECT id, area, puesto FROM employees")

    # Todos los IDs existentes se reservan para evitar duplicados, incluso los de GG/Gerencia General.
    all_used = {as_text(r.get("id")) for r in rows}
    all_used.update(used_extra)

    max_operativo = 0
    for row in rows:
        if is_gerencia_general_record(row):
            continue
        max_operativo = max(max_operativo, employee_id_number(row.get("id")))

    # IDs asignados durante una misma importación también deben empujar el consecutivo.
    for employee_id in used_extra:
        max_operativo = max(max_operativo, employee_id_number(employee_id))

    candidate = max_operativo + 1
    while str(candidate) in all_used:
        candidate += 1
    return str(candidate)


def clean_guard_code(value: str) -> str:
    """Normaliza códigos de vigilante. Acepta QR como VIG:VIG-ALTIMA-1 o VIG-ALTIMA-1."""
    value = (value or "").strip().upper()
    if value.startswith("VIG:"):
        value = value.split(":", 1)[1]
    if value.startswith("GUARD:"):
        value = value.split(":", 1)[1]
    value = re.sub(r"[^A-Z0-9_\-]", "", value)
    return value[:60]


def clean_username(value: str) -> str:
    # Usuario seguro pero respetando mayúsculas/minúsculas solicitadas.
    value = (value or "").strip()
    value = re.sub(r"[^A-Za-z0-9_.@\-]", "", value)
    return value[:60]


def safe_filename(value: str) -> str:
    value = clean_id(value) or "ARCHIVO"
    return value


def filename_slug(value: str, fallback: str = "archivo") -> str:
    """Nombre seguro para archivos descargables, conservando lectura humana."""
    value = as_text(value) if "as_text" in globals() else str(value or "").strip()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return (value or fallback)[:90]


def employee_file_base(emp: dict) -> str:
    return f"{filename_slug(emp.get('nombre') or 'Empleado')}_{filename_slug(emp.get('id') or 'SIN_ID')}"


def bool_from_excel(value) -> int:
    if value is None:
        return 0
    text_value = str(value).strip().lower()
    return 1 if text_value in {"si", "sí", "s", "yes", "y", "true", "1", "x"} else 0


def parse_excel_date_value(value) -> Optional[date]:
    """Convierte una fecha de Excel/texto a date. Soporta fechas de Excel, datetime y texto común."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(raw).date()
    except Exception:
        return None


def parse_excel_time_value(value) -> Optional[time]:
    """Convierte hora desde Excel/texto a time."""
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).time().replace(microsecond=0)
        except Exception:
            # Excel puede guardar horas como fracción de día. Gracias por tanto, Excel.
            total_seconds = int(round(float(value) * 24 * 60 * 60))
            total_seconds %= 24 * 60 * 60
            return time(total_seconds // 3600, (total_seconds % 3600) // 60)
    raw = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(raw, fmt).time().replace(microsecond=0)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(raw).time().replace(microsecond=0)
    except Exception:
        return None


def parse_excel_datetime_value(value, shift_day: date, config: dict, is_exit: bool = False) -> Optional[datetime]:
    """Convierte celda de entrada/salida a datetime con zona MX.
    Si viene solo hora, se combina con fecha_turno y se cruza al día siguiente cuando aplica.
    """
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=TZ) if value.tzinfo is None else value.astimezone(TZ)
    # Si es texto con fecha y hora completa
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y %H:%M:%S"):
            try:
                return datetime.strptime(raw, fmt).replace(tzinfo=TZ)
            except Exception:
                pass
        try:
            parsed = datetime.fromisoformat(raw)
            return parsed.replace(tzinfo=TZ) if parsed.tzinfo is None else parsed.astimezone(TZ)
        except Exception:
            pass
    cell_time = parse_excel_time_value(value)
    if not cell_time:
        return None
    entry_t = parse_time_value(config.get("entry_time"), time(8, 0))
    day = shift_day
    if is_exit and cell_time < entry_t:
        day = shift_day + timedelta(days=1)
    return datetime.combine(day, cell_time, tzinfo=TZ)


def as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def fetch_one(conn: Connection, sql: str, params: Optional[dict] = None):
    row = conn.execute(text(sql), params or {}).mappings().first()
    return dict(row) if row else None


def fetch_all(conn: Connection, sql: str, params: Optional[dict] = None):
    rows = conn.execute(text(sql), params or {}).mappings().all()
    return [dict(r) for r in rows]


def get_table_columns(conn: Connection, table_name: str) -> set[str]:
    if engine.dialect.name == "sqlite":
        rows = conn.exec_driver_sql(f"PRAGMA table_info({table_name})").fetchall()
        return {row[1] for row in rows}
    rows = conn.execute(
        text(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = :table_name
            """
        ),
        {"table_name": table_name},
    ).fetchall()
    return {row[0] for row in rows}


def init_db() -> None:
    is_sqlite = engine.dialect.name == "sqlite"
    attendance_id = "INTEGER PRIMARY KEY AUTOINCREMENT" if is_sqlite else "SERIAL PRIMARY KEY"
    audit_id = "INTEGER PRIMARY KEY AUTOINCREMENT" if is_sqlite else "SERIAL PRIMARY KEY"

    statements = [
        """
        CREATE TABLE IF NOT EXISTS employees (
            id TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            area TEXT DEFAULT '',
            puesto TEXT DEFAULT '',
            turno TEXT DEFAULT 'Día',
            estado TEXT DEFAULT 'Activo',
            tiene_vehiculo INTEGER DEFAULT 0,
            requiere_fotos_vehiculo INTEGER DEFAULT 0,
            foto_path TEXT DEFAULT '',
            qr_activo INTEGER DEFAULT 1,
            observaciones TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS attendance (
            id {attendance_id},
            employee_id TEXT NOT NULL,
            shift_date TEXT NOT NULL,
            turno TEXT NOT NULL,
            entry_at TEXT,
            exit_at TEXT,
            entry_guard TEXT DEFAULT '',
            exit_guard TEXT DEFAULT '',
            entry_status TEXT DEFAULT '',
            exit_status TEXT DEFAULT '',
            late_reason TEXT DEFAULT '',
            early_reason TEXT DEFAULT '',
            vehicle_expected INTEGER DEFAULT 0,
            vehicle_entered INTEGER DEFAULT 0,
            vehicle_front_entry TEXT DEFAULT '',
            vehicle_trunk_entry TEXT DEFAULT '',
            vehicle_front_exit TEXT DEFAULT '',
            vehicle_trunk_exit TEXT DEFAULT '',
            incident TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(employee_id) REFERENCES employees(id)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS shift_settings (
            name TEXT PRIMARY KEY,
            entry_time TEXT NOT NULL,
            exit_time TEXT NOT NULL,
            crosses_midnight INTEGER DEFAULT 0,
            work_days TEXT DEFAULT '1,2,3,4,5',
            entry_tolerance_minutes INTEGER DEFAULT 10,
            exit_tolerance_minutes INTEGER DEFAULT 10,
            extra_after_minutes INTEGER DEFAULT 30,
            active INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'Admin',
            active INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS guards (
            code TEXT PRIMARY KEY,
            alias TEXT DEFAULT '',
            nombre TEXT DEFAULT '',
            active INTEGER DEFAULT 1,
            qr_activo INTEGER DEFAULT 1,
            observaciones TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS active_guard_state (
            id TEXT PRIMARY KEY,
            guard_code TEXT NOT NULL,
            guard_display TEXT NOT NULL,
            started_at TEXT NOT NULL,
            changed_by TEXT DEFAULT '',
            updated_at TEXT NOT NULL
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS user_sessions (
            token_hash TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            expires_at TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(username) REFERENCES users(username)
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS correction_batches (
            batch_id TEXT PRIMARY KEY,
            user_name TEXT NOT NULL,
            status TEXT DEFAULT 'preview',
            file_path TEXT DEFAULT '',
            total_rows INTEGER DEFAULT 0,
            total_changes INTEGER DEFAULT 0,
            total_errors INTEGER DEFAULT 0,
            created_at TEXT NOT NULL,
            applied_at TEXT DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS system_events (
            id {audit_id},
            level TEXT DEFAULT 'info',
            module TEXT DEFAULT '',
            event_type TEXT DEFAULT '',
            user_name TEXT DEFAULT '',
            message TEXT DEFAULT '',
            detail TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS audit_log (
            id {audit_id},
            table_name TEXT NOT NULL,
            record_id TEXT NOT NULL,
            action TEXT NOT NULL,
            user_name TEXT DEFAULT '',
            field_name TEXT DEFAULT '',
            old_value TEXT DEFAULT '',
            new_value TEXT DEFAULT '',
            reason TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_attendance_employee_open ON attendance(employee_id, exit_at)",
        "CREATE INDEX IF NOT EXISTS idx_attendance_shift_date ON attendance(shift_date)",
        "CREATE INDEX IF NOT EXISTS idx_audit_record ON audit_log(table_name, record_id)",
        "CREATE INDEX IF NOT EXISTS idx_sessions_username ON user_sessions(username)",
        "CREATE INDEX IF NOT EXISTS idx_sessions_expires ON user_sessions(expires_at)",
        "CREATE INDEX IF NOT EXISTS idx_system_events_created ON system_events(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_system_events_level ON system_events(level)",
        "CREATE INDEX IF NOT EXISTS idx_guards_active ON guards(active, qr_activo)",
    ]

    with engine.begin() as conn:
        for statement in statements:
            conn.exec_driver_sql(statement)

        # Migraciones suaves para despliegues existentes en Render. No tiramos la base,
        # porque aparentemente destruir datos para arreglar una columna no es administración.
        existing_columns = get_table_columns(conn, "attendance")
        add_columns = {
            "scheduled_entry_at": "TEXT DEFAULT ''",
            "scheduled_exit_at": "TEXT DEFAULT ''",
            "entry_limit_at": "TEXT DEFAULT ''",
            "exit_early_limit_at": "TEXT DEFAULT ''",
            "extra_limit_at": "TEXT DEFAULT ''",
            "entry_tolerance_minutes": "INTEGER DEFAULT 10",
            "exit_tolerance_minutes": "INTEGER DEFAULT 10",
            "extra_after_minutes": "INTEGER DEFAULT 30",
            "late_minutes": "INTEGER DEFAULT 0",
            "early_minutes": "INTEGER DEFAULT 0",
            "extra_minutes": "INTEGER DEFAULT 0",
            "late_justified": "INTEGER DEFAULT 0",
            "early_justified": "INTEGER DEFAULT 0",
            "extra_authorized": "INTEGER DEFAULT 0",
            "provisional_exit": "INTEGER DEFAULT 0",
            "review_required": "INTEGER DEFAULT 0",
            "review_status": "TEXT DEFAULT ''",
            "auto_closed_at": "TEXT DEFAULT ''",
            "anulled": "INTEGER DEFAULT 0",
            "lunch_taken": "TEXT DEFAULT ''",
            "worked_minutes": "INTEGER DEFAULT 0",
        }
        for col, definition in add_columns.items():
            if col not in existing_columns:
                conn.exec_driver_sql(f"ALTER TABLE attendance ADD COLUMN {col} {definition}")

        shift_columns = get_table_columns(conn, "shift_settings")
        shift_add_columns = {
            "auto_close_enabled": "INTEGER DEFAULT 1",
            "provisional_close_time": "TEXT DEFAULT '02:00'",
            "max_open_minutes": "INTEGER DEFAULT 1080",
        }
        for col, definition in shift_add_columns.items():
            if col not in shift_columns:
                conn.exec_driver_sql(f"ALTER TABLE shift_settings ADD COLUMN {col} {definition}")

        user_columns = get_table_columns(conn, "users")
        user_add_columns = {
            "display_name": "TEXT DEFAULT ''",
            "failed_login_attempts": "INTEGER DEFAULT 0",
            "locked_until": "TEXT DEFAULT ''",
            "last_login_at": "TEXT DEFAULT ''",
            "password_changed_at": "TEXT DEFAULT ''",
        }
        for col, definition in user_add_columns.items():
            if col not in user_columns:
                conn.exec_driver_sql(f"ALTER TABLE users ADD COLUMN {col} {definition}")

        ts = now_mx().isoformat()
        defaults = [
            {"name": "Día", "entry_time": "08:00", "exit_time": "19:00", "crosses_midnight": 0},
            {"name": "Noche", "entry_time": "19:00", "exit_time": "08:00", "crosses_midnight": 1},
        ]
        for item in defaults:
            exists = fetch_one(conn, "SELECT name FROM shift_settings WHERE name = :name", {"name": item["name"]})
            if not exists:
                conn.execute(
                    text(
                        """
                        INSERT INTO shift_settings (
                            name, entry_time, exit_time, crosses_midnight, work_days,
                            entry_tolerance_minutes, exit_tolerance_minutes, extra_after_minutes,
                            active, auto_close_enabled, provisional_close_time, max_open_minutes, created_at, updated_at
                        ) VALUES (
                            :name, :entry_time, :exit_time, :crosses_midnight, '1,2,3,4,5',
                            10, 10, 30, 1, 1, :provisional_close_time, :max_open_minutes, :created_at, :updated_at
                        )
                        """
                    ),
                    {**item, "provisional_close_time": "02:00" if item["name"] == "Día" else "14:00", "max_open_minutes": 1080, "created_at": ts, "updated_at": ts},
                )


def audit(conn: Connection, table: str, record_id: str, action: str,
          user: str = "Sistema", field: str = "", old: str = "", new: str = "", reason: str = "") -> None:
    conn.execute(
        text(
            """
            INSERT INTO audit_log (table_name, record_id, action, user_name, field_name, old_value, new_value, reason, created_at)
            VALUES (:table_name, :record_id, :action, :user_name, :field_name, :old_value, :new_value, :reason, :created_at)
            """
        ),
        {
            "table_name": table,
            "record_id": record_id,
            "action": action,
            "user_name": user,
            "field_name": field,
            "old_value": old,
            "new_value": new,
            "reason": reason,
            "created_at": now_mx().isoformat(),
        }
    )


def log_event(conn: Connection, level: str, module: str, event_type: str, message: str,
              user: str = "Sistema", detail: str = "") -> None:
    try:
        conn.execute(
            text(
                """
                INSERT INTO system_events (level, module, event_type, user_name, message, detail, created_at)
                VALUES (:level, :module, :event_type, :user_name, :message, :detail, :created_at)
                """
            ),
            {
                "level": level,
                "module": module,
                "event_type": event_type,
                "user_name": user,
                "message": message,
                "detail": detail,
                "created_at": now_mx().isoformat(),
            },
        )
    except Exception:
        # La bitácora técnica no debe tumbar el registro principal. Qué detalle tan civilizado.
        pass


def get_employee(conn: Connection, employee_id: str):
    return fetch_one(conn, "SELECT * FROM employees WHERE id = :id", {"id": employee_id})


def get_guard(conn: Connection, guard_code: str):
    return fetch_one(conn, "SELECT * FROM guards WHERE code = :code", {"code": clean_guard_code(guard_code)})


def find_guard(conn: Connection, value: str):
    """Busca vigilante por código QR, alias o alias+nombre para hacer la captura menos delicada."""
    cleaned = clean_guard_code(value)
    if not cleaned:
        return None
    exact = get_guard(conn, cleaned)
    if exact:
        return exact
    rows = fetch_all(conn, "SELECT * FROM guards")
    for row in rows:
        options = [row.get("alias") or "", row.get("nombre") or "", f"{row.get('alias') or ''}-{row.get('nombre') or ''}"]
        if any(clean_guard_code(opt) == cleaned for opt in options):
            return row
    return None


def guard_display(guard: Optional[dict]) -> str:
    if not guard:
        return ""
    alias = as_text(guard.get("alias")) or as_text(guard.get("code"))
    nombre = as_text(guard.get("nombre"))
    if nombre and nombre.lower() != "pendiente":
        return f"{alias} - {nombre}"
    return alias


def get_active_guard(conn: Connection):
    row = fetch_one(conn, "SELECT * FROM active_guard_state WHERE id = 'default'")
    if not row:
        return None
    guard = get_guard(conn, row.get("guard_code", ""))
    if not guard or not int(guard.get("active") or 0) or not int(guard.get("qr_activo") or 0):
        return None
    row["guard"] = guard
    row["display"] = row.get("guard_display") or guard_display(guard)
    return row


def set_active_guard(conn: Connection, guard: dict, changed_by: str) -> dict:
    ts = now_mx().isoformat()
    display = guard_display(guard)
    previous = get_active_guard(conn)
    conn.execute(text("DELETE FROM active_guard_state WHERE id = 'default'"))
    conn.execute(
        text("""
            INSERT INTO active_guard_state (id, guard_code, guard_display, started_at, changed_by, updated_at)
            VALUES ('default', :guard_code, :guard_display, :started_at, :changed_by, :updated_at)
        """),
        {"guard_code": guard["code"], "guard_display": display, "started_at": ts, "changed_by": changed_by, "updated_at": ts},
    )
    audit(conn, "active_guard_state", "default", "CHANGE", changed_by, "guard_code", previous.get("guard_code", "") if previous else "", guard["code"], "Cambio de vigilante activo por QR")
    log_event(conn, "info", "vigilancia", "guard_change", f"Vigilante activo: {display}", changed_by)
    return {"guard_code": guard["code"], "display": display, "started_at": ts}


def parse_time_value(value: str, fallback: time) -> time:
    try:
        parts = str(value or "").strip().split(":")
        return time(int(parts[0]), int(parts[1]))
    except Exception:
        return fallback


def default_shift_config(turno: str) -> dict:
    if (turno or "").lower().startswith("n"):
        return {
            "name": "Noche",
            "entry_time": "19:00",
            "exit_time": "08:00",
            "crosses_midnight": 1,
            "entry_tolerance_minutes": 10,
            "exit_tolerance_minutes": 10,
            "extra_after_minutes": 30,
        }
    return {
        "name": "Día",
        "entry_time": "08:00",
        "exit_time": "19:00",
        "crosses_midnight": 0,
        "entry_tolerance_minutes": 10,
        "exit_tolerance_minutes": 10,
        "extra_after_minutes": 30,
    }


def get_shift_config(conn: Connection, turno: str) -> dict:
    row = fetch_one(conn, "SELECT * FROM shift_settings WHERE name = :name", {"name": turno})
    if not row and turno == "Dia":
        row = fetch_one(conn, "SELECT * FROM shift_settings WHERE name = 'Día'")
    if not row:
        return default_shift_config(turno)
    return row


def current_shift_date(turno: str, dt: Optional[datetime] = None, config: Optional[dict] = None) -> str:
    dt = dt or now_mx()
    config = config or default_shift_config(turno)
    exit_t = parse_time_value(config.get("exit_time"), time(8, 0))
    crosses = bool(config.get("crosses_midnight"))
    # Si cruza medianoche y todavía estamos antes/igual de la hora de salida, pertenece al día anterior.
    if crosses and dt.time() <= exit_t:
        return (dt.date() - timedelta(days=1)).isoformat()
    return dt.date().isoformat()


def schedule_times_from_config(config: dict, shift_date_iso: str):
    shift_day = date.fromisoformat(shift_date_iso)
    entry_t = parse_time_value(config.get("entry_time"), time(8, 0))
    exit_t = parse_time_value(config.get("exit_time"), time(19, 0))
    start_dt = datetime.combine(shift_day, entry_t, tzinfo=TZ)
    exit_day = shift_day + timedelta(days=1) if bool(config.get("crosses_midnight")) else shift_day
    end_dt = datetime.combine(exit_day, exit_t, tzinfo=TZ)
    return start_dt, end_dt


def expected_times(turno: str, shift_date_iso: str):
    return schedule_times_from_config(default_shift_config(turno), shift_date_iso)


def minutes_between_late(a: datetime, b: datetime) -> int:
    return max(0, int((a - b).total_seconds() // 60))


def evaluate_entry(config: dict, shift_date_iso: str, dt: datetime) -> dict:
    start_dt, end_dt = schedule_times_from_config(config, shift_date_iso)
    tol = int(config.get("entry_tolerance_minutes") or 0)
    entry_limit = start_dt + timedelta(minutes=tol)

    # Regla operativa MS:
    # La tolerancia solo decide si existe retardo.
    # Si existe retardo, los minutos se cuentan desde la hora programada,
    # no desde el límite de tolerancia.
    # Ejemplo: entrada 08:00, tolerancia 5 min, llegada 08:08 => 8 min tarde.
    is_late = dt > entry_limit
    late_minutes = minutes_between_late(dt, start_dt) if is_late else 0
    status = "Retardo" if is_late else "Correcta"
    return {
        "status": status,
        "scheduled_entry_at": start_dt.isoformat(),
        "scheduled_exit_at": end_dt.isoformat(),
        "entry_limit_at": entry_limit.isoformat(),
        "exit_early_limit_at": (end_dt - timedelta(minutes=int(config.get("exit_tolerance_minutes") or 0))).isoformat(),
        "extra_limit_at": (end_dt + timedelta(minutes=int(config.get("extra_after_minutes") or 30))).isoformat(),
        "entry_tolerance_minutes": tol,
        "exit_tolerance_minutes": int(config.get("exit_tolerance_minutes") or 0),
        "extra_after_minutes": int(config.get("extra_after_minutes") or 30),
        "late_minutes": late_minutes,
    }


def evaluate_exit(config: dict, shift_date_iso: str, dt: datetime) -> dict:
    start_dt, end_dt = schedule_times_from_config(config, shift_date_iso)
    exit_tol = int(config.get("exit_tolerance_minutes") or 0)
    extra_after = int(config.get("extra_after_minutes") or 30)
    early_limit = end_dt - timedelta(minutes=exit_tol)
    extra_limit = end_dt + timedelta(minutes=extra_after)

    # La tolerancia de salida temprana solo decide si existe incidencia.
    # Si existe, los minutos se cuentan contra la hora programada de salida.
    # Ejemplo: salida 19:00, tolerancia 10 min, salida real 18:45 => 15 min temprano.
    is_early = dt < early_limit
    is_extra = dt > extra_limit
    early_minutes = max(0, int((end_dt - dt).total_seconds() // 60)) if is_early else 0
    extra_minutes = max(0, int((dt - end_dt).total_seconds() // 60)) if is_extra else 0
    if is_early:
        status = "Salida temprana"
    elif is_extra:
        status = "Extra"
    else:
        status = "Correcta"
    return {
        "status": status,
        "scheduled_entry_at": start_dt.isoformat(),
        "scheduled_exit_at": end_dt.isoformat(),
        "entry_limit_at": (start_dt + timedelta(minutes=int(config.get("entry_tolerance_minutes") or 0))).isoformat(),
        "exit_early_limit_at": early_limit.isoformat(),
        "extra_limit_at": extra_limit.isoformat(),
        "entry_tolerance_minutes": int(config.get("entry_tolerance_minutes") or 0),
        "exit_tolerance_minutes": exit_tol,
        "extra_after_minutes": extra_after,
        "early_minutes": early_minutes,
        "extra_minutes": extra_minutes,
    }




def is_open_attendance_row(row: Optional[dict]) -> bool:
    if not row:
        return False
    return as_text(row.get("exit_at")).strip() == ""


def open_attendance_sql(alias: str = "") -> str:
    field = f"{alias}.exit_at" if alias else "exit_at"
    return f"({field} IS NULL OR TRIM(COALESCE({field}, '')) = '')"


def find_open_attendance(conn: Connection, employee_id: str) -> Optional[dict]:
    return fetch_one(
        conn,
        f"""
        SELECT * FROM attendance
        WHERE employee_id = :employee_id
          AND COALESCE(anulled,0)=0
          AND {open_attendance_sql()}
        ORDER BY id DESC
        LIMIT 1
        """,
        {"employee_id": employee_id},
    )


def find_latest_shift_attendance(conn: Connection, employee_id: str, shift_date: str, turno: str = "") -> Optional[dict]:
    params = {"employee_id": employee_id, "shift_date": shift_date}
    clause = ""
    if turno:
        clause = " AND turno = :turno"
        params["turno"] = turno
    return fetch_one(
        conn,
        f"""
        SELECT * FROM attendance
        WHERE employee_id = :employee_id
          AND shift_date = :shift_date
          {clause}
          AND COALESCE(anulled,0)=0
        ORDER BY id DESC
        LIMIT 1
        """,
        params,
    )


def is_exit_window(config: dict, shift_date_iso: str, dt: datetime) -> bool:
    _start_dt, end_dt = schedule_times_from_config(config, shift_date_iso)
    return (end_dt - timedelta(minutes=EXIT_WINDOW_BEFORE_MINUTES)) <= dt <= (end_dt + timedelta(minutes=EXIT_WINDOW_AFTER_MINUTES))


def build_entry_confirmation(config: dict, shift_date: str, dt: datetime, latest: Optional[dict] = None) -> Optional[dict]:
    """Devuelve una advertencia cuando la entrada requiere doble confirmación.

    No bloquea turno noche: la ventana se calcula contra el turno del empleado.
    Si alguien realmente entra a las 19:00 porque es turno noche, su turno no estará en ventana de salida.
    """
    _start_dt, end_dt = schedule_times_from_config(config, shift_date)
    reasons = []
    if latest and not is_open_attendance_row(latest):
        reasons.append("este empleado ya tiene un registro cerrado para este día")
    if is_exit_window(config, shift_date, dt):
        reasons.append(f"estás cerca de la salida programada del turno ({end_dt.strftime('%H:%M')})")
    if not reasons:
        return None
    return {
        "requires_confirmation": True,
        "confirmation_code": "CONFIRMAR",
        "reason": "; ".join(reasons),
        "message": "Confirmación requerida: " + "; ".join(reasons) + ". Si realmente es una nueva entrada/reentrada, confirma antes de guardar.",
    }


def entry_confirmation_required_response(config: dict, shift_date: str, dt: datetime, latest: Optional[dict] = None) -> JSONResponse:
    info = build_entry_confirmation(config, shift_date, dt, latest) or {}
    return JSONResponse(
        status_code=409,
        content={
            "ok": False,
            "requires_confirmation": True,
            "message": info.get("message") or "Confirmación requerida para registrar esta entrada.",
            "shift_date": shift_date,
        },
    )



def close_shift_day_safely(conn: Connection, shift_date: str, turno: str, exit_time_value: str, user_name: str) -> dict:
    """Cierra salidas de un turno/día sin inventar entradas.

    1. Anula entradas accidentales creadas en ventana de salida cuando ya existe un registro
       anterior del mismo empleado para el mismo día.
    2. Cierra registros abiertos usando la hora de salida indicada.
    3. Deja sin tocar registros que no puede interpretar con seguridad.
    """
    shift_date = as_text(shift_date)
    turno = as_text(turno) or "Día"
    if not shift_date:
        shift_date = now_mx().date().isoformat()
    config = get_shift_config(conn, turno)
    start_dt, scheduled_end_dt = schedule_times_from_config(config, shift_date)
    exit_t = parse_time_value(exit_time_value, scheduled_end_dt.time())
    exit_day = date.fromisoformat(shift_date) + timedelta(days=1) if bool(config.get("crosses_midnight")) else date.fromisoformat(shift_date)
    close_dt = datetime.combine(exit_day, exit_t, tzinfo=TZ)
    ts = now_mx().isoformat()
    result = {
        "fecha": shift_date,
        "turno": turno,
        "hora_salida": close_dt.strftime("%H:%M"),
        "revisados": 0,
        "anulados_accidentales": 0,
        "cerrados": 0,
        "ya_cerrados": 0,
        "sin_entrada": 0,
        "omitidos": 0,
        "mensajes": [],
    }

    records = fetch_all(
        conn,
        """
        SELECT * FROM attendance
        WHERE shift_date = :shift_date
          AND turno = :turno
          AND COALESCE(anulled,0)=0
        ORDER BY employee_id, id
        """,
        {"shift_date": shift_date, "turno": turno},
    )
    result["revisados"] = len(records)

    by_emp = {}
    for r in records:
        by_emp.setdefault(r["employee_id"], []).append(r)

    accidental_ids = set()
    accidental_from = close_dt - timedelta(minutes=ACCIDENTAL_EXIT_ENTRY_BEFORE_MINUTES)
    prior_cutoff = close_dt - timedelta(minutes=60)

    for emp_id, emp_records in by_emp.items():
        prior_records = []
        for r in emp_records:
            entry_dt = parse_dt(r.get("entry_at") or "")
            if entry_dt and entry_dt < prior_cutoff:
                prior_records.append(r)

        if not prior_records:
            continue

        for r in emp_records:
            entry_dt = parse_dt(r.get("entry_at") or "")
            if not entry_dt:
                continue
            if is_open_attendance_row(r) and entry_dt >= accidental_from:
                accidental_ids.add(r["id"])

    for rid in accidental_ids:
        old = fetch_one(conn, "SELECT * FROM attendance WHERE id=:id", {"id": rid})
        if not old:
            continue
        incident = as_text(old.get("incident"))
        note = "ANULADO: entrada accidental capturada en ventana de salida; ya existía registro previo del mismo día."
        new_incident = (incident + " | " + note).strip(" |") if incident else note
        conn.execute(
            text("""
            UPDATE attendance
            SET anulled=1,
                review_required=1,
                review_status='Anulado por cierre seguro',
                incident=:incident,
                updated_at=:updated_at
            WHERE id=:id
            """),
            {"id": rid, "incident": new_incident, "updated_at": ts},
        )
        audit(conn, "attendance", str(rid), "ANUL_ACCIDENTAL_EXIT_ENTRY", user_name, reason=note)
        result["anulados_accidentales"] += 1

    open_rows = fetch_all(
        conn,
        f"""
        SELECT * FROM attendance
        WHERE shift_date = :shift_date
          AND turno = :turno
          AND COALESCE(anulled,0)=0
          AND {open_attendance_sql()}
        ORDER BY employee_id, id
        """,
        {"shift_date": shift_date, "turno": turno},
    )

    for row in open_rows:
        entry_dt = parse_dt(row.get("entry_at") or "")
        if not entry_dt:
            result["sin_entrada"] += 1
            continue

        eval_data = evaluate_exit(config, shift_date, close_dt)
        status = eval_data["status"]
        lunch_value = "SI"
        worked_minutes, lunch_value, payroll_note = compute_worked_minutes(
            row.get("entry_at"), close_dt.isoformat(), eval_data.get("scheduled_entry_at"), eval_data.get("scheduled_exit_at"), lunch_value, status
        )
        incident = as_text(row.get("incident"))
        note = "Salida cerrada por cierre seguro administrativo."
        if payroll_note:
            note += f" {payroll_note}"
        new_incident = (incident + " | " + note).strip(" |") if incident else note

        conn.execute(
            text("""
            UPDATE attendance
            SET exit_at=:exit_at,
                exit_guard=:exit_guard,
                exit_status=:exit_status,
                early_reason=:early_reason,
                scheduled_entry_at=COALESCE(NULLIF(scheduled_entry_at, ''), :scheduled_entry_at),
                scheduled_exit_at=COALESCE(NULLIF(scheduled_exit_at, ''), :scheduled_exit_at),
                entry_limit_at=COALESCE(NULLIF(entry_limit_at, ''), :entry_limit_at),
                exit_early_limit_at=:exit_early_limit_at,
                extra_limit_at=:extra_limit_at,
                entry_tolerance_minutes=:entry_tolerance_minutes,
                exit_tolerance_minutes=:exit_tolerance_minutes,
                extra_after_minutes=:extra_after_minutes,
                early_minutes=:early_minutes,
                extra_minutes=:extra_minutes,
                lunch_taken=:lunch_taken,
                worked_minutes=:worked_minutes,
                incident=:incident,
                updated_at=:updated_at
            WHERE id=:id
            """),
            {
                "id": row["id"],
                "exit_at": close_dt.isoformat(),
                "exit_guard": f"CIERRE SEGURO / {user_name}",
                "exit_status": status,
                "early_reason": "",
                "lunch_taken": lunch_value,
                "worked_minutes": worked_minutes,
                "incident": new_incident,
                "updated_at": ts,
                **eval_data,
            },
        )
        audit(conn, "attendance", str(row["id"]), "SAFE_SHIFT_CLOSE", user_name, reason=f"Cierre seguro {shift_date} {turno} {close_dt.strftime('%H:%M')}")
        result["cerrados"] += 1

    total_closed = fetch_all(
        conn,
        f"""
        SELECT id FROM attendance
        WHERE shift_date = :shift_date
          AND turno = :turno
          AND COALESCE(anulled,0)=0
          AND NOT {open_attendance_sql()}
        """,
        {"shift_date": shift_date, "turno": turno},
    )
    result["ya_cerrados"] = max(0, len(total_closed) - result["cerrados"])
    log_event(conn, "info", "configuracion", "safe_shift_close", f"Cierre seguro {shift_date} {turno}: {result}", user_name)
    return result



def recalculate_attendance_records(conn: Connection, fecha_inicio: str, fecha_fin: str, turno: str = "", user_name: str = "Sistema") -> dict:
    """Recalcula retardo/salida temprana/extra usando la configuración actual de turnos.

    Esto NO debe ocurrir automáticamente al cambiar un turno; se aplica por rango para que
    el Admin decida cuándo una regla nueva debe afectar registros ya capturados/importados.
    """
    params = {"inicio": fecha_inicio, "fin": fecha_fin}
    clause = "a.shift_date BETWEEN :inicio AND :fin AND COALESCE(a.anulled,0)=0"
    if turno:
        clause += " AND a.turno = :turno"
        params["turno"] = turno
    rows = fetch_all(conn, f"SELECT a.* FROM attendance a WHERE {clause} ORDER BY a.shift_date, a.id", params)
    changed_records = 0
    changed_fields = 0
    scanned = len(rows)
    late_added = 0
    late_removed = 0
    early_added = 0
    early_removed = 0
    extra_added = 0
    extra_removed = 0
    errors = []
    ts = now_mx().isoformat()

    def norm(v):
        return "" if v is None else str(v)

    def same(a, b):
        return norm(a) == norm(b)

    for row in rows:
        rid = row.get("id")
        config = get_shift_config(conn, row.get("turno") or "Día")
        updates = {"id": rid, "updated_at": ts}
        field_changes = []
        reason = f"Recalculo masivo por cambio de reglas ({fecha_inicio} a {fecha_fin})"
        entry_dt = parse_dt(row.get("entry_at") or "")
        exit_dt = parse_dt(row.get("exit_at") or "")
        if not entry_dt:
            errors.append(f"Registro {rid}: no tiene entrada válida para recalcular")
            continue

        try:
            entry_eval = evaluate_entry(config, row.get("shift_date"), entry_dt)
        except Exception as exc:
            errors.append(f"Registro {rid}: error evaluando entrada: {exc}")
            continue

        old_entry_status = row.get("entry_status") or ""
        new_entry_status = entry_eval["status"]
        if old_entry_status in ("Tarde", "Retardo") and new_entry_status == "Correcta":
            late_removed += 1
        if old_entry_status not in ("Tarde", "Retardo") and new_entry_status == "Retardo":
            late_added += 1

        # Campos de entrada y snapshots de reglas actuales.
        entry_fields = {
            "scheduled_entry_at": entry_eval["scheduled_entry_at"],
            "scheduled_exit_at": entry_eval["scheduled_exit_at"],
            "entry_limit_at": entry_eval["entry_limit_at"],
            "exit_early_limit_at": entry_eval["exit_early_limit_at"],
            "extra_limit_at": entry_eval["extra_limit_at"],
            "entry_tolerance_minutes": entry_eval["entry_tolerance_minutes"],
            "exit_tolerance_minutes": entry_eval["exit_tolerance_minutes"],
            "extra_after_minutes": entry_eval["extra_after_minutes"],
            "entry_status": new_entry_status,
            "late_minutes": entry_eval["late_minutes"],
        }
        for field, new_value in entry_fields.items():
            if not same(row.get(field), new_value):
                updates[field] = new_value
                field_changes.append((field, row.get(field), new_value))

        if new_entry_status == "Retardo":
            if not (row.get("late_reason") or "").strip():
                default_reason = "NO REGISTRADO - RECALCULO DE TOLERANCIA"
                updates["late_reason"] = default_reason
                updates["late_justified"] = 1
                field_changes.append(("late_reason", row.get("late_reason") or "", default_reason))
                field_changes.append(("late_justified", row.get("late_justified") or 0, 1))
        else:
            # Si ya no es retardo, limpiamos el motivo para no dejar basura de reporte.
            if (row.get("late_reason") or "") or int(row.get("late_justified") or 0) != 0:
                updates["late_reason"] = ""
                updates["late_justified"] = 0
                field_changes.append(("late_reason", row.get("late_reason") or "", ""))
                field_changes.append(("late_justified", row.get("late_justified") or 0, 0))

        if exit_dt:
            try:
                exit_eval = evaluate_exit(config, row.get("shift_date"), exit_dt)
            except Exception as exc:
                errors.append(f"Registro {rid}: error evaluando salida: {exc}")
                continue
            old_exit_status = row.get("exit_status") or ""
            new_exit_status = exit_eval["status"]
            if old_exit_status == "Salida temprana" and new_exit_status != "Salida temprana":
                early_removed += 1
            if old_exit_status != "Salida temprana" and new_exit_status == "Salida temprana":
                early_added += 1
            if old_exit_status == "Extra" and new_exit_status != "Extra":
                extra_removed += 1
            if old_exit_status != "Extra" and new_exit_status == "Extra":
                extra_added += 1

            exit_fields = {
                "scheduled_entry_at": exit_eval["scheduled_entry_at"],
                "scheduled_exit_at": exit_eval["scheduled_exit_at"],
                "entry_limit_at": exit_eval["entry_limit_at"],
                "exit_early_limit_at": exit_eval["exit_early_limit_at"],
                "extra_limit_at": exit_eval["extra_limit_at"],
                "entry_tolerance_minutes": exit_eval["entry_tolerance_minutes"],
                "exit_tolerance_minutes": exit_eval["exit_tolerance_minutes"],
                "extra_after_minutes": exit_eval["extra_after_minutes"],
                "exit_status": new_exit_status,
                "early_minutes": exit_eval["early_minutes"],
                "extra_minutes": exit_eval["extra_minutes"],
            }
            for field, new_value in exit_fields.items():
                if not same(row.get(field), new_value):
                    updates[field] = new_value
                    field_changes.append((field, row.get(field), new_value))

            if new_exit_status == "Salida temprana":
                if not (row.get("early_reason") or "").strip():
                    default_reason = "NO REGISTRADO - RECALCULO DE TOLERANCIA"
                    updates["early_reason"] = default_reason
                    updates["early_justified"] = 1
                    field_changes.append(("early_reason", row.get("early_reason") or "", default_reason))
                    field_changes.append(("early_justified", row.get("early_justified") or 0, 1))
            else:
                if (row.get("early_reason") or "") or int(row.get("early_justified") or 0) != 0:
                    updates["early_reason"] = ""
                    updates["early_justified"] = 0
                    field_changes.append(("early_reason", row.get("early_reason") or "", ""))
                    field_changes.append(("early_justified", row.get("early_justified") or 0, 0))
        else:
            # Registro abierto: no inventamos salida; solo actualizamos reglas de entrada.
            if not same(row.get("exit_status") or "", ""):
                updates["exit_status"] = ""
                field_changes.append(("exit_status", row.get("exit_status") or "", ""))

        if not field_changes:
            continue

        set_clause = ", ".join([f"{k}=:{k}" for k in updates.keys() if k != "id"])
        conn.execute(text(f"UPDATE attendance SET {set_clause} WHERE id=:id"), updates)
        changed_records += 1
        changed_fields += len(field_changes)
        for field, old, new in field_changes:
            audit(conn, "attendance", str(rid), "RECALCULATE_RULES", user_name, field, norm(old), norm(new), reason)

    log_event(
        conn,
        "info",
        "asistencia",
        "recalculate_rules",
        f"Recalculo aplicado: {changed_records} registros modificados",
        user_name,
        json.dumps({"inicio": fecha_inicio, "fin": fecha_fin, "turno": turno, "scanned": scanned, "changed_fields": changed_fields}, ensure_ascii=False),
    )
    return {
        "scanned": scanned,
        "changed_records": changed_records,
        "changed_fields": changed_fields,
        "late_added": late_added,
        "late_removed": late_removed,
        "early_added": early_added,
        "early_removed": early_removed,
        "extra_added": extra_added,
        "extra_removed": extra_removed,
        "errors": errors[:100],
        "total_errors": len(errors),
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "turno": turno,
    }


def provisional_close_deadline(config: dict, shift_date_iso: str) -> datetime:
    start_dt, _ = schedule_times_from_config(config, shift_date_iso)
    close_t = parse_time_value(config.get("provisional_close_time") or "02:00", time(2, 0))
    close_day = date.fromisoformat(shift_date_iso)
    entry_t = parse_time_value(config.get("entry_time"), time(8, 0))
    # Si la hora límite es menor/igual a la entrada, se entiende como el día siguiente.
    if close_t <= entry_t or bool(config.get("crosses_midnight")):
        close_day = close_day + timedelta(days=1)
    close_dt = datetime.combine(close_day, close_t, tzinfo=TZ)
    max_minutes = int(config.get("max_open_minutes") or 0)
    if max_minutes > 0:
        max_dt = start_dt + timedelta(minutes=max_minutes)
        # Se usa el límite más lejano para no cerrar turnos legítimos demasiado pronto.
        if max_dt > close_dt:
            close_dt = max_dt
    return close_dt


def auto_close_overdue_records(conn: Connection, user: str = "Sistema") -> int:
    now_dt = now_mx()
    open_rows = fetch_all(conn, "SELECT * FROM attendance WHERE (exit_at IS NULL OR TRIM(COALESCE(exit_at, '')) = '') ORDER BY id ASC")
    closed = 0
    for row in open_rows:
        config = get_shift_config(conn, row.get("turno") or "Día")
        if int(config.get("auto_close_enabled") or 1) != 1:
            continue
        deadline = provisional_close_deadline(config, row["shift_date"])
        if now_dt < deadline:
            continue
        _, scheduled_exit = schedule_times_from_config(config, row["shift_date"])
        eval_data = evaluate_exit(config, row["shift_date"], scheduled_exit)
        old_exit = row.get("exit_at") or ""
        ts = now_dt.isoformat()
        conn.execute(
            text(
                """
                UPDATE attendance
                SET exit_at=:exit_at,
                    exit_guard='Sistema',
                    exit_status='Salida provisional',
                    early_reason='',
                    scheduled_entry_at=COALESCE(NULLIF(scheduled_entry_at, ''), :scheduled_entry_at),
                    scheduled_exit_at=COALESCE(NULLIF(scheduled_exit_at, ''), :scheduled_exit_at),
                    entry_limit_at=COALESCE(NULLIF(entry_limit_at, ''), :entry_limit_at),
                    exit_early_limit_at=:exit_early_limit_at,
                    extra_limit_at=:extra_limit_at,
                    entry_tolerance_minutes=:entry_tolerance_minutes,
                    exit_tolerance_minutes=:exit_tolerance_minutes,
                    extra_after_minutes=:extra_after_minutes,
                    early_minutes=0,
                    extra_minutes=0,
                    provisional_exit=1,
                    review_required=1,
                    review_status='Pendiente',
                    auto_closed_at=:auto_closed_at,
                    updated_at=:updated_at,
                    incident=:incident
                WHERE id=:id
                """
            ),
            {
                "id": row["id"],
                "exit_at": scheduled_exit.isoformat(),
                "auto_closed_at": ts,
                "updated_at": ts,
                "incident": (row.get("incident") or "") + (" | " if row.get("incident") else "") + "Salida provisional automática: no registró salida dentro del límite configurado.",
                **eval_data,
            },
        )
        audit(conn, "attendance", str(row["id"]), "AUTO_CLOSE", user, "exit_at", old_exit, scheduled_exit.isoformat(), "Cierre provisional automático por registro abierto vencido")
        log_event(conn, "alerta", "asistencia", "auto_close", f"Registro {row['id']} cerrado provisionalmente", user, f"Empleado {row['employee_id']} · límite {deadline.isoformat()}")
        closed += 1
    return closed


def entry_status(turno: str, shift_date_iso: str, dt: datetime, tolerance_minutes: int = 10) -> str:
    config = default_shift_config(turno)
    config["entry_tolerance_minutes"] = tolerance_minutes
    return evaluate_entry(config, shift_date_iso, dt)["status"]


def exit_status(turno: str, shift_date_iso: str, dt: datetime, tolerance_minutes: int = 10) -> str:
    config = default_shift_config(turno)
    config["exit_tolerance_minutes"] = tolerance_minutes
    return evaluate_exit(config, shift_date_iso, dt)["status"]



async def save_upload(file: Optional[UploadFile], folder: Path, prefix: str) -> str:
    if not file or not file.filename:
        return ""
    suffix = Path(file.filename).suffix.lower() or ".jpg"
    if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
        suffix = ".jpg"
    name = f"{safe_filename(prefix)}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{suffix}"
    path = folder / name
    content = await file.read()
    path.write_bytes(content)
    return str(path.relative_to(UPLOADS_DIR)).replace("\\", "/")


# -----------------------------
# Usuarios, login y sesiones
# -----------------------------

def hash_password(password: str, salt: Optional[str] = None) -> str:
    salt = salt or secrets.token_hex(16)
    iterations = 200_000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_text, salt, expected = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), int(iterations_text)).hex()
        return secrets.compare_digest(digest, expected)
    except Exception:
        return False


def session_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def parse_dt(value: str) -> Optional[datetime]:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=TZ)
        return dt.astimezone(TZ)
    except Exception:
        return None


def seed_admin_user() -> None:
    """Crea/actualiza usuarios base solicitados.

    Para esta beta sincronizamos rol y clave en cada arranque para asegurar que Render
    quede con los accesos correctos. Las claves quedan hasheadas en base de datos.
    """
    with engine.begin() as conn:
        ts = now_mx().isoformat()
        for seed in DEFAULT_SYSTEM_USERS:
            existing = fetch_one(conn, "SELECT * FROM users WHERE username = :username", {"username": seed["username"]})
            password_hash = hash_password(seed["password"])
            if existing:
                old_role = existing.get("role") or ""
                conn.execute(
                    text("""
                        UPDATE users
                        SET password_hash=:password_hash, role=:role, active=1, display_name=:display_name,
                            failed_login_attempts=0, locked_until='', password_changed_at=:password_changed_at, updated_at=:updated_at
                        WHERE username=:username
                    """),
                    {
                        "username": seed["username"],
                        "password_hash": password_hash,
                        "role": seed["role"],
                        "display_name": seed["display_name"],
                        "password_changed_at": ts,
                        "updated_at": ts,
                    },
                )
                if old_role != seed["role"]:
                    audit(conn, "users", seed["username"], "UPDATE", "Sistema", "role", old_role, seed["role"], "Sincronización de usuarios base")
            else:
                conn.execute(
                    text("""
                        INSERT INTO users (username, password_hash, role, active, display_name, failed_login_attempts, locked_until, last_login_at, password_changed_at, created_at, updated_at)
                        VALUES (:username, :password_hash, :role, 1, :display_name, 0, '', '', :password_changed_at, :created_at, :updated_at)
                    """),
                    {
                        "username": seed["username"],
                        "password_hash": password_hash,
                        "role": seed["role"],
                        "display_name": seed["display_name"],
                        "password_changed_at": ts,
                        "created_at": ts,
                        "updated_at": ts,
                    },
                )
                audit(conn, "users", seed["username"], "CREATE", "Sistema", reason=f"Usuario base {seed['role']}")


def seed_guards() -> None:
    """Crea/actualiza vigilantes base solicitados para que los QR queden correctos en Render."""
    with engine.begin() as conn:
        ts = now_mx().isoformat()
        for guard in DEFAULT_GUARDS:
            existing = get_guard(conn, guard["code"])
            if existing:
                old_display = guard_display(existing)
                conn.execute(
                    text("""
                        UPDATE guards
                        SET alias=:alias, nombre=:nombre, active=:active, qr_activo=:qr_activo, updated_at=:updated_at
                        WHERE code=:code
                    """),
                    {"code": guard["code"], "alias": guard["alias"], "nombre": guard["nombre"], "active": guard["active"], "qr_activo": guard["qr_activo"], "updated_at": ts},
                )
                if old_display != f"{guard['alias']} - {guard['nombre']}":
                    audit(conn, "guards", guard["code"], "UPDATE", "Sistema", "nombre", old_display, f"{guard['alias']} - {guard['nombre']}", "Sincronización de vigilantes base")
            else:
                conn.execute(
                    text("""
                        INSERT INTO guards (code, alias, nombre, active, qr_activo, observaciones, created_at, updated_at)
                        VALUES (:code, :alias, :nombre, :active, :qr_activo, :observaciones, :created_at, :updated_at)
                    """),
                    {**guard, "observaciones": "Vigilante base", "created_at": ts, "updated_at": ts},
                )
                audit(conn, "guards", guard["code"], "CREATE", "Sistema", reason="Vigilante base")


def is_safe_next_url(value: str) -> bool:
    value = as_text(value)
    return bool(value.startswith("/") and not value.startswith("//") and "\\" not in value)


def get_client_key(request: Request, username: str) -> str:
    forwarded = request.headers.get("x-forwarded-for", "").split(",")[0].strip()
    ip = forwarded or (request.client.host if request.client else "unknown")
    return f"{ip}:{username.strip().lower()}"


def login_is_locked(conn: Connection, username: str) -> tuple[bool, str]:
    row = fetch_one(conn, "SELECT failed_login_attempts, locked_until FROM users WHERE username=:username", {"username": username})
    if not row:
        return False, ""
    locked_until = parse_dt(row.get("locked_until") or "")
    if locked_until and locked_until > now_mx():
        return True, locked_until.strftime("%H:%M")
    return False, ""


def register_login_failure(conn: Connection, username: str, request: Request) -> None:
    row = fetch_one(conn, "SELECT failed_login_attempts FROM users WHERE username=:username", {"username": username})
    if not row:
        return
    attempts = int(row.get("failed_login_attempts") or 0) + 1
    locked_until = ""
    if attempts >= LOGIN_MAX_ATTEMPTS:
        locked_until = (now_mx() + timedelta(minutes=LOGIN_LOCK_MINUTES)).isoformat()
    conn.execute(
        text("UPDATE users SET failed_login_attempts=:attempts, locked_until=:locked_until, updated_at=:updated_at WHERE username=:username"),
        {"username": username, "attempts": attempts, "locked_until": locked_until, "updated_at": now_mx().isoformat()},
    )
    log_event(conn, "alerta", "auth", "login_failure", f"Intento fallido de acceso para {username}", username, get_client_key(request, username))


def register_login_success(conn: Connection, username: str) -> None:
    conn.execute(
        text("UPDATE users SET failed_login_attempts=0, locked_until='', last_login_at=:last_login_at, updated_at=:updated_at WHERE username=:username"),
        {"username": username, "last_login_at": now_mx().isoformat(), "updated_at": now_mx().isoformat()},
    )

def get_current_user(request: Request):
    token = request.cookies.get(AUTH_COOKIE_NAME, "")
    if not token:
        return None
    token_digest = session_hash(token)
    with engine.begin() as conn:
        row = fetch_one(
            conn,
            """
            SELECT s.token_hash, s.username, s.expires_at, u.role, u.active
            FROM user_sessions s
            JOIN users u ON u.username = s.username
            WHERE s.token_hash = :token_hash
            """,
            {"token_hash": token_digest},
        )
        if not row or not row.get("active"):
            return None
        expires = parse_dt(row.get("expires_at", ""))
        if not expires or expires < now_mx():
            conn.execute(text("DELETE FROM user_sessions WHERE token_hash = :token_hash"), {"token_hash": token_digest})
            return None
        return {"username": row["username"], "role": row["role"]}


def admin_login_redirect(request: Request) -> RedirectResponse:
    target = request.url.path
    if request.url.query:
        target += "?" + request.url.query
    return RedirectResponse(url=f"/login?next={quote(target)}", status_code=303)


def require_role_page(request: Request, roles: set[str]):
    user = get_current_user(request)
    if not user:
        return None
    if user.get("role") not in roles:
        raise HTTPException(status_code=403, detail="No tienes permiso para este módulo")
    return user


def require_role_http(request: Request, roles: set[str]):
    user = require_role_page(request, roles)
    if not user:
        raise HTTPException(status_code=401, detail="Inicia sesión")
    return user


def require_admin_page(request: Request):
    return require_role_page(request, {"Admin", "Supremo"})


def require_admin_http(request: Request):
    return require_role_http(request, {"Admin", "Supremo"})


def require_supremo_page(request: Request):
    return require_role_page(request, {"Supremo"})


def require_supremo_http(request: Request):
    return require_role_http(request, {"Supremo"})


def require_rh_page(request: Request):
    return require_role_page(request, {"RH", "Admin", "Supremo"})


def require_rh_http(request: Request):
    return require_role_http(request, {"RH", "Admin", "Supremo"})


def require_vigilancia_http(request: Request):
    return require_role_http(request, {"Admin", "Supremo", "Vigilancia"})


templates.env.globals["get_current_user"] = get_current_user
templates.env.globals["short_datetime"] = short_datetime if "short_datetime" in globals() else lambda x: x


# -----------------------------
# Excel export helpers
# -----------------------------

def bool_text(value) -> str:
    return "Sí" if bool(value) else "No"


def short_datetime(value) -> str:
    if not value:
        return ""
    text_value = str(value)
    try:
        dt = datetime.fromisoformat(text_value.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return text_value[:16]




def format_minutes_duration(value) -> str:
    """Visualiza minutos como duración HH:MM.
    Ejemplos MS: 50 -> 00:50, 70 -> 1:10.
    La base conserva minutos numéricos para cálculos y nómina.
    """
    try:
        minutes = int(float(value or 0))
    except Exception:
        minutes = 0
    if minutes < 0:
        minutes = 0
    hours = minutes // 60
    mins = minutes % 60
    if hours == 0:
        return f"00:{mins:02d}"
    return f"{hours}:{mins:02d}"


# Se asigna aquí también porque short_datetime se define después de los helpers de sesión.
templates.env.globals["short_datetime"] = short_datetime
templates.env.globals["fmt_minutes"] = format_minutes_duration


def date_filter_clause(alias: str = "a", fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    conditions = []
    params = {}
    if fecha_inicio:
        conditions.append(f"{alias}.shift_date >= :fecha_inicio")
        params["fecha_inicio"] = fecha_inicio
    if fecha_fin:
        conditions.append(f"{alias}.shift_date <= :fecha_fin")
        params["fecha_fin"] = fecha_fin
    return (" AND " + " AND ".join(conditions) if conditions else "", params)


def style_worksheet(ws, title: str = "") -> None:
    header_fill = PatternFill("solid", fgColor="051A39")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="051A39", bold=True, size=16)
    thin = Side(style="thin", color="D9DEE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if title:
        ws.insert_rows(1)
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A3"
        header_row = 2
    else:
        ws.freeze_panes = "A2"
        header_row = 1

    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < header_row:
        return

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 42))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 38)


def workbook_response(wb: Workbook, filename: str):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def make_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list], title: str = ""):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_worksheet(ws, title or name)
    return ws


def remove_default_sheet(wb: Workbook) -> None:
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]


def get_employees_export_rows(conn: Connection):
    employees = fetch_all(conn, "SELECT * FROM employees ORDER BY nombre")
    headers = [
        "ID empleado", "Nombre", "Área", "Puesto", "Turno", "Estado", "Tiene vehículo",
        "QR activo", "Observaciones", "Creado", "Actualizado"
    ]
    rows = [
        [
            emp["id"], emp["nombre"], emp["area"], emp["puesto"], emp["turno"], emp["estado"],
            bool_text(emp["tiene_vehiculo"]), bool_text(emp["qr_activo"]),
            emp["observaciones"], short_datetime(emp["created_at"]), short_datetime(emp["updated_at"]),
        ]
        for emp in employees
    ]
    return headers, rows


def get_attendance_export_rows(conn: Connection, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    clause, params = date_filter_clause("a", fecha_inicio, fecha_fin)
    records = fetch_all(conn, f"""
        SELECT a.*, e.nombre, e.area, e.puesto
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        WHERE 1=1 {clause}
        ORDER BY a.shift_date DESC, a.entry_at DESC
    """, params)
    headers = [
        "ID registro", "ID empleado", "Nombre", "Área", "Puesto", "Fecha turno", "Turno",
        "Entrada", "Salida", "Entrada programada", "Salida programada", "Límite retardo", "Límite extra",
        "Guardia entrada", "Guardia salida", "Estado entrada", "Estado salida",
        "Min retardo", "Min salida temprana", "Min extra", "Comida tomada", "Min trabajados",
        "Motivo retardo", "Motivo salida temprana", "Salida provisional", "Pendiente revisión", "Estado revisión", "Anulado",
        "Vehículo esperado", "Vehículo registrado", "Incidencia / observaciones", "Creado", "Actualizado"
    ]
    rows = [
        [
            row["id"], row["employee_id"], row.get("nombre") or "", row.get("area") or "", row.get("puesto") or "",
            row["shift_date"], row["turno"], short_datetime(row["entry_at"]), short_datetime(row["exit_at"]),
            short_datetime(row.get("scheduled_entry_at")), short_datetime(row.get("scheduled_exit_at")),
            short_datetime(row.get("entry_limit_at")), short_datetime(row.get("extra_limit_at")),
            row["entry_guard"], row["exit_guard"], row["entry_status"], row["exit_status"],
            row.get("late_minutes") or 0, row.get("early_minutes") or 0, row.get("extra_minutes") or 0, row.get("lunch_taken") or "", row.get("worked_minutes") or 0,
            row["late_reason"], row["early_reason"], bool_text(row.get("provisional_exit") or 0), bool_text(row.get("review_required") or 0), row.get("review_status") or "", bool_text(row.get("anulled") or 0),
            bool_text(row["vehicle_expected"]), bool_text(row["vehicle_entered"]), row["incident"], short_datetime(row["created_at"]), short_datetime(row["updated_at"]),
        ]
        for row in records
    ]
    return headers, rows


def get_incidents_export_rows(conn: Connection, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    clause, params = date_filter_clause("a", fecha_inicio, fecha_fin)
    records = fetch_all(conn, f"""
        SELECT a.*, e.nombre, e.area, e.puesto
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        WHERE 1=1 {clause}
          AND (
            COALESCE(a.entry_status, '') IN ('Tarde', 'Retardo')
            OR COALESCE(a.exit_status, '') IN ('Salida temprana', 'Extra')
            OR COALESCE(a.incident, '') != ''
          )
        ORDER BY a.updated_at DESC
    """, params)
    headers = [
        "ID registro", "ID empleado", "Nombre", "Área", "Fecha turno", "Turno", "Tipo incidencia",
        "Entrada", "Salida", "Min retardo", "Min temprano", "Min extra", "Motivo", "Observaciones", "Vehículo registrado"
    ]
    rows = []
    for row in records:
        tipo = []
        if row["entry_status"] in {"Tarde", "Retardo"}:
            tipo.append("Retardo")
        if row["exit_status"] == "Salida temprana":
            tipo.append("Salida temprana")
        if row["exit_status"] == "Extra":
            tipo.append("Extra")
        if row["incident"]:
            tipo.append("Observación")
        rows.append([
            row["id"], row["employee_id"], row.get("nombre") or "", row.get("area") or "", row["shift_date"], row["turno"],
            ", ".join(tipo) or "Incidencia", short_datetime(row["entry_at"]), short_datetime(row["exit_at"]),
            row.get("late_minutes") or 0, row.get("early_minutes") or 0, row.get("extra_minutes") or 0,
            row["late_reason"] or row["early_reason"] or "", row["incident"], bool_text(row["vehicle_entered"]),
        ])
    return headers, rows


def get_audit_export_rows(conn: Connection):
    records = fetch_all(conn, "SELECT * FROM audit_log ORDER BY created_at DESC")
    headers = ["ID", "Tabla", "ID registro", "Acción", "Usuario", "Campo", "Valor anterior", "Valor nuevo", "Motivo", "Fecha"]
    rows = [
        [
            row["id"], row["table_name"], row["record_id"], row["action"], row["user_name"], row["field_name"],
            row["old_value"], row["new_value"], row["reason"], short_datetime(row["created_at"]),
        ]
        for row in records
    ]
    return headers, rows


def add_summary_sheet(wb: Workbook, conn: Connection, fecha_inicio: Optional[str], fecha_fin: Optional[str]) -> None:
    ws = wb.create_sheet("Resumen")
    ws.append(["Reporte general de asistencia"])
    ws.append(["Generado", now_mx().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Fecha inicio", fecha_inicio or "Sin filtro"])
    ws.append(["Fecha fin", fecha_fin or "Sin filtro"])
    ws.append([])

    clause, params = date_filter_clause("a", fecha_inicio, fecha_fin)
    total_empleados = conn.execute(text("SELECT COUNT(*) FROM employees")).scalar() or 0
    activos = conn.execute(text("SELECT COUNT(*) FROM employees WHERE estado = 'Activo'")).scalar() or 0
    total_asistencias = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE 1=1 {clause}"), params).scalar() or 0
    retardos = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE entry_status IN ('Tarde','Retardo') {clause}"), params).scalar() or 0
    salidas_tempranas = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE exit_status = 'Salida temprana' {clause}"), params).scalar() or 0
    extras = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE exit_status = 'Extra' {clause}"), params).scalar() or 0
    vehiculos = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE vehicle_entered = 1 {clause}"), params).scalar() or 0
    abiertas = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE (a.exit_at IS NULL OR TRIM(COALESCE(a.exit_at, '')) = '') {clause}"), params).scalar() or 0
    provisionales = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE provisional_exit = 1 {clause}"), params).scalar() or 0
    pendientes_revision = conn.execute(text(f"SELECT COUNT(*) FROM attendance a WHERE review_required = 1 {clause}"), params).scalar() or 0

    ws.append(["Indicador", "Valor"])
    for label, value in [
        ("Empleados registrados", total_empleados),
        ("Empleados activos", activos),
        ("Registros de asistencia", total_asistencias),
        ("Retardos", retardos),
        ("Salidas tempranas", salidas_tempranas),
        ("Extras", extras),
        ("Entradas con vehículo", vehiculos),
        ("Entradas abiertas", abiertas),
        ("Salidas provisionales", provisionales),
        ("Pendientes revisión", pendientes_revision),
    ]:
        ws.append([label, value])

    ws.append([])
    ws.append(["Nota", "El QR contiene únicamente el ID del empleado. Esta versión no captura imágenes ni placas; la evidencia visual se mantiene por WhatsApp."])
    style_worksheet(ws, "")
    ws["A1"].font = Font(color="051A39", bold=True, size=16)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42



# -----------------------------
# Correcciones masivas desde Excel
# -----------------------------

def normalize_header(value) -> str:
    txt = as_text(value).lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n",
        "_": " ", "-": " ", "/": " ", ".": " ",
    }
    for old, new in replacements.items():
        txt = txt.replace(old, new)
    return " ".join(txt.split())


def parse_date_value(value) -> Optional[str]:
    if value is None or as_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    txt = as_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(txt, fmt).date().isoformat()
        except Exception:
            pass
    return txt


def parse_time_or_datetime(value, shift_date_iso: str, turno: str, kind: str, entry_dt: Optional[datetime] = None):
    if value is None or as_text(value) == "":
        return None, ""
    if as_text(value).upper() == "BORRAR":
        return "__CLEAR__", ""
    try:
        if isinstance(value, datetime):
            dt = value
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=TZ)
            return dt.astimezone(TZ), ""
        if isinstance(value, time):
            t = value
        elif isinstance(value, (int, float)):
            # Excel puede guardar una hora como fracción del día: 0.5 = 12:00. Gracias, Excel, muy normal todo.
            if 0 <= float(value) < 1:
                total_seconds = int(round(float(value) * 24 * 60 * 60))
                t = (datetime.min + timedelta(seconds=total_seconds)).time()
            else:
                return None, f"Valor numérico de hora inválido: {value}"
        else:
            txt = as_text(value)
            # Si viene fecha y hora completa, usarla tal cual.
            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
                try:
                    dt = datetime.strptime(txt, fmt).replace(tzinfo=TZ)
                    return dt, ""
                except Exception:
                    pass
            # Si solo viene hora, combinar con fecha de turno.
            for fmt in ("%H:%M", "%H:%M:%S"):
                try:
                    t = datetime.strptime(txt, fmt).time()
                    break
                except Exception:
                    t = None
            if t is None:
                return None, f"Formato de hora inválido: {txt}"

        base_date = date.fromisoformat(shift_date_iso)
        if kind == "salida" and (turno or "").lower().startswith("n"):
            base_date = base_date + timedelta(days=1)
        dt = datetime.combine(base_date, t, tzinfo=TZ)
        if kind == "salida" and entry_dt and dt < entry_dt:
            dt = dt + timedelta(days=1)
        return dt, ""
    except Exception as exc:
        return None, f"No se pudo interpretar fecha/hora: {exc}"


def correction_export_headers() -> list[str]:
    return [
        "id_registro", "id_empleado", "nombre", "area", "fecha_turno_actual", "turno_actual",
        "entrada_actual", "salida_actual", "estado_entrada_actual", "estado_salida_actual",
        "motivo_retardo_actual", "motivo_salida_temprana_actual", "observaciones_actuales", "updated_at_actual",
        "entrada_nueva", "salida_nueva", "turno_nuevo", "fecha_turno_nueva",
        "motivo_retardo_nuevo", "motivo_salida_temprana_nuevo", "observaciones_nuevas", "motivo_correccion",
    ]


def build_editable_corrections_workbook(conn: Connection, fecha_inicio: Optional[str], fecha_fin: Optional[str]) -> Workbook:
    clause, params = date_filter_clause("a", fecha_inicio, fecha_fin)
    records = fetch_all(conn, f"""
        SELECT a.*, e.nombre, e.area
        FROM attendance a
        LEFT JOIN employees e ON e.id = a.employee_id
        WHERE 1=1 {clause}
        ORDER BY a.shift_date DESC, a.entry_at DESC
    """, params)
    wb = Workbook()
    ws = wb.active
    ws.title = "correcciones"
    headers = correction_export_headers()
    ws.append(headers)
    for row in records:
        ws.append([
            row["id"], row["employee_id"], row.get("nombre") or "", row.get("area") or "",
            row["shift_date"], row["turno"], short_datetime(row["entry_at"]), short_datetime(row["exit_at"]),
            row["entry_status"], row["exit_status"], row["late_reason"], row["early_reason"], row["incident"], row["updated_at"],
            "", "", "", "", "", "", "", "",
        ])
    style_worksheet(ws, "Corrección masiva editable")
    note = wb.create_sheet("instrucciones")
    note.append(["Uso"])
    note.append(["Edita solo columnas que terminan en _nueva y motivo_correccion."])
    note.append(["Una celda nueva vacía significa: no modificar."])
    note.append(["Para borrar un valor editable, escribe BORRAR."])
    note.append(["Al reimportar, el sistema compara contra la base, detecta alteraciones, valida y audita cada cambio."])
    style_worksheet(note, "Instrucciones")
    return wb




def find_excel_header_row(ws, required_headers: list[str], max_scan_rows: int = 15):
    """Busca la fila real de encabezados en un Excel.

    Las plantillas del sistema pueden traer una fila de título arriba de los encabezados.
    El importador debe encontrar id_registro, id_empleado y updated_at_actual aunque estén en fila 2.
    Porque, increíblemente, el propio sistema también merece ser compatible consigo mismo.
    """
    required_normalized = [normalize_header(h) for h in required_headers]
    rows_to_scan = min(ws.max_row or 1, max_scan_rows)

    for row_idx in range(1, rows_to_scan + 1):
        raw_headers = [cell.value for cell in ws[row_idx]]
        headers = {
            normalize_header(value): idx
            for idx, value in enumerate(raw_headers)
            if normalize_header(value)
        }
        if all(req in headers for req in required_normalized):
            return row_idx, headers

    # Regresa la fila 1 como respaldo para poder mostrar un error claro.
    raw_headers = [cell.value for cell in ws[1]]
    headers = {
        normalize_header(value): idx
        for idx, value in enumerate(raw_headers)
        if normalize_header(value)
    }
    return 1, headers


def analyze_corrections_excel(content: bytes, admin_username: str):
    errors = []
    changes = []
    total_rows = 0
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        return {"ok": False, "errors": [f"No se pudo leer el Excel: {exc}"], "changes": [], "total_rows": 0}
    ws = wb["correcciones"] if "correcciones" in wb.sheetnames else wb.active
    required = ["id_registro", "id_empleado", "updated_at_actual"]
    header_row, headers = find_excel_header_row(ws, required)

    def col(name: str):
        return headers.get(normalize_header(name))

    missing = [name for name in required if col(name) is None]
    if missing:
        return {
            "ok": False,
            "errors": [
                "Faltan columnas requeridas: " + ", ".join(missing) +
                ". Descarga nuevamente la plantilla desde Correcciones o verifica que no hayas borrado los encabezados."
            ],
            "changes": [],
            "total_rows": 0,
        }

    data_start_row = header_row + 1
    seen_records = set()
    with engine.begin() as conn:
        for row_number, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if not any(row):
                continue
            total_rows += 1
            record_id = as_text(row[col("id_registro")])
            employee_id = clean_employee_id(as_text(row[col("id_empleado")]))
            version_excel = as_text(row[col("updated_at_actual")])
            motivo = as_text(row[col("motivo_correccion")]) if col("motivo_correccion") is not None else ""
            if not record_id:
                errors.append(f"Fila {row_number}: falta id_registro")
                continue
            if record_id in seen_records:
                errors.append(f"Fila {row_number}: registro duplicado en el Excel: {record_id}")
                continue
            seen_records.add(record_id)
            db = fetch_one(conn, "SELECT * FROM attendance WHERE id = :id", {"id": record_id})
            if not db:
                errors.append(f"Fila {row_number}: registro {record_id} no existe")
                continue
            if clean_id(db["employee_id"]) != employee_id:
                errors.append(f"Fila {row_number}: el empleado no coincide con el registro {record_id}")
                continue
            if as_text(db["updated_at"]) != version_excel:
                errors.append(f"Fila {row_number}: el registro {record_id} fue modificado después de exportar. Exporta de nuevo esa fila.")
                continue

            new_shift_date = db["shift_date"]
            if col("fecha_turno_nueva") is not None:
                val = row[col("fecha_turno_nueva")]
                if as_text(val):
                    if as_text(val).upper() == "BORRAR":
                        errors.append(f"Fila {row_number}: fecha_turno_nueva no se puede borrar")
                        continue
                    parsed = parse_date_value(val)
                    if not parsed:
                        errors.append(f"Fila {row_number}: fecha_turno_nueva inválida")
                        continue
                    new_shift_date = parsed

            new_turno = db["turno"]
            if col("turno_nuevo") is not None:
                val = as_text(row[col("turno_nuevo")])
                if val:
                    if val.lower() in {"dia", "día"}:
                        new_turno = "Día"
                    elif val.lower() == "noche":
                        new_turno = "Noche"
                    else:
                        errors.append(f"Fila {row_number}: turno_nuevo inválido: {val}")
                        continue

            row_changes = []
            def add_change(field, old, new):
                old_txt = "" if old is None else str(old)
                new_txt = "" if new is None else str(new)
                if old_txt != new_txt:
                    row_changes.append({"record_id": record_id, "field": field, "old": old_txt, "new": new_txt, "row_number": row_number, "reason": motivo})

            # Horas
            entry_dt_current = parse_dt(db["entry_at"]) if db.get("entry_at") else None
            entry_value = db.get("entry_at") or ""
            if col("entrada_nueva") is not None and as_text(row[col("entrada_nueva")]):
                parsed, err = parse_time_or_datetime(row[col("entrada_nueva")], new_shift_date, new_turno, "entrada")
                if err:
                    errors.append(f"Fila {row_number}: {err}")
                    continue
                entry_value = "" if parsed == "__CLEAR__" else parsed.isoformat()
                entry_dt_current = None if parsed == "__CLEAR__" else parsed
                add_change("entry_at", db.get("entry_at") or "", entry_value)

            exit_value = db.get("exit_at") or ""
            if col("salida_nueva") is not None and as_text(row[col("salida_nueva")]):
                parsed, err = parse_time_or_datetime(row[col("salida_nueva")], new_shift_date, new_turno, "salida", entry_dt_current)
                if err:
                    errors.append(f"Fila {row_number}: {err}")
                    continue
                exit_value = "" if parsed == "__CLEAR__" else parsed.isoformat()
                add_change("exit_at", db.get("exit_at") or "", exit_value)

            if entry_value and exit_value:
                entry_dt = parse_dt(entry_value)
                exit_dt = parse_dt(exit_value)
                if entry_dt and exit_dt and exit_dt < entry_dt:
                    errors.append(f"Fila {row_number}: la salida queda antes que la entrada")
                    continue

            add_change("shift_date", db["shift_date"], new_shift_date)
            add_change("turno", db["turno"], new_turno)

            # Textos editables
            field_map = [
                ("motivo_retardo_nuevo", "late_reason"),
                ("motivo_salida_temprana_nuevo", "early_reason"),
                ("observaciones_nuevas", "incident"),
            ]
            for excel_col, db_field in field_map:
                if col(excel_col) is not None and as_text(row[col(excel_col)]):
                    val = as_text(row[col(excel_col)])
                    new_val = "" if val.upper() == "BORRAR" else val
                    add_change(db_field, db.get(db_field) or "", new_val)

            # Recalcular estados si se movieron horas/turno/fecha.
            entry_for_status = parse_dt(entry_value) if entry_value else None
            exit_for_status = parse_dt(exit_value) if exit_value else None
            if entry_for_status:
                add_change("entry_status", db.get("entry_status") or "", entry_status(new_turno, new_shift_date, entry_for_status))
            if exit_for_status:
                add_change("exit_status", db.get("exit_status") or "", exit_status(new_turno, new_shift_date, exit_for_status))

            if row_changes and not motivo:
                errors.append(f"Fila {row_number}: falta motivo_correccion")
                continue
            changes.extend(row_changes)

    return {"ok": len(errors) == 0, "errors": errors, "changes": changes, "total_rows": total_rows}


def save_correction_preview(analysis: dict, admin_username: str, original_filename: str = "") -> str:
    batch_id = f"CORR-{now_mx().strftime('%Y%m%d-%H%M%S')}-{secrets.token_hex(3).upper()}"
    path = CORRECTIONS_DIR / f"{batch_id}.json"
    payload = {"batch_id": batch_id, "admin": admin_username, "original_filename": original_filename, "analysis": analysis, "created_at": now_mx().isoformat()}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO correction_batches (batch_id, user_name, status, file_path, total_rows, total_changes, total_errors, created_at)
            VALUES (:batch_id, :user_name, 'preview', :file_path, :total_rows, :total_changes, :total_errors, :created_at)
        """), {
            "batch_id": batch_id,
            "user_name": admin_username,
            "file_path": str(path),
            "total_rows": analysis.get("total_rows", 0),
            "total_changes": len(analysis.get("changes", [])),
            "total_errors": len(analysis.get("errors", [])),
            "created_at": now_mx().isoformat(),
        })
    return batch_id


def apply_correction_batch(batch_id: str, admin_username: str) -> dict:
    batch_id = clean_id(batch_id)
    path = CORRECTIONS_DIR / f"{batch_id}.json"
    if not path.exists():
        return {"ok": False, "message": "Lote no encontrado"}
    payload = json.loads(path.read_text(encoding="utf-8"))
    analysis = payload.get("analysis", {})
    if analysis.get("errors"):
        return {"ok": False, "message": "No se puede aplicar un lote con errores"}
    changes = analysis.get("changes", [])
    if not changes:
        return {"ok": False, "message": "No hay cambios para aplicar"}

    grouped = {}
    for change in changes:
        grouped.setdefault(str(change["record_id"]), []).append(change)

    allowed_fields = {"entry_at", "exit_at", "shift_date", "turno", "late_reason", "early_reason", "incident", "entry_status", "exit_status"}
    applied = 0
    ts = now_mx().isoformat()
    with engine.begin() as conn:
        batch = fetch_one(conn, "SELECT * FROM correction_batches WHERE batch_id = :batch_id", {"batch_id": batch_id})
        if batch and batch.get("status") == "applied":
            return {"ok": False, "message": "Este lote ya fue aplicado"}
        for record_id, record_changes in grouped.items():
            set_parts = []
            params = {"id": record_id, "updated_at": ts}
            for change in record_changes:
                field = change["field"]
                if field not in allowed_fields:
                    continue
                param_name = f"v_{field}"
                set_parts.append(f"{field}=:{param_name}")
                if field in {"entry_at", "exit_at"} and change["new"] == "":
                    params[param_name] = None
                else:
                    params[param_name] = change["new"]
            if not set_parts:
                continue
            set_parts.append("updated_at=:updated_at")
            conn.execute(text(f"UPDATE attendance SET {', '.join(set_parts)} WHERE id=:id"), params)
            for change in record_changes:
                audit(conn, "attendance", record_id, "BULK_UPDATE", admin_username, field=change["field"], old=change["old"], new=change["new"], reason=f"{batch_id}: {change.get('reason','')}")
                applied += 1
        conn.execute(text("UPDATE correction_batches SET status='applied', applied_at=:applied_at WHERE batch_id=:batch_id"), {"applied_at": ts, "batch_id": batch_id})
    return {"ok": True, "message": f"Lote aplicado. Cambios: {applied}", "applied": applied}

def seed_demo_if_empty() -> None:
    with engine.begin() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM employees")).scalar() or 0
        if count:
            return
        ts = now_mx().isoformat()
        demo = [
            {
                "id": "123", "nombre": "Juan Pérez", "area": "Producción", "puesto": "Operador", "turno": "Día",
                "estado": "Activo", "tiene_vehiculo": 1, "requiere_fotos_vehiculo": 0, "foto_path": "", "qr_activo": 1,
                "observaciones": "Demo con vehículo", "created_at": ts, "updated_at": ts,
            },
            {
                "id": "124", "nombre": "María López", "area": "Calidad", "puesto": "Inspectora", "turno": "Día",
                "estado": "Activo", "tiene_vehiculo": 0, "requiere_fotos_vehiculo": 0, "foto_path": "", "qr_activo": 1,
                "observaciones": "Demo sin vehículo", "created_at": ts, "updated_at": ts,
            },
            {
                "id": "125", "nombre": "Carlos Ramos", "area": "Almacén", "puesto": "Auxiliar", "turno": "Noche",
                "estado": "Activo", "tiene_vehiculo": 1, "requiere_fotos_vehiculo": 0, "foto_path": "", "qr_activo": 1,
                "observaciones": "Demo turno noche", "created_at": ts, "updated_at": ts,
            },
        ]
        for item in demo:
            conn.execute(
                text(
                    """
                    INSERT INTO employees (id, nombre, area, puesto, turno, estado, tiene_vehiculo, requiere_fotos_vehiculo, foto_path, qr_activo, observaciones, created_at, updated_at)
                    VALUES (:id, :nombre, :area, :puesto, :turno, :estado, :tiene_vehiculo, :requiere_fotos_vehiculo, :foto_path, :qr_activo, :observaciones, :created_at, :updated_at)
                    """
                ), item
            )


@app.on_event("startup")
def startup() -> None:
    init_db()
    seed_admin_user()
    seed_guards()
    seed_demo_if_empty()

# Inicialización defensiva para pruebas locales y algunos runners.
init_db()
seed_admin_user()
seed_guards()
seed_demo_if_empty()


# -----------------------------
# Pages
# -----------------------------

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, next: str = "/monitor", error: str = ""):
    return templates.TemplateResponse("login.html", {"request": request, "next": next, "error": error})


@app.post("/login")
def login_submit(request: Request, username: str = Form(...), password: str = Form(...), next: str = Form("/monitor")):
    username = clean_username(username)
    with engine.begin() as conn:
        locked, locked_time = login_is_locked(conn, username)
        if locked:
            return templates.TemplateResponse(
                "login.html",
                {"request": request, "next": next or "/monitor", "error": f"Usuario bloqueado temporalmente por intentos fallidos. Intenta después de las {locked_time}."},
                status_code=429,
            )
        user = fetch_one(conn, "SELECT * FROM users WHERE username = :username AND active = 1", {"username": username})
        if not user or not verify_password(password, user["password_hash"]):
            if user:
                register_login_failure(conn, username, request)
            return templates.TemplateResponse(
                "login.html",
                {"request": request, "next": next or "/monitor", "error": "Usuario o clave incorrectos."},
                status_code=401,
            )
        register_login_success(conn, user["username"])
        token = secrets.token_urlsafe(32)
        expires_at = (now_mx() + timedelta(days=SESSION_DAYS)).isoformat()
        conn.execute(
            text(
                """
                INSERT INTO user_sessions (token_hash, username, expires_at, created_at)
                VALUES (:token_hash, :username, :expires_at, :created_at)
                """
            ),
            {"token_hash": session_hash(token), "username": user["username"], "expires_at": expires_at, "created_at": now_mx().isoformat()},
        )
        audit(conn, "users", user["username"], "LOGIN", user["username"], reason="Inicio de sesión")

    role = user.get("role")
    if not is_safe_next_url(next) or (next == "/monitor" and role in {"Vigilancia", "RH"}):
        if role == "Vigilancia":
            safe_next = "/captura"
        elif role == "RH":
            safe_next = "/rh"
        else:
            safe_next = "/monitor"
    else:
        safe_next = next
    response = RedirectResponse(url=safe_next, status_code=303)
    response.set_cookie(
        AUTH_COOKIE_NAME,
        token,
        httponly=True,
        secure=SESSION_COOKIE_SECURE,
        samesite="strict",
        max_age=SESSION_DAYS * 24 * 60 * 60,
    )
    return response


@app.get("/logout")
def logout(request: Request):
    token = request.cookies.get(AUTH_COOKIE_NAME, "")
    if token:
        with engine.begin() as conn:
            user = get_current_user(request)
            conn.execute(text("DELETE FROM user_sessions WHERE token_hash = :token_hash"), {"token_hash": session_hash(token)})
            if user:
                audit(conn, "users", user["username"], "LOGOUT", user["username"], reason="Cierre de sesión")
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(AUTH_COOKIE_NAME, path="/", secure=SESSION_COOKIE_SECURE, samesite="strict")
    return response


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    user = get_current_user(request)
    if user:
        if user.get("role") in {"Admin", "Supremo"}:
            return RedirectResponse(url="/monitor")
        if user.get("role") == "RH":
            return RedirectResponse(url="/rh")
    return RedirectResponse(url="/captura")


def render_vigilancia_page(request: Request):
    user = require_role_page(request, {"Admin", "Supremo", "Vigilancia"})
    if not user:
        return admin_login_redirect(request)
    with engine.begin() as conn:
        auto_close_overdue_records(conn, user.get("username", "Sistema"))
        active_guard = get_active_guard(conn)
    return templates.TemplateResponse("vigilancia.html", {"request": request, "active_user": user, "active_guard": active_guard})


@app.get("/captura", response_class=HTMLResponse)
def captura(request: Request):
    return render_vigilancia_page(request)


@app.get("/vigilancia", response_class=HTMLResponse)
def vigilancia(request: Request):
    return render_vigilancia_page(request)


@app.get("/vigilantes", response_class=HTMLResponse)
def vigilantes_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    with engine.begin() as conn:
        rows = fetch_all(conn, "SELECT * FROM guards ORDER BY alias, code")
        active_guard = get_active_guard(conn)
    return templates.TemplateResponse("vigilantes.html", {"request": request, "guards": rows, "active_guard": active_guard})


@app.post("/vigilantes/guardar")
def vigilante_guardar(
    request: Request,
    code: str = Form(...),
    alias: str = Form(...),
    nombre: str = Form(""),
    active: str = Form("1"),
    qr_activo: str = Form("1"),
    observaciones: str = Form(""),
):
    admin_user = require_admin_http(request)
    code = clean_guard_code(code)
    alias = as_text(alias) or code
    if not code:
        raise HTTPException(status_code=400, detail="Código de vigilante obligatorio")
    ts = now_mx().isoformat()
    with engine.begin() as conn:
        existing = get_guard(conn, code)
        params = {
            "code": code, "alias": alias[:80], "nombre": as_text(nombre)[:120],
            "active": 1 if active == "1" else 0, "qr_activo": 1 if qr_activo == "1" else 0,
            "observaciones": as_text(observaciones)[:500], "updated_at": ts,
        }
        if existing:
            conn.execute(text("""
                UPDATE guards
                SET alias=:alias, nombre=:nombre, active=:active, qr_activo=:qr_activo, observaciones=:observaciones, updated_at=:updated_at
                WHERE code=:code
            """), params)
            audit(conn, "guards", code, "UPDATE", admin_user["username"], reason="Edición de vigilante")
        else:
            conn.execute(text("""
                INSERT INTO guards (code, alias, nombre, active, qr_activo, observaciones, created_at, updated_at)
                VALUES (:code, :alias, :nombre, :active, :qr_activo, :observaciones, :created_at, :updated_at)
            """), {**params, "created_at": ts})
            audit(conn, "guards", code, "CREATE", admin_user["username"], reason="Alta de vigilante")
    return RedirectResponse(url="/vigilantes", status_code=303)


@app.post("/vigilantes/{guard_code}/estado")
def vigilante_estado(request: Request, guard_code: str, active: str = Form("0")):
    admin_user = require_admin_http(request)
    guard_code = clean_guard_code(guard_code)
    with engine.begin() as conn:
        guard = get_guard(conn, guard_code)
        if not guard:
            raise HTTPException(status_code=404, detail="Vigilante no encontrado")
        conn.execute(text("UPDATE guards SET active=:active, updated_at=:updated_at WHERE code=:code"), {"active": 1 if active == "1" else 0, "updated_at": now_mx().isoformat(), "code": guard_code})
        audit(conn, "guards", guard_code, "UPDATE", admin_user["username"], "active", str(guard.get("active")), active, "Cambio de estado de vigilante")
    return RedirectResponse(url="/vigilantes", status_code=303)


@app.post("/vigilantes/{guard_code}/borrar")
def vigilante_borrar(request: Request, guard_code: str, confirmacion: str = Form("")):
    supremo = require_supremo_http(request)
    guard_code = clean_guard_code(guard_code)
    if as_text(confirmacion).upper() != "BORRAR":
        raise HTTPException(status_code=400, detail="Para borrar definitivamente escribe BORRAR")
    with engine.begin() as conn:
        guard = get_guard(conn, guard_code)
        if not guard:
            raise HTTPException(status_code=404, detail="Vigilante no encontrado")
        # Si está activo, primero se limpia el estado.
        conn.execute(text("DELETE FROM active_guard_state WHERE guard_code=:code"), {"code": guard_code})
        conn.execute(text("DELETE FROM guards WHERE code=:code"), {"code": guard_code})
        audit(conn, "guards", guard_code, "DELETE", supremo["username"], reason="Borrado definitivo de vigilante en beta")
        log_event(conn, "alert", "vigilantes", "delete_guard", f"Vigilante borrado: {guard_code}", supremo["username"])
    return RedirectResponse(url="/vigilantes", status_code=303)


@app.get("/monitor", response_class=HTMLResponse)
def monitor(
    request: Request,
    period: str = "day",
    fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    today_date = now_mx().date()
    today = today_date.isoformat()
    period, start, end, period_label, selected_month, selected_year = resolve_dashboard_period_range(
        period=period,
        fecha=fecha,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        mes=mes,
        anio=anio,
    )
    month_start, month_end = month_bounds()
    year_start, year_end = year_bounds()
    month_options = dashboard_month_options(selected_year)
    year_options = dashboard_year_options(selected_year)
    with engine.begin() as conn:
        auto_close_overdue_records(conn, admin_user["username"])
        total_employees = conn.execute(text("SELECT COUNT(*) FROM employees WHERE estado = 'Activo'")).scalar() or 0
        present_today = conn.execute(
            text("SELECT COUNT(*) FROM attendance WHERE shift_date = :today AND entry_at IS NOT NULL AND COALESCE(anulled,0)=0"),
            {"today": today}
        ).scalar() or 0
        absent_today = fetch_all(conn,
            """
            SELECT e.id, e.nombre, e.area, e.puesto, e.turno
            FROM employees e
            WHERE e.estado = 'Activo'
              AND NOT EXISTS (
                SELECT 1 FROM attendance a
                WHERE a.employee_id = e.id AND a.shift_date = :today AND COALESCE(a.anulled,0)=0
              )
            ORDER BY e.area, e.nombre
            """,
            {"today": today}
        )
        absent_count = len(absent_today)
        day_stats = period_attendance_stats(conn, today_date, today_date)
        month_stats = period_attendance_stats(conn, month_start, month_end)
        year_stats = period_attendance_stats(conn, year_start, year_end)
        period_stats = period_attendance_stats(conn, start, end)
        late_today = day_stats["retardos"]
        extra_today = day_stats["extras"]
        vehicles_inside = conn.execute(
            text("SELECT COUNT(*) FROM attendance WHERE (exit_at IS NULL OR TRIM(COALESCE(exit_at, '')) = '') AND vehicle_entered = 1 AND COALESCE(anulled,0)=0")
        ).scalar() or 0
        inside = fetch_all(conn,
            """
            SELECT a.*, e.nombre, e.area, e.puesto
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            WHERE (a.exit_at IS NULL OR TRIM(COALESCE(a.exit_at, '')) = '') AND COALESCE(a.anulled,0)=0
            ORDER BY a.entry_at DESC
            """
        )
        incidents = fetch_all(conn,
            """
            SELECT a.*, e.nombre, e.area
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            WHERE a.shift_date = :today
              AND COALESCE(a.anulled,0)=0
              AND (a.entry_status IN ('Tarde', 'Retardo')
                   OR a.exit_status IN ('Salida temprana', 'Extra')
                   OR a.incident != ''
                   OR a.review_required = 1)
            ORDER BY a.updated_at DESC
            LIMIT 40
            """,
            {"today": today}
        )
        ranking_rows = employee_period_rows(conn, start, end)
    ranking_retardos_dash = sorted(
        [r for r in ranking_rows if int(r.get("retardos") or 0) > 0 or int(r.get("min_retardo") or 0) > 0],
        key=lambda r: (int(r.get("min_retardo") or 0), int(r.get("retardos") or 0), str(r.get("nombre") or "")),
        reverse=True,
    )[:8]
    ranking_puntuales = sorted(
        [r for r in ranking_rows if int(r.get("registros") or 0) > 0 and int(r.get("retardos") or 0) == 0 and int(r.get("faltas") or 0) == 0],
        key=lambda r: (int(r.get("registros") or 0), -int(r.get("salidas_tempranas") or 0), str(r.get("nombre") or "")),
        reverse=True,
    )[:8]
    return templates.TemplateResponse(
        "monitor.html",
        {
            "request": request,
            "total_employees": total_employees,
            "present_today": present_today,
            "late_today": late_today,
            "extra_today": extra_today,
            "vehicles_inside": vehicles_inside,
            "absent_count": absent_count,
            "absent_today": absent_today,
            "inside": inside,
            "incidents": incidents,
            "day_stats": day_stats,
            "month_stats": month_stats,
            "year_stats": year_stats,
            "period_stats": period_stats,
            "period": period,
            "period_label": period_label,
            "start": start,
            "end": end,
            "selected_month": selected_month,
            "selected_year": selected_year,
            "month_options": month_options,
            "year_options": year_options,
            "ranking_retardos_dash": ranking_retardos_dash,
            "ranking_puntuales": ranking_puntuales,
            "today": today,
        }
    )


@app.get("/configuracion", response_class=HTMLResponse)
def configuracion_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    today = now_mx().date().isoformat()
    with engine.begin() as conn:
        turnos = fetch_all(conn, "SELECT * FROM shift_settings ORDER BY name")
        empleados_activos = conn.execute(text("SELECT COUNT(*) FROM employees WHERE estado = 'Activo'")).scalar() or 0
        registros_abiertos = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE (exit_at IS NULL OR TRIM(COALESCE(exit_at, '')) = '')")).scalar() or 0
        retardos_hoy = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date = :today AND entry_status IN ('Tarde','Retardo')"), {"today": today}).scalar() or 0
    return templates.TemplateResponse(
        "configuracion.html",
        {
            "request": request,
            "turnos": turnos,
            "empleados_activos": empleados_activos,
            "registros_abiertos": registros_abiertos,
            "retardos_hoy": retardos_hoy,
            "today": today,
        },
    )


@app.post("/configuracion/cierre-seguro", response_class=HTMLResponse)
def configuracion_cierre_seguro(
    request: Request,
    fecha: str = Form(""),
    turno: str = Form("Día"),
    hora_salida: str = Form("19:00"),
    confirmacion: str = Form(""),
):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    today = now_mx().date().isoformat()
    cierre_result = None
    cierre_error = ""
    if as_text(confirmacion).upper() != "CERRAR":
        cierre_error = "Para aplicar el cierre seguro escribe CERRAR."
    else:
        with engine.begin() as conn:
            cierre_result = close_shift_day_safely(conn, fecha or today, turno, hora_salida, admin_user["username"])
    with engine.begin() as conn:
        turnos = fetch_all(conn, "SELECT * FROM shift_settings ORDER BY name")
        empleados_activos = conn.execute(text("SELECT COUNT(*) FROM employees WHERE estado = 'Activo'")).scalar() or 0
        registros_abiertos = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE (exit_at IS NULL OR TRIM(COALESCE(exit_at, '')) = '')")).scalar() or 0
        retardos_hoy = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date = :today AND entry_status IN ('Tarde','Retardo')"), {"today": today}).scalar() or 0
    return templates.TemplateResponse(
        "configuracion.html",
        {
            "request": request,
            "turnos": turnos,
            "empleados_activos": empleados_activos,
            "registros_abiertos": registros_abiertos,
            "retardos_hoy": retardos_hoy,
            "today": today,
            "cierre_result": cierre_result,
            "cierre_error": cierre_error,
            "cierre_fecha": fecha or today,
            "cierre_turno": turno,
            "cierre_hora": hora_salida,
        },
    )



@app.post("/configuracion/estado-planta", response_class=HTMLResponse)
def configuracion_estado_planta(
    request: Request,
    employee_id: str = Form(""),
    accion: str = Form("dentro"),
    fecha: str = Form(""),
    turno: str = Form(""),
    hora: str = Form(""),
    confirmacion: str = Form(""),
):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)

    today = now_mx().date().isoformat()
    estado_result = None
    estado_error = ""
    employee_id = clean_employee_id(employee_id)
    accion = as_text(accion).lower()
    if as_text(confirmacion).upper() != "CONFIRMAR":
        estado_error = "Para aplicar el ajuste de planta escribe CONFIRMAR."
    elif not employee_id:
        estado_error = "ID de empleado obligatorio."
    else:
        with engine.begin() as conn:
            emp = get_employee(conn, employee_id)
            if not emp:
                estado_error = "Empleado no encontrado."
            elif emp.get("estado") != "Activo":
                estado_error = f"Empleado no activo: {emp.get('estado')}"
            else:
                turno_final = as_text(turno) or emp.get("turno") or "Día"
                fecha_final = fecha or current_shift_date(turno_final, now_mx(), get_shift_config(conn, turno_final))
                config = get_shift_config(conn, turno_final)
                start_dt, end_dt = schedule_times_from_config(config, fecha_final)
                hora_default = start_dt.time() if accion == "dentro" else end_dt.time()
                mark_time = parse_time_value(hora, hora_default)
                mark_day = date.fromisoformat(fecha_final) + timedelta(days=1) if (accion == "fuera" and bool(config.get("crosses_midnight")) and mark_time <= start_dt.time()) else date.fromisoformat(fecha_final)
                mark_dt = datetime.combine(mark_day, mark_time, tzinfo=TZ)
                ts = now_mx().isoformat()

                if accion == "dentro":
                    open_att = find_open_attendance(conn, employee_id)
                    if open_att:
                        estado_result = {"tipo": "info", "mensaje": f"{emp['nombre']} ya aparece dentro de planta. No se creó otro registro."}
                    else:
                        eval_data = evaluate_entry(config, fecha_final, mark_dt)
                        result = conn.execute(text("""
                            INSERT INTO attendance (
                                employee_id, shift_date, turno, entry_at, entry_guard, entry_status, late_reason,
                                vehicle_expected, vehicle_entered, vehicle_front_entry, vehicle_trunk_entry, incident,
                                scheduled_entry_at, scheduled_exit_at, entry_limit_at, exit_early_limit_at, extra_limit_at,
                                entry_tolerance_minutes, exit_tolerance_minutes, extra_after_minutes, late_minutes,
                                early_minutes, extra_minutes, late_justified, early_justified, extra_authorized, created_at, updated_at
                            ) VALUES (
                                :employee_id, :shift_date, :turno, :entry_at, :entry_guard, :entry_status, :late_reason,
                                :vehicle_expected, :vehicle_entered, '', '', :incident,
                                :scheduled_entry_at, :scheduled_exit_at, :entry_limit_at, :exit_early_limit_at, :extra_limit_at,
                                :entry_tolerance_minutes, :exit_tolerance_minutes, :extra_after_minutes, :late_minutes,
                                0, 0, 0, 0, 0, :created_at, :updated_at
                            ) RETURNING id
                        """), {
                            "employee_id": employee_id,
                            "shift_date": fecha_final,
                            "turno": turno_final,
                            "entry_at": mark_dt.isoformat(),
                            "entry_guard": f"AJUSTE PLANTA / {admin_user['username']}",
                            "entry_status": eval_data["status"],
                            "late_reason": "Ajuste operativo" if eval_data["status"] == "Retardo" else "",
                            "vehicle_expected": 1 if emp.get("tiene_vehiculo") else 0,
                            "vehicle_entered": 0,
                            "incident": "Ajuste operativo: marcado dentro de planta. El siguiente escaneo será salida.",
                            **eval_data,
                            "created_at": ts,
                            "updated_at": ts,
                        })
                        rid = str(result.scalar_one())
                        audit(conn, "attendance", rid, "MARK_INSIDE", admin_user["username"], reason="Ajuste operativo: marcar dentro de planta")
                        estado_result = {"tipo": "ok", "mensaje": f"{emp['nombre']} quedó marcado dentro. El siguiente escaneo será SALIDA.", "registro": rid}

                elif accion == "fuera":
                    open_att = find_open_attendance(conn, employee_id)
                    if not open_att:
                        estado_error = f"{emp['nombre']} no tiene entrada abierta. No se registró salida."
                    else:
                        eval_data = evaluate_exit(config, open_att.get("shift_date") or fecha_final, mark_dt)
                        status = eval_data["status"]
                        lunch_value = "SI"
                        worked_minutes, lunch_value, payroll_note = compute_worked_minutes(open_att.get("entry_at"), mark_dt.isoformat(), eval_data.get("scheduled_entry_at"), eval_data.get("scheduled_exit_at"), lunch_value, status)
                        incident = as_text(open_att.get("incident"))
                        note = "Ajuste operativo: marcado fuera de planta."
                        new_incident = (incident + " | " + note).strip(" |") if incident else note
                        conn.execute(text("""
                            UPDATE attendance
                            SET exit_at=:exit_at,
                                exit_guard=:exit_guard,
                                exit_status=:exit_status,
                                early_reason=:early_reason,
                                scheduled_entry_at=COALESCE(NULLIF(scheduled_entry_at, ''), :scheduled_entry_at),
                                scheduled_exit_at=COALESCE(NULLIF(scheduled_exit_at, ''), :scheduled_exit_at),
                                entry_limit_at=COALESCE(NULLIF(entry_limit_at, ''), :entry_limit_at),
                                exit_early_limit_at=:exit_early_limit_at,
                                extra_limit_at=:extra_limit_at,
                                entry_tolerance_minutes=:entry_tolerance_minutes,
                                exit_tolerance_minutes=:exit_tolerance_minutes,
                                extra_after_minutes=:extra_after_minutes,
                                early_minutes=:early_minutes,
                                extra_minutes=:extra_minutes,
                                lunch_taken=:lunch_taken,
                                worked_minutes=:worked_minutes,
                                incident=:incident,
                                updated_at=:updated_at
                            WHERE id=:id
                        """), {
                            "id": open_att["id"],
                            "exit_at": mark_dt.isoformat(),
                            "exit_guard": f"AJUSTE PLANTA / {admin_user['username']}",
                            "exit_status": status,
                            "early_reason": "",
                            "lunch_taken": lunch_value,
                            "worked_minutes": worked_minutes,
                            "incident": new_incident,
                            "updated_at": ts,
                            **eval_data,
                        })
                        audit(conn, "attendance", str(open_att["id"]), "MARK_OUTSIDE", admin_user["username"], reason="Ajuste operativo: marcar fuera de planta")
                        estado_result = {"tipo": "ok", "mensaje": f"{emp['nombre']} quedó marcado fuera de planta.", "registro": open_att["id"]}
                else:
                    estado_error = "Acción inválida. Usa dentro o fuera."

    with engine.begin() as conn:
        turnos = fetch_all(conn, "SELECT * FROM shift_settings ORDER BY name")
        empleados_activos = conn.execute(text("SELECT COUNT(*) FROM employees WHERE estado = 'Activo'")).scalar() or 0
        registros_abiertos = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE (exit_at IS NULL OR TRIM(COALESCE(exit_at, '')) = '')")).scalar() or 0
        retardos_hoy = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date = :today AND entry_status IN ('Tarde','Retardo')"), {"today": today}).scalar() or 0
    return templates.TemplateResponse(
        "configuracion.html",
        {
            "request": request,
            "turnos": turnos,
            "empleados_activos": empleados_activos,
            "registros_abiertos": registros_abiertos,
            "retardos_hoy": retardos_hoy,
            "today": today,
            "estado_result": estado_result,
            "estado_error": estado_error,
            "estado_employee_id": employee_id,
            "estado_accion": accion,
            "estado_fecha": fecha or today,
            "estado_turno": turno,
            "estado_hora": hora,
        },
    )


@app.get("/turnos", response_class=HTMLResponse)
def turnos_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    with engine.begin() as conn:
        rows = fetch_all(conn, "SELECT * FROM shift_settings ORDER BY name")
    return templates.TemplateResponse("turnos.html", {"request": request, "turnos": rows})


@app.post("/turnos/guardar")
def turnos_guardar(
    request: Request,
    name: str = Form(...),
    entry_time: str = Form(...),
    exit_time: str = Form(...),
    crosses_midnight: str = Form("0"),
    entry_tolerance_minutes: int = Form(10),
    exit_tolerance_minutes: int = Form(10),
    extra_after_minutes: int = Form(30),
    auto_close_enabled: str = Form("1"),
    provisional_close_time: str = Form("02:00"),
    max_open_minutes: int = Form(1080),
    active: str = Form("1"),
):
    admin_user = require_admin_http(request)
    name = as_text(name) or "Día"
    # límites simples para evitar que una pantalla de configuración se convierta en ruleta rusa.
    entry_tolerance_minutes = max(0, min(int(entry_tolerance_minutes), 240))
    exit_tolerance_minutes = max(0, min(int(exit_tolerance_minutes), 240))
    extra_after_minutes = max(0, min(int(extra_after_minutes), 240))
    max_open_minutes = max(0, min(int(max_open_minutes), 2880))
    ts = now_mx().isoformat()
    with engine.begin() as conn:
        existing = fetch_one(conn, "SELECT * FROM shift_settings WHERE name = :name", {"name": name})
        if existing:
            conn.execute(
                text(
                    """
                    UPDATE shift_settings
                    SET entry_time=:entry_time, exit_time=:exit_time, crosses_midnight=:crosses_midnight,
                        entry_tolerance_minutes=:entry_tolerance_minutes,
                        exit_tolerance_minutes=:exit_tolerance_minutes,
                        extra_after_minutes=:extra_after_minutes,
                        auto_close_enabled=:auto_close_enabled,
                        provisional_close_time=:provisional_close_time,
                        max_open_minutes=:max_open_minutes,
                        active=:active, updated_at=:updated_at
                    WHERE name=:name
                    """
                ),
                {
                    "name": name,
                    "entry_time": entry_time,
                    "exit_time": exit_time,
                    "crosses_midnight": 1 if crosses_midnight == "1" else 0,
                    "entry_tolerance_minutes": entry_tolerance_minutes,
                    "exit_tolerance_minutes": exit_tolerance_minutes,
                    "extra_after_minutes": extra_after_minutes,
                    "auto_close_enabled": 1 if auto_close_enabled == "1" else 0,
                    "provisional_close_time": provisional_close_time,
                    "max_open_minutes": max_open_minutes,
                    "active": 1 if active == "1" else 0,
                    "updated_at": ts,
                },
            )
            audit(conn, "shift_settings", name, "UPDATE", admin_user["username"], reason="Configuración de turno")
        else:
            conn.execute(
                text(
                    """
                    INSERT INTO shift_settings (
                        name, entry_time, exit_time, crosses_midnight, work_days,
                        entry_tolerance_minutes, exit_tolerance_minutes, extra_after_minutes,
                        active, auto_close_enabled, provisional_close_time, max_open_minutes, created_at, updated_at
                    ) VALUES (
                        :name, :entry_time, :exit_time, :crosses_midnight, '1,2,3,4,5',
                        :entry_tolerance_minutes, :exit_tolerance_minutes, :extra_after_minutes,
                        :active, :auto_close_enabled, :provisional_close_time, :max_open_minutes, :created_at, :updated_at
                    )
                    """
                ),
                {
                    "name": name,
                    "entry_time": entry_time,
                    "exit_time": exit_time,
                    "crosses_midnight": 1 if crosses_midnight == "1" else 0,
                    "entry_tolerance_minutes": entry_tolerance_minutes,
                    "exit_tolerance_minutes": exit_tolerance_minutes,
                    "extra_after_minutes": extra_after_minutes,
                    "auto_close_enabled": 1 if auto_close_enabled == "1" else 0,
                    "provisional_close_time": provisional_close_time,
                    "max_open_minutes": max_open_minutes,
                    "active": 1 if active == "1" else 0,
                    "created_at": ts,
                    "updated_at": ts,
                },
            )
            audit(conn, "shift_settings", name, "CREATE", admin_user["username"], reason="Alta de turno")
    return RedirectResponse(url="/turnos", status_code=303)


@app.get("/recalcular", response_class=HTMLResponse)
def recalcular_page(request: Request, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None, turno: str = ""):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    today = now_mx().date().isoformat()
    fecha_inicio = fecha_inicio or today
    fecha_fin = fecha_fin or fecha_inicio
    with engine.begin() as conn:
        turnos = fetch_all(conn, "SELECT * FROM shift_settings ORDER BY name")
        params = {"inicio": fecha_inicio, "fin": fecha_fin}
        clause = "shift_date BETWEEN :inicio AND :fin AND COALESCE(anulled,0)=0"
        if turno:
            clause += " AND turno=:turno"
            params["turno"] = turno
        resumen = {
            "registros": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause}"), params).scalar() or 0,
            "retardos": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause} AND entry_status IN ('Tarde','Retardo')"), params).scalar() or 0,
            "extras": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause} AND exit_status='Extra'"), params).scalar() or 0,
            "tempranas": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause} AND exit_status='Salida temprana'"), params).scalar() or 0,
        }
    return templates.TemplateResponse(
        "recalcular.html",
        {"request": request, "fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin, "turno": turno, "turnos": turnos, "resumen": resumen, "result": None},
    )


@app.post("/recalcular")
def recalcular_aplicar(
    request: Request,
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(...),
    turno: str = Form(""),
    confirmar: str = Form(""),
):
    admin_user = require_admin_http(request)
    fecha_inicio = as_text(fecha_inicio)
    fecha_fin = as_text(fecha_fin) or fecha_inicio
    turno = as_text(turno)
    if confirmar.strip().upper() != "RECALCULAR":
        raise HTTPException(status_code=400, detail="Para aplicar debes escribir RECALCULAR")
    try:
        date.fromisoformat(fecha_inicio)
        date.fromisoformat(fecha_fin)
    except Exception:
        raise HTTPException(status_code=400, detail="Fechas inválidas")
    if fecha_fin < fecha_inicio:
        raise HTTPException(status_code=400, detail="La fecha fin no puede ser menor que la fecha inicio")
    with engine.begin() as conn:
        result = recalculate_attendance_records(conn, fecha_inicio, fecha_fin, turno, admin_user["username"])
        turnos = fetch_all(conn, "SELECT * FROM shift_settings ORDER BY name")
        params = {"inicio": fecha_inicio, "fin": fecha_fin}
        clause = "shift_date BETWEEN :inicio AND :fin AND COALESCE(anulled,0)=0"
        if turno:
            clause += " AND turno=:turno"
            params["turno"] = turno
        resumen = {
            "registros": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause}"), params).scalar() or 0,
            "retardos": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause} AND entry_status IN ('Tarde','Retardo')"), params).scalar() or 0,
            "extras": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause} AND exit_status='Extra'"), params).scalar() or 0,
            "tempranas": conn.execute(text(f"SELECT COUNT(*) FROM attendance WHERE {clause} AND exit_status='Salida temprana'"), params).scalar() or 0,
        }
    return templates.TemplateResponse(
        "recalcular.html",
        {"request": request, "fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin, "turno": turno, "turnos": turnos, "resumen": resumen, "result": result},
    )


@app.get("/retardos", response_class=HTMLResponse)
def retardos_page(request: Request, fecha: Optional[str] = None):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    fecha = fecha or now_mx().date().isoformat()
    with engine.begin() as conn:
        late_rows = fetch_all(conn,
            """
            SELECT a.*, e.nombre, e.area, e.puesto
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            WHERE a.shift_date = :fecha AND a.entry_status IN ('Tarde', 'Retardo')
            ORDER BY a.entry_at ASC
            """,
            {"fecha": fecha}
        )
        early_rows = fetch_all(conn,
            """
            SELECT a.*, e.nombre, e.area, e.puesto
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            WHERE a.shift_date = :fecha AND a.exit_status = 'Salida temprana'
            ORDER BY a.exit_at ASC
            """,
            {"fecha": fecha}
        )
        extra_rows = fetch_all(conn,
            """
            SELECT a.*, e.nombre, e.area, e.puesto
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            WHERE a.shift_date = :fecha AND a.exit_status = 'Extra'
            ORDER BY a.exit_at ASC
            """,
            {"fecha": fecha}
        )
        kpis = {
            "retardos": len(late_rows),
            "min_retardo": sum(int(r.get("late_minutes") or 0) for r in late_rows),
            "salidas_tempranas": len(early_rows),
            "min_temprano": sum(int(r.get("early_minutes") or 0) for r in early_rows),
            "extras": len(extra_rows),
            "min_extra": sum(int(r.get("extra_minutes") or 0) for r in extra_rows),
            "sin_motivo": sum(1 for r in late_rows if not (r.get("late_reason") or "").strip()),
            "justificados": sum(1 for r in late_rows if int(r.get("late_justified") or 0) == 1),
            "pendientes": sum(1 for r in late_rows if int(r.get("late_justified") or 0) != 1),
        }
    return templates.TemplateResponse(
        "retardos.html",
        {"request": request, "fecha": fecha, "late_rows": late_rows, "early_rows": early_rows, "extra_rows": extra_rows, "kpis": kpis},
    )


@app.post("/retardos/justificar")
def retardos_justificar(
    request: Request,
    attendance_id: int = Form(...),
    late_reason: str = Form(""),
    motivo_admin: str = Form("Justificación de retardo"),
):
    admin_user = require_admin_http(request)
    reason = as_text(late_reason)
    admin_reason = as_text(motivo_admin) or "Justificación de retardo"
    if not reason:
        raise HTTPException(status_code=400, detail="El motivo del retardo es obligatorio")
    with engine.begin() as conn:
        row = fetch_one(conn, "SELECT * FROM attendance WHERE id = :id", {"id": attendance_id})
        if not row:
            raise HTTPException(status_code=404, detail="Registro no encontrado")
        if row.get("entry_status") not in ("Tarde", "Retardo"):
            raise HTTPException(status_code=400, detail="Este registro no está marcado como retardo")
        old_reason = row.get("late_reason") or ""
        old_justified = str(row.get("late_justified") or 0)
        conn.execute(
            text("""
                UPDATE attendance
                SET late_reason=:late_reason, late_justified=1, updated_at=:updated_at
                WHERE id=:id
            """),
            {"id": attendance_id, "late_reason": reason, "updated_at": now_mx().isoformat()},
        )
        if old_reason != reason:
            audit(conn, "attendance", str(attendance_id), "JUSTIFY_LATE", admin_user["username"], "late_reason", old_reason, reason, admin_reason)
        audit(conn, "attendance", str(attendance_id), "JUSTIFY_LATE", admin_user["username"], "late_justified", old_justified, "1", admin_reason)
    return RedirectResponse(url=f"/retardos?fecha={row['shift_date']}", status_code=303)


@app.get("/usuarios", response_class=HTMLResponse)
def usuarios_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    with engine.begin() as conn:
        rows_raw = fetch_all(conn, "SELECT username, role, active, display_name, created_at, updated_at FROM users ORDER BY username")
    # Seguridad: no se guardan contraseñas reales en texto plano en la base.
    # Solo un Supremo puede ver las claves semilla conocidas del sistema.
    # Para usuarios personalizados, se debe restablecer la clave en lugar de "revelarla".
    is_supremo = (admin_user.get("role") == "Supremo")
    rows = []
    for row in rows_raw:
        item = dict(row)
        item["visible_password"] = DEFAULT_SYSTEM_PASSWORDS.get(item.get("username"), "") if is_supremo else ""
        rows.append(item)
    return templates.TemplateResponse("usuarios.html", {"request": request, "users": rows, "error": "", "is_supremo": is_supremo})


@app.post("/usuarios/guardar")
def usuarios_guardar(
    request: Request,
    username: str = Form(...),
    password: str = Form(""),
    display_name: str = Form(""),
    role: str = Form("Vigilancia"),
    active: str = Form("1"),
):
    admin_user = require_admin_http(request)
    username = clean_username(username)
    role = role if role in {"Supremo", "Admin", "Vigilancia", "RH"} else "Vigilancia"
    # Solo Supremo puede crear, editar o degradar usuarios Supremo. Evita escalamiento interno accidental.
    if role == "Supremo" and admin_user.get("role") != "Supremo":
        raise HTTPException(status_code=403, detail="Solo un usuario Supremo puede asignar el rol Supremo")
    if not username:
        raise HTTPException(status_code=400, detail="Usuario obligatorio")
    ts = now_mx().isoformat()
    with engine.begin() as conn:
        existing = fetch_one(conn, "SELECT * FROM users WHERE username = :username", {"username": username})
        if existing and existing.get("role") == "Supremo" and admin_user.get("role") != "Supremo":
            raise HTTPException(status_code=403, detail="Solo un usuario Supremo puede modificar usuarios Supremo")
        if existing:
            updates = ["role=:role", "active=:active", "display_name=:display_name", "updated_at=:updated_at"]
            params = {"username": username, "role": role, "active": 1 if active == "1" else 0, "display_name": as_text(display_name), "updated_at": ts}
            if password.strip():
                updates.append("password_hash=:password_hash")
                params["password_hash"] = hash_password(password.strip())
            conn.execute(text(f"UPDATE users SET {', '.join(updates)} WHERE username=:username"), params)
            audit(conn, "users", username, "UPDATE", admin_user["username"], reason="Edición de usuario")
        else:
            if not password.strip():
                raise HTTPException(status_code=400, detail="Clave obligatoria para usuario nuevo")
            conn.execute(
                text("""
                    INSERT INTO users (username, password_hash, role, active, display_name, created_at, updated_at)
                    VALUES (:username, :password_hash, :role, :active, :display_name, :created_at, :updated_at)
                """),
                {"username": username, "password_hash": hash_password(password.strip()), "role": role, "active": 1 if active == "1" else 0, "display_name": as_text(display_name), "created_at": ts, "updated_at": ts},
            )
            audit(conn, "users", username, "CREATE", admin_user["username"], reason="Alta de usuario")
    return RedirectResponse(url="/usuarios", status_code=303)


@app.get("/sistema", response_class=HTMLResponse)
def sistema_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    db_ok = True
    db_error = ""
    write_ok = False
    with engine.begin() as conn:
        try:
            conn.execute(text("SELECT 1")).scalar()
            log_event(conn, "info", "sistema", "health_check", "Revisión manual de estado", admin_user["username"])
            write_ok = True
        except Exception as exc:
            db_ok = False
            db_error = str(exc)
        total_open = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE (exit_at IS NULL OR TRIM(COALESCE(exit_at, '')) = '')")).scalar() or 0
        pending_review = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE review_required = 1")).scalar() or 0
        last_attendance = fetch_one(conn, "SELECT * FROM attendance ORDER BY updated_at DESC LIMIT 1")
        recent_events = fetch_all(conn, "SELECT * FROM system_events ORDER BY created_at DESC LIMIT 80")
        recent_errors = fetch_all(conn, "SELECT * FROM system_events WHERE level IN ('error','alerta') ORDER BY created_at DESC LIMIT 20")
    return templates.TemplateResponse("sistema.html", {
        "request": request,
        "db_ok": db_ok,
        "db_error": db_error,
        "write_ok": write_ok,
        "db_name": engine.dialect.name,
        "server_time": now_mx().strftime("%Y-%m-%d %H:%M:%S"),
        "total_open": total_open,
        "pending_review": pending_review,
        "last_attendance": last_attendance,
        "recent_events": recent_events,
        "recent_errors": recent_errors,
    })


def week_bounds(week_start: Optional[str] = None):
    if week_start:
        start = date.fromisoformat(week_start)
    else:
        today = now_mx().date()
        start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    return start, end


def safe_date(value: Optional[str], fallback: date) -> date:
    try:
        return date.fromisoformat(as_text(value))
    except Exception:
        return fallback


def month_bounds(base: Optional[str] = None):
    today = now_mx().date()
    d = safe_date(base, today)
    start = d.replace(day=1)
    if d.month == 12:
        next_month = d.replace(year=d.year + 1, month=1, day=1)
    else:
        next_month = d.replace(month=d.month + 1, day=1)
    end = next_month - timedelta(days=1)
    if start <= today <= end:
        end = today
    return start, end


def year_bounds(base: Optional[str] = None):
    today = now_mx().date()
    d = safe_date(base, today)
    start = date(d.year, 1, 1)
    end = date(d.year, 12, 31)
    if d.year == today.year:
        end = today
    return start, end

SPANISH_MONTHS = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def clamp_int(value, fallback: int, min_value: int, max_value: int) -> int:
    try:
        number = int(value)
    except Exception:
        return fallback
    if number < min_value or number > max_value:
        return fallback
    return number


def month_bounds_by_parts(year_value: Optional[int] = None, month_value: Optional[int] = None):
    today = now_mx().date()
    year = clamp_int(year_value, today.year, 2000, today.year + 5)
    month = clamp_int(month_value, today.month, 1, 12)
    start = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    end = next_month - timedelta(days=1)
    if start <= today <= end:
        end = today
    return start, end, month, year


def dashboard_month_options(selected_year: Optional[int] = None):
    today = now_mx().date()
    year = selected_year or today.year
    return [
        {"value": i + 1, "label": f"{name} {year}"}
        for i, name in enumerate(SPANISH_MONTHS)
    ]


def dashboard_year_options(selected_year: Optional[int] = None):
    today = now_mx().date()
    selected_year = selected_year or today.year
    years = sorted({today.year - 1, today.year, today.year + 1, selected_year})
    return years


def resolve_dashboard_period_range(
    period: str = "day",
    fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
):
    """
    Reglas del dashboard:
    - Semana = semana actual, sin depender de fechas viejas del formulario.
    - Mes = mes actual o el mes elegido en la lista.
    - Rango = exclusivamente fecha inicio a fecha fin.
    """
    today = now_mx().date()
    period = (period or "day").lower()
    selected_month = clamp_int(mes, today.month, 1, 12)
    selected_year = clamp_int(anio, today.year, 2000, today.year + 5)

    if period in {"week", "semana"}:
        start, end = week_bounds(None)
        period = "week"
        label = f"Semana actual {start.isoformat()} a {min(end, today).isoformat()}"
    elif period in {"month", "mes"}:
        start, end, selected_month, selected_year = month_bounds_by_parts(selected_year, selected_month)
        month_name = SPANISH_MONTHS[selected_month - 1]
        period = "month"
        label = f"{month_name} {selected_year}"
    elif period in {"year", "ano", "año"}:
        start, end = year_bounds(str(selected_year) + "-01-01")
        period = "year"
        label = f"Año {selected_year}"
    elif period in {"range", "rango"}:
        start = safe_date(fecha_inicio, today)
        end = safe_date(fecha_fin, start)
        if end < start:
            start, end = end, start
        period = "range"
        label = f"Rango {start.isoformat()} a {min(end, today).isoformat()}"
    else:
        start = safe_date(fecha, today)
        end = start
        period = "day"
        label = f"Día {start.isoformat()}"

    if end > today:
        end = today
    return period, start, end, label, selected_month, selected_year



def resolve_period_range(period: str = "week", fecha: Optional[str] = None, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None, semana: Optional[str] = None):
    today = now_mx().date()
    period = (period or "week").lower()
    if semana and not fecha_inicio and period in {"week", "semana"}:
        fecha_inicio = semana
    if period in {"day", "dia", "día"}:
        start = safe_date(fecha or fecha_inicio, today)
        end = start
        label = f"Día {start.isoformat()}"
    elif period in {"month", "mes"}:
        start, end = month_bounds(fecha or fecha_inicio)
        label = f"Mes {start.strftime('%m/%Y')}"
    elif period in {"year", "ano", "año"}:
        start, end = year_bounds(fecha or fecha_inicio)
        label = f"Año {start.year}"
    elif period in {"range", "rango"}:
        start = safe_date(fecha_inicio, today)
        end = safe_date(fecha_fin, start)
        if end < start:
            start, end = end, start
        label = f"Rango {start.isoformat()} a {end.isoformat()}"
    else:
        start, end = week_bounds(fecha_inicio or fecha or semana)
        label = f"Semana {start.isoformat()} a {end.isoformat()}"
        period = "week"
    # No proyectamos faltas al futuro. Bastante raro es el presente.
    if end > today:
        end = today
    return period, start, end, label


def iter_dates(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def active_employee_rows(conn: Connection, area: str = "", turno: str = ""):
    params = {}
    filters = ["estado='Activo'"]
    if area:
        filters.append("area=:area")
        params["area"] = area
    if turno:
        filters.append("turno=:turno")
        params["turno"] = turno
    return fetch_all(conn, f"SELECT id, nombre, area, puesto, turno FROM employees WHERE {' AND '.join(filters)} ORDER BY nombre", params)


def absence_map(conn: Connection, start: date, end: date, area: str = "", turno: str = ""):
    employees = active_employee_rows(conn, area, turno)
    by_emp = {e["id"]: 0 for e in employees}
    total = 0
    if not employees:
        return by_emp, total
    for d in iter_dates(start, end):
        # Por ahora usamos lunes-viernes como regla operativa general. Los turnos específicos se pueden afinar después sin meter otro monstruo.
        if d.weekday() >= 5:
            continue
        params = {"fecha": d.isoformat()}
        attendance_ids = {r["employee_id"] for r in fetch_all(conn, "SELECT DISTINCT employee_id FROM attendance WHERE shift_date=:fecha AND COALESCE(anulled,0)=0", params)}
        for emp in employees:
            if emp["id"] not in attendance_ids:
                by_emp[emp["id"]] += 1
                total += 1
    return by_emp, total


def period_attendance_stats(conn: Connection, start: date, end: date, area: str = "", turno: str = ""):
    params = {"start": start.isoformat(), "end": end.isoformat()}
    joins = ""
    filters = ["a.shift_date >= :start", "a.shift_date <= :end", "COALESCE(a.anulled,0)=0"]
    if area or turno:
        joins = "JOIN employees e ON e.id=a.employee_id"
    if area:
        filters.append("e.area=:area")
        params["area"] = area
    if turno:
        filters.append("a.turno=:turno")
        params["turno"] = turno
    where = " AND ".join(filters)
    stats = {
        "registros": conn.execute(text(f"SELECT COUNT(*) FROM attendance a {joins} WHERE {where}"), params).scalar() or 0,
        "retardos": conn.execute(text(f"SELECT COUNT(*) FROM attendance a {joins} WHERE {where} AND a.entry_status IN ('Tarde','Retardo')"), params).scalar() or 0,
        "min_retardo": conn.execute(text(f"SELECT COALESCE(SUM(a.late_minutes),0) FROM attendance a {joins} WHERE {where} AND a.entry_status IN ('Tarde','Retardo')"), params).scalar() or 0,
        "tempranas": conn.execute(text(f"SELECT COUNT(*) FROM attendance a {joins} WHERE {where} AND a.exit_status='Salida temprana'"), params).scalar() or 0,
        "min_temprano": conn.execute(text(f"SELECT COALESCE(SUM(a.early_minutes),0) FROM attendance a {joins} WHERE {where} AND a.exit_status='Salida temprana'"), params).scalar() or 0,
        "extras": conn.execute(text(f"SELECT COUNT(*) FROM attendance a {joins} WHERE {where} AND a.exit_status='Extra'"), params).scalar() or 0,
        "min_extra": conn.execute(text(f"SELECT COALESCE(SUM(a.extra_minutes),0) FROM attendance a {joins} WHERE {where} AND a.exit_status='Extra'"), params).scalar() or 0,
    }
    _, faltas = absence_map(conn, start, end, area, turno)
    stats["faltas"] = faltas
    return stats


def employee_period_rows(conn: Connection, start: date, end: date, area: str = "", turno: str = ""):
    params = {"start": start.isoformat(), "end": end.isoformat()}
    filters = ["a.shift_date >= :start", "a.shift_date <= :end", "COALESCE(a.anulled,0)=0"]
    emp_filters = ["e.estado='Activo'"]
    if area:
        emp_filters.append("e.area=:area")
        params["area"] = area
    if turno:
        filters.append("a.turno=:turno")
        emp_filters.append("e.turno=:turno")
        params["turno"] = turno
    where_att = " AND ".join(filters)
    where_emp = " AND ".join(emp_filters)
    rows = fetch_all(conn, f"""
        SELECT e.id AS employee_id, e.nombre, e.area, e.puesto, e.turno AS turno_empleado,
               COUNT(a.id) AS registros,
               SUM(CASE WHEN a.entry_status IN ('Tarde','Retardo') THEN 1 ELSE 0 END) AS retardos,
               SUM(COALESCE(a.late_minutes,0)) AS min_retardo,
               SUM(CASE WHEN a.exit_status = 'Salida temprana' THEN 1 ELSE 0 END) AS salidas_tempranas,
               SUM(COALESCE(a.early_minutes,0)) AS min_temprano,
               SUM(CASE WHEN a.exit_status = 'Extra' THEN 1 ELSE 0 END) AS extras,
               SUM(COALESCE(a.extra_minutes,0)) AS min_extra,
               SUM(CASE WHEN a.provisional_exit = 1 THEN 1 ELSE 0 END) AS salidas_provisionales,
               SUM(CASE WHEN a.review_required = 1 THEN 1 ELSE 0 END) AS pendientes_revision,
               SUM(CASE WHEN (a.exit_at IS NULL OR TRIM(COALESCE(a.exit_at, '')) = '') THEN 1 ELSE 0 END) AS abiertas,
               SUM(CASE WHEN a.anulled = 1 THEN 1 ELSE 0 END) AS anulados
        FROM employees e
        LEFT JOIN attendance a ON a.employee_id=e.id AND {where_att}
        WHERE {where_emp}
        GROUP BY e.id, e.nombre, e.area, e.puesto, e.turno
        ORDER BY min_retardo DESC, retardos DESC, e.nombre
    """, params)
    faltas_by_employee, _ = absence_map(conn, start, end, area, turno)
    for r in rows:
        r["turno"] = r.get("turno_empleado") or ""
        r["faltas"] = faltas_by_employee.get(r["employee_id"], 0)
    return rows


def compute_worked_minutes(entry_at: str, exit_at: str, scheduled_entry_at: str, scheduled_exit_at: str, lunch_taken: str = "", exit_status_value: str = "") -> tuple[int, str, str]:
    entry_dt = parse_dt(entry_at or "")
    exit_dt = parse_dt(exit_at or "")
    scheduled_entry = parse_dt(scheduled_entry_at or "")
    scheduled_exit = parse_dt(scheduled_exit_at or "")
    if not entry_dt or not exit_dt or not scheduled_entry or not scheduled_exit:
        return 0, as_text(lunch_taken).upper() or "", "Faltan fechas para calcular"
    start_dt = max(entry_dt, scheduled_entry)
    end_dt = min(exit_dt, scheduled_exit)
    base = max(0, int((end_dt - start_dt).total_seconds() // 60))
    lunch = as_text(lunch_taken).upper()
    note = ""
    if exit_status_value == "Salida temprana":
        if lunch == "SI":
            base = max(0, base - 60)
        elif lunch == "NO":
            pass
        else:
            note = "Comida no definida en salida temprana"
    else:
        if not lunch:
            lunch = "SI"
        if lunch == "SI":
            base = max(0, base - 60)
    return base, lunch, note


@app.get("/reportes/semanal", response_class=HTMLResponse)
def reporte_semanal_page(
    request: Request,
    semana: Optional[str] = None,
    period: str = "week",
    fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    area: str = "",
    turno: str = "",
):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    period, start, end, period_label = resolve_period_range(period, fecha, fecha_inicio, fecha_fin, semana)
    with engine.begin() as conn:
        rows = employee_period_rows(conn, start, end, area, turno)
        areas = fetch_all(conn, "SELECT DISTINCT area FROM employees WHERE area != '' ORDER BY area")
        turnos = fetch_all(conn, "SELECT name FROM shift_settings WHERE active = 1 ORDER BY name")
    ranking_retardos = [r for r in rows if int(r.get("retardos") or 0) > 0 or int(r.get("min_retardo") or 0) > 0]
    ranking_retardos = sorted(
        ranking_retardos,
        key=lambda r: (int(r.get("min_retardo") or 0), int(r.get("retardos") or 0), str(r.get("nombre") or "")),
        reverse=True,
    )[:10]
    ranking_puntuales = [r for r in rows if int(r.get("registros") or 0) > 0 and int(r.get("retardos") or 0) == 0 and int(r.get("faltas") or 0) == 0]
    ranking_puntuales = sorted(
        ranking_puntuales,
        key=lambda r: (int(r.get("registros") or 0), -int(r.get("salidas_tempranas") or 0), str(r.get("nombre") or "")),
        reverse=True,
    )[:10]
    totals = {
        "retardos": sum(int(r.get("retardos") or 0) for r in rows),
        "min_retardo": sum(int(r.get("min_retardo") or 0) for r in rows),
        "faltas": sum(int(r.get("faltas") or 0) for r in rows),
        "tempranas": sum(int(r.get("salidas_tempranas") or 0) for r in rows),
        "min_temprano": sum(int(r.get("min_temprano") or 0) for r in rows),
        "extras": sum(int(r.get("extras") or 0) for r in rows),
        "min_extra": sum(int(r.get("min_extra") or 0) for r in rows),
        "pendientes": sum(int(r.get("pendientes_revision") or 0) for r in rows),
    }
    qs = f"period={period}&fecha_inicio={start.isoformat()}&fecha_fin={end.isoformat()}&area={quote(area)}&turno={quote(turno)}"
    return templates.TemplateResponse("reporte_semanal.html", {
        "request": request,
        "rows": rows,
        "ranking_retardos": ranking_retardos,
        "ranking_puntuales": ranking_puntuales,
        "start": start,
        "end": end,
        "period": period,
        "period_label": period_label,
        "areas": areas,
        "turnos": turnos,
        "area": area,
        "turno": turno,
        "totals": totals,
        "excel_query": qs,
    })


@app.get("/reportes/semanal.xlsx")
def reporte_semanal_excel(
    request: Request,
    semana: Optional[str] = None,
    period: str = "week",
    fecha: Optional[str] = None,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    area: str = "",
    turno: str = "",
):
    require_rh_http(request)
    period, start, end, period_label = resolve_period_range(period, fecha, fecha_inicio, fecha_fin, semana)
    with engine.begin() as conn:
        rows = employee_period_rows(conn, start, end, area, turno)
    wb = Workbook()
    ws = wb.active
    ws.title = "Acumulado"
    headers = ["ID", "Empleado", "Área", "Turno", "Registros", "Faltas", "Retardos", "Min retardo", "Salidas tempranas", "Min temprano", "Extras", "Min extra", "Salidas provisionales", "Pendientes revisión", "Abiertas", "Anulados"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get("employee_id"), r.get("nombre"), r.get("area"), r.get("turno"), r.get("registros") or 0, r.get("faltas") or 0, r.get("retardos") or 0, r.get("min_retardo") or 0, r.get("salidas_tempranas") or 0, r.get("min_temprano") or 0, r.get("extras") or 0, r.get("min_extra") or 0, r.get("salidas_provisionales") or 0, r.get("pendientes_revision") or 0, r.get("abiertas") or 0, r.get("anulados") or 0])
    style_worksheet(ws, f"Reporte acumulado {period_label}")

    ranking = [r for r in rows if int(r.get("retardos") or 0) > 0 or int(r.get("min_retardo") or 0) > 0]
    ranking = sorted(
        ranking,
        key=lambda r: (int(r.get("min_retardo") or 0), int(r.get("retardos") or 0), str(r.get("nombre") or "")),
        reverse=True,
    )
    ws_rank = wb.create_sheet("Ranking retardos")
    ws_rank.append(["Lugar", "ID", "Empleado", "Área", "Turno", "Retardos", "Min retardo", "Promedio min/retardo", "Registros", "Faltas"])
    for idx, r in enumerate(ranking, start=1):
        retardos = int(r.get("retardos") or 0)
        min_retardo = int(r.get("min_retardo") or 0)
        promedio = round(min_retardo / retardos, 2) if retardos else 0
        ws_rank.append([idx, r.get("employee_id"), r.get("nombre"), r.get("area"), r.get("turno"), retardos, min_retardo, promedio, r.get("registros") or 0, r.get("faltas") or 0])
    style_worksheet(ws_rank, f"Ranking de retardos {start.isoformat()} a {end.isoformat()}")

    punctual = [r for r in rows if int(r.get("registros") or 0) > 0 and int(r.get("retardos") or 0) == 0 and int(r.get("faltas") or 0) == 0]
    punctual = sorted(punctual, key=lambda r: (int(r.get("registros") or 0), -int(r.get("salidas_tempranas") or 0), str(r.get("nombre") or "")), reverse=True)
    ws_p = wb.create_sheet("Ranking puntuales")
    ws_p.append(["Lugar", "ID", "Empleado", "Área", "Turno", "Registros", "Retardos", "Faltas", "Salidas tempranas"])
    for idx, r in enumerate(punctual, start=1):
        ws_p.append([idx, r.get("employee_id"), r.get("nombre"), r.get("area"), r.get("turno"), r.get("registros") or 0, r.get("retardos") or 0, r.get("faltas") or 0, r.get("salidas_tempranas") or 0])
    style_worksheet(ws_p, f"Ranking de puntualidad {start.isoformat()} a {end.isoformat()}")

    return workbook_response(wb, f"reporte_acumulado_{start.isoformat()}_{end.isoformat()}.xlsx")


@app.get("/rh", response_class=HTMLResponse)
def rh_dashboard(request: Request, fecha: Optional[str] = None):
    rh_user = require_rh_page(request)
    if not rh_user:
        return admin_login_redirect(request)
    fecha = fecha or now_mx().date().isoformat()
    start, end = week_bounds(None)
    with engine.begin() as conn:
        auto_close_overdue_records(conn, rh_user["username"])
        summary = {
            "presentes": conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date=:fecha AND entry_at IS NOT NULL"), {"fecha": fecha}).scalar() or 0,
            "faltas": conn.execute(text("""
                SELECT COUNT(*) FROM employees e
                WHERE e.estado='Activo' AND NOT EXISTS (
                  SELECT 1 FROM attendance a WHERE a.employee_id=e.id AND a.shift_date=:fecha
                )
            """), {"fecha": fecha}).scalar() or 0,
            "retardos": conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date=:fecha AND entry_status IN ('Tarde','Retardo')"), {"fecha": fecha}).scalar() or 0,
            "min_retardo": conn.execute(text("SELECT COALESCE(SUM(late_minutes),0) FROM attendance WHERE shift_date=:fecha AND entry_status IN ('Tarde','Retardo')"), {"fecha": fecha}).scalar() or 0,
            "salidas_tempranas": conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date=:fecha AND exit_status='Salida temprana'"), {"fecha": fecha}).scalar() or 0,
            "extras": conn.execute(text("SELECT COUNT(*) FROM attendance WHERE shift_date=:fecha AND exit_status='Extra'"), {"fecha": fecha}).scalar() or 0,
            "pendientes_revision": conn.execute(text("SELECT COUNT(*) FROM attendance WHERE review_required=1"), {}).scalar() or 0,
        }
        daily_rows = fetch_all(conn, """
            SELECT a.*, e.nombre, e.area, e.puesto
            FROM attendance a
            JOIN employees e ON e.id = a.employee_id
            WHERE a.shift_date=:fecha
            ORDER BY COALESCE(a.entry_at, a.created_at) ASC
        """, {"fecha": fecha})
        absent_rows = fetch_all(conn, """
            SELECT e.id, e.nombre, e.area, e.puesto, e.turno
            FROM employees e
            WHERE e.estado='Activo' AND NOT EXISTS (
              SELECT 1 FROM attendance a WHERE a.employee_id=e.id AND a.shift_date=:fecha
            )
            ORDER BY e.area, e.nombre
        """, {"fecha": fecha})
        weekly_rows = fetch_all(conn, """
            SELECT e.id AS employee_id, e.nombre, e.area, COALESCE(a.turno, e.turno) AS turno,
                   COUNT(a.id) AS registros,
                   SUM(CASE WHEN a.entry_status IN ('Tarde','Retardo') THEN 1 ELSE 0 END) AS retardos,
                   SUM(COALESCE(a.late_minutes,0)) AS min_retardo,
                   SUM(CASE WHEN a.exit_status = 'Salida temprana' THEN 1 ELSE 0 END) AS salidas_tempranas,
                   SUM(CASE WHEN a.exit_status = 'Extra' THEN 1 ELSE 0 END) AS extras,
                   SUM(COALESCE(a.extra_minutes,0)) AS min_extra,
                   SUM(CASE WHEN a.review_required = 1 THEN 1 ELSE 0 END) AS pendientes_revision
            FROM employees e
            LEFT JOIN attendance a ON a.employee_id = e.id AND a.shift_date >= :start AND a.shift_date <= :end
            WHERE e.estado = 'Activo'
            GROUP BY e.id, e.nombre, e.area, COALESCE(a.turno, e.turno)
            ORDER BY min_retardo DESC, retardos DESC, e.nombre
            LIMIT 80
        """, {"start": start.isoformat(), "end": end.isoformat()})
    return templates.TemplateResponse("rh.html", {
        "request": request, "fecha": fecha, "summary": summary, "daily_rows": daily_rows, "absent_rows": absent_rows,
        "weekly_rows": weekly_rows, "start": start, "end": end,
    })


@app.get("/supremo/registros", response_class=HTMLResponse)
def supremo_registros_page(request: Request):
    supremo = require_supremo_page(request)
    if not supremo:
        return admin_login_redirect(request)
    today = now_mx().date().isoformat()
    with engine.begin() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM attendance")).scalar() or 0
        recent = fetch_all(conn, """
            SELECT a.*, e.nombre
            FROM attendance a
            LEFT JOIN employees e ON e.id = a.employee_id
            ORDER BY a.updated_at DESC
            LIMIT 40
        """)
    return templates.TemplateResponse("supremo_registros.html", {"request": request, "total": total, "recent": recent, "today": today, "message": ""})


@app.post("/supremo/registros/borrar", response_class=HTMLResponse)
def supremo_registros_borrar(
    request: Request,
    mode: str = Form(...),
    attendance_id: str = Form(""),
    fecha_inicio: str = Form(""),
    fecha_fin: str = Form(""),
    confirmacion: str = Form(""),
):
    supremo = require_supremo_http(request)
    deleted = 0
    message = ""
    if as_text(confirmacion).upper() != "BORRAR":
        raise HTTPException(status_code=400, detail="Para borrar registros debes escribir BORRAR en confirmación")
    with engine.begin() as conn:
        if mode == "id":
            rid = int(attendance_id)
            row = fetch_one(conn, "SELECT * FROM attendance WHERE id=:id", {"id": rid})
            if row:
                audit(conn, "attendance", str(rid), "DELETE_BETA", supremo["username"], reason="Borrado beta por usuario Supremo")
                conn.execute(text("DELETE FROM attendance WHERE id=:id"), {"id": rid})
                deleted = 1
        elif mode == "rango":
            if not fecha_inicio or not fecha_fin:
                raise HTTPException(status_code=400, detail="Fecha inicio y fin son obligatorias")
            ids = fetch_all(conn, "SELECT id FROM attendance WHERE shift_date >= :fi AND shift_date <= :ff", {"fi": fecha_inicio, "ff": fecha_fin})
            for row in ids:
                audit(conn, "attendance", str(row["id"]), "DELETE_BETA", supremo["username"], reason=f"Borrado beta por rango {fecha_inicio} a {fecha_fin}")
            conn.execute(text("DELETE FROM attendance WHERE shift_date >= :fi AND shift_date <= :ff"), {"fi": fecha_inicio, "ff": fecha_fin})
            deleted = len(ids)
        elif mode == "todos":
            ids = fetch_all(conn, "SELECT id FROM attendance")
            for row in ids:
                audit(conn, "attendance", str(row["id"]), "DELETE_BETA", supremo["username"], reason="Borrado beta total de registros de asistencia")
            conn.execute(text("DELETE FROM attendance"))
            deleted = len(ids)
        else:
            raise HTTPException(status_code=400, detail="Modo inválido")
        log_event(conn, "alerta", "supremo", "delete_beta", f"Borrado beta de {deleted} registros de asistencia", supremo["username"], mode)
        total = conn.execute(text("SELECT COUNT(*) FROM attendance")).scalar() or 0
        recent = fetch_all(conn, """
            SELECT a.*, e.nombre
            FROM attendance a
            LEFT JOIN employees e ON e.id = a.employee_id
            ORDER BY a.updated_at DESC
            LIMIT 40
        """)
        message = f"Se borraron {deleted} registros de asistencia. Auditoría conservada."
    return templates.TemplateResponse("supremo_registros.html", {"request": request, "total": total, "recent": recent, "today": now_mx().date().isoformat(), "message": message})


@app.get("/personal", response_class=HTMLResponse)
def empleados(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    with engine.begin() as conn:
        rows = fetch_all(conn, "SELECT * FROM employees ORDER BY nombre")
    return templates.TemplateResponse("personal.html", {"request": request, "employees": rows})


@app.get("/personal/nuevo", response_class=HTMLResponse)
def empleado_nuevo(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    
    with engine.begin() as conn:
        turnos = fetch_all(conn, "SELECT name FROM shift_settings WHERE active = 1 ORDER BY name")
        next_id = next_numeric_employee_id(conn)
    return templates.TemplateResponse("empleado_form.html", {"request": request, "employee": None, "turnos": turnos, "next_id": next_id})


@app.get("/personal/{employee_id}", response_class=HTMLResponse)
def empleado_editar(request: Request, employee_id: str):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    admin_user = require_admin_http(request)
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
    
    with engine.begin() as conn:
        turnos = fetch_all(conn, "SELECT name FROM shift_settings WHERE active = 1 ORDER BY name")
    return templates.TemplateResponse("empleado_form.html", {"request": request, "employee": emp, "turnos": turnos})


@app.post("/personal/guardar")
async def empleado_guardar(
    request: Request,
    id: str = Form(...),
    nombre: str = Form(...),
    area: str = Form(""),
    puesto: str = Form(""),
    turno: str = Form("Día"),
    estado: str = Form("Activo"),
    tiene_vehiculo: str = Form("0"),
    requiere_fotos_vehiculo: str = Form("0"),
    qr_activo: str = Form("1"),
    observaciones: str = Form(""),
    foto: Optional[UploadFile] = File(None),
):
    admin_user = require_admin_http(request)
    requested_id = clean_employee_id(id)
    if id and not requested_id:
        raise HTTPException(status_code=400, detail="El ID de empleado debe ser únicamente numérico")
    if not nombre.strip():
        raise HTTPException(status_code=400, detail="El nombre es obligatorio")
    ts = now_mx().isoformat()
    foto_path = ""
    with engine.begin() as conn:
        employee_id = requested_id or next_numeric_employee_id(conn)
        exists = get_employee(conn, employee_id)
        # Si la pantalla quedó abierta y alguien creó ese ID mientras tanto, avanzamos al siguiente disponible.
        if exists:
            employee_id = next_numeric_employee_id(conn)
            exists = get_employee(conn, employee_id)
        if exists:
            raise HTTPException(status_code=400, detail="No se pudo asignar ID automático. Intenta de nuevo.")
        conn.execute(
            text(
                """
                INSERT INTO employees (id, nombre, area, puesto, turno, estado, tiene_vehiculo, requiere_fotos_vehiculo, foto_path, qr_activo, observaciones, created_at, updated_at)
                VALUES (:id, :nombre, :area, :puesto, :turno, :estado, :tiene_vehiculo, :requiere_fotos_vehiculo, :foto_path, :qr_activo, :observaciones, :created_at, :updated_at)
                """
            ),
            {
                "id": employee_id,
                "nombre": nombre.strip(),
                "area": area.strip(),
                "puesto": puesto.strip(),
                "turno": turno,
                "estado": estado,
                "tiene_vehiculo": 1 if tiene_vehiculo == "1" else 0,
                "requiere_fotos_vehiculo": 1 if requiere_fotos_vehiculo == "1" else 0,
                "foto_path": foto_path,
                "qr_activo": 1 if qr_activo == "1" else 0,
                "observaciones": observaciones.strip(),
                "created_at": ts,
                "updated_at": ts,
            }
        )
        audit(conn, "employees", employee_id, "CREATE", admin_user["username"])
    return RedirectResponse(url="/personal", status_code=303)


@app.post("/personal/{employee_id}/guardar")
async def empleado_actualizar(
    request: Request,
    employee_id: str,
    nombre: str = Form(...),
    area: str = Form(""),
    puesto: str = Form(""),
    turno: str = Form("Día"),
    estado: str = Form("Activo"),
    tiene_vehiculo: str = Form("0"),
    requiere_fotos_vehiculo: str = Form("0"),
    qr_activo: str = Form("1"),
    observaciones: str = Form(""),
    foto: Optional[UploadFile] = File(None),
):
    admin_user = require_admin_http(request)
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        foto_path = emp["foto_path"] or ""
        # Versión sin imágenes: no se cargan ni reemplazan fotos desde expediente.
        conn.execute(
            text(
                """
                UPDATE employees
                SET nombre=:nombre, area=:area, puesto=:puesto, turno=:turno, estado=:estado,
                    tiene_vehiculo=:tiene_vehiculo, requiere_fotos_vehiculo=:requiere_fotos_vehiculo,
                    foto_path=:foto_path, qr_activo=:qr_activo, observaciones=:observaciones, updated_at=:updated_at
                WHERE id=:id
                """
            ),
            {
                "nombre": nombre.strip(),
                "area": area.strip(),
                "puesto": puesto.strip(),
                "turno": turno,
                "estado": estado,
                "tiene_vehiculo": 1 if tiene_vehiculo == "1" else 0,
                "requiere_fotos_vehiculo": 1 if requiere_fotos_vehiculo == "1" else 0,
                "foto_path": foto_path,
                "qr_activo": 1 if qr_activo == "1" else 0,
                "observaciones": observaciones.strip(),
                "updated_at": now_mx().isoformat(),
                "id": employee_id,
            }
        )
        audit(conn, "employees", employee_id, "UPDATE", admin_user["username"], reason="Edición manual de expediente")
    return RedirectResponse(url="/personal", status_code=303)


@app.post("/personal/{employee_id}/inactivar")
def empleado_inactivar(request: Request, employee_id: str):
    admin_user = require_admin_http(request)
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        conn.execute(text("UPDATE employees SET estado='Baja', qr_activo=0, updated_at=:updated_at WHERE id=:id"), {"updated_at": now_mx().isoformat(), "id": employee_id})
        audit(conn, "employees", employee_id, "UPDATE", admin_user["username"], "estado", emp.get("estado", ""), "Baja", "Baja lógica de empleado")
    return RedirectResponse(url="/personal", status_code=303)


@app.post("/personal/{employee_id}/reactivar")
def empleado_reactivar(request: Request, employee_id: str):
    admin_user = require_admin_http(request)
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        conn.execute(text("UPDATE employees SET estado='Activo', qr_activo=1, updated_at=:updated_at WHERE id=:id"), {"updated_at": now_mx().isoformat(), "id": employee_id})
        audit(conn, "employees", employee_id, "UPDATE", admin_user["username"], "estado", emp.get("estado", ""), "Activo", "Reactivación de empleado")
    return RedirectResponse(url="/personal", status_code=303)


@app.post("/personal/{employee_id}/borrar")
def empleado_borrar(request: Request, employee_id: str, confirmacion: str = Form("")):
    supremo = require_supremo_http(request)
    employee_id = clean_id(employee_id)
    if as_text(confirmacion).upper() != "BORRAR":
        raise HTTPException(status_code=400, detail="Para borrar definitivamente escribe BORRAR")
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        count = conn.execute(text("SELECT COUNT(*) FROM attendance WHERE employee_id=:id"), {"id": employee_id}).scalar() or 0
        # En beta Supremo puede borrar empleado y sus asistencias. En producción se recomienda baja lógica.
        conn.execute(text("DELETE FROM attendance WHERE employee_id=:id"), {"id": employee_id})
        conn.execute(text("DELETE FROM employees WHERE id=:id"), {"id": employee_id})
        audit(conn, "employees", employee_id, "DELETE", supremo["username"], reason=f"Borrado definitivo beta. Asistencias eliminadas: {count}")
        log_event(conn, "alert", "personal", "delete_employee", f"Empleado borrado: {employee_id}; asistencias: {count}", supremo["username"])
    return RedirectResponse(url="/personal", status_code=303)


@app.get("/importar", response_class=HTMLResponse)
def importar_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    return templates.TemplateResponse("importar.html", {"request": request, "result": None})


@app.get("/plantilla-empleados.xlsx")
def plantilla_empleados(request: Request):
    require_admin_http(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "empleados"
    headers = [
        "ID empleado (opcional)",
        "Nombre completo",
        "Area",
        "Puesto",
        "Turno",
        "Estado",
        "Tiene vehiculo",
        "Observaciones",
    ]
    ws.append(headers)
    ws.append(["", "Nombre Apellido", "Producción", "Operador", "Día", "Activo", "Sí", "ID vacío = automático"])
    ws.append(["", "Nombre Apellido", "Calidad", "Inspectora", "Noche", "Activo", "No", "ID vacío = automático"])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 24
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla-empleados.xlsx"},
    )


@app.post("/importar", response_class=HTMLResponse)
async def importar_empleados(request: Request, archivo: UploadFile = File(...), actualizar: Optional[str] = Form(None)):
    admin_user = require_admin_http(request)
    content = await archivo.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        return templates.TemplateResponse("importar.html", {"request": request, "result": {"ok": False, "errors": [f"No se pudo leer el Excel: {exc}"]}})

    ws = wb.active

    aliases = {
        "id": ["id empleado", "id empleado (opcional)", "id", "empleado", "codigo", "código"],
        "nombre": ["nombre completo", "nombre"],
        "area": ["area", "área"],
        "puesto": ["puesto"],
        "turno": ["turno"],
        "estado": ["estado"],
        "tiene_vehiculo": ["tiene vehiculo", "tiene vehículo", "vehiculo", "vehículo"],
        "requiere_fotos_vehiculo": ["requiere fotos vehiculo", "requiere fotos vehículo", "requiere foto vehiculo", "requiere foto vehículo"],
        "observaciones": ["observaciones", "obs"],
    }

    required = ["nombre", "turno", "estado"]
    header_row, idx_import = detect_import_header(ws, aliases, required)

    def index_for(key: str) -> Optional[int]:
        return idx_import.get(key)

    missing = [key for key in required if index_for(key) is None]
    if missing:
        return templates.TemplateResponse("importar.html", {"request": request, "result": {"ok": False, "errors": ["Faltan columnas requeridas: " + ", ".join(missing)]}})

    errors = []
    rows_to_import = []
    seen = set()
    with engine.begin() as conn:
        allowed_turnos = {r["name"] for r in fetch_all(conn, "SELECT name FROM shift_settings WHERE active = 1")}
        base_used = {as_text(r["id"]) for r in fetch_all(conn, "SELECT id FROM employees")}
        next_candidate = int(next_numeric_employee_id(conn))
    allowed_turnos.update({"Día", "Dia", "Noche"})

    id_idx = index_for("id")
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(row):
            continue
        raw_employee_id = as_text(row[id_idx]) if id_idx is not None else ""
        if raw_employee_id:
            employee_id = clean_employee_id(raw_employee_id)
            if not employee_id:
                errors.append(f"Fila {row_number}: el ID debe ser únicamente numérico: {raw_employee_id}")
        else:
            while str(next_candidate) in base_used or str(next_candidate) in seen:
                next_candidate += 1
            employee_id = str(next_candidate)
            next_candidate += 1
        nombre = as_text(row[index_for("nombre")])
        turno = as_text(row[index_for("turno")]) or "Día"
        estado = as_text(row[index_for("estado")]) or "Activo"
        area = as_text(row[index_for("area")]) if index_for("area") is not None else ""
        puesto = as_text(row[index_for("puesto")]) if index_for("puesto") is not None else ""
        observaciones = as_text(row[index_for("observaciones")]) if index_for("observaciones") is not None else ""
        tiene_vehiculo = bool_from_excel(row[index_for("tiene_vehiculo")]) if index_for("tiene_vehiculo") is not None else 0
        requiere_fotos = 0

        if not nombre:
            errors.append(f"Fila {row_number}: falta nombre")
        if turno not in allowed_turnos:
            errors.append(f"Fila {row_number}: turno inválido: {turno}. Debe existir en Turnos.")
        if estado not in {"Activo", "Inactivo", "Baja", "Suspendido"}:
            errors.append(f"Fila {row_number}: estado inválido: {estado}")
        if employee_id in seen:
            errors.append(f"Fila {row_number}: ID duplicado dentro del Excel: {employee_id}")
        seen.add(employee_id)
        if turno == "Dia":
            turno = "Día"
        rows_to_import.append((employee_id, nombre, area, puesto, turno, estado, tiene_vehiculo, requiere_fotos, observaciones))

    with engine.begin() as conn:
        for employee_id, *_ in rows_to_import:
            exists = get_employee(conn, employee_id)
            if exists and not actualizar:
                errors.append(f"ID {employee_id}: ya existe. Marca 'actualizar existentes' si quieres modificarlo.")

    if errors:
        return templates.TemplateResponse("importar.html", {"request": request, "result": {"ok": False, "errors": errors[:200], "total_errors": len(errors)}})

    nuevos = 0
    actualizados = 0
    ts = now_mx().isoformat()
    with engine.begin() as conn:
        for employee_id, nombre, area, puesto, turno, estado, tiene_vehiculo, requiere_fotos, observaciones in rows_to_import:
            exists = get_employee(conn, employee_id)
            if exists:
                conn.execute(
                    text(
                        """
                        UPDATE employees
                        SET nombre=:nombre, area=:area, puesto=:puesto, turno=:turno, estado=:estado,
                            tiene_vehiculo=:tiene_vehiculo, requiere_fotos_vehiculo=:requiere_fotos_vehiculo,
                            observaciones=:observaciones, updated_at=:updated_at
                        WHERE id=:id
                        """
                    ),
                    {
                        "id": employee_id,
                        "nombre": nombre,
                        "area": area,
                        "puesto": puesto,
                        "turno": turno,
                        "estado": estado,
                        "tiene_vehiculo": tiene_vehiculo,
                        "requiere_fotos_vehiculo": requiere_fotos,
                        "observaciones": observaciones,
                        "updated_at": ts,
                    }
                )
                audit(conn, "employees", employee_id, "UPDATE", admin_user["username"], reason="Importación Excel")
                actualizados += 1
            else:
                conn.execute(
                    text(
                        """
                        INSERT INTO employees (id, nombre, area, puesto, turno, estado, tiene_vehiculo, requiere_fotos_vehiculo, foto_path, qr_activo, observaciones, created_at, updated_at)
                        VALUES (:id, :nombre, :area, :puesto, :turno, :estado, :tiene_vehiculo, :requiere_fotos_vehiculo, '', 1, :observaciones, :created_at, :updated_at)
                        """
                    ),
                    {
                        "id": employee_id,
                        "nombre": nombre,
                        "area": area,
                        "puesto": puesto,
                        "turno": turno,
                        "estado": estado,
                        "tiene_vehiculo": tiene_vehiculo,
                        "requiere_fotos_vehiculo": requiere_fotos,
                        "observaciones": observaciones,
                        "created_at": ts,
                        "updated_at": ts,
                    }
                )
                audit(conn, "employees", employee_id, "CREATE", admin_user["username"], reason="Importación Excel")
                nuevos += 1

    return templates.TemplateResponse("importar.html", {"request": request, "result": {"ok": True, "leidos": len(rows_to_import), "nuevos": nuevos, "actualizados": actualizados}})


@app.get("/plantilla-asistencias.xlsx")
def plantilla_asistencias(request: Request):
    require_admin_http(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "asistencias"
    headers = [
        "ID registro", "ID empleado", "Fecha turno", "Turno", "Entrada", "Salida",
        "Guardia entrada", "Guardia salida", "Motivo retardo", "Motivo salida temprana",
        "Vehiculo", "Observaciones",
    ]
    ws.append(headers)
    ws.append(["", "126", now_mx().date().isoformat(), "Día", "08:00", "19:00", "IMPORTACION_HISTORICA", "IMPORTACION_HISTORICA", "", "", "No", "Ejemplo histórico: motivos pueden ir vacíos"] )
    ws.append(["", "127", now_mx().date().isoformat(), "Noche", "19:00", "08:00", "IMPORTACION_HISTORICA", "IMPORTACION_HISTORICA", "", "", "No", "Salida del día siguiente si solo pones hora"] )
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    style_worksheet(ws, "Plantilla de asistencia histórica")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla-asistencias.xlsx"'},
    )


def _idx_from_aliases(normalized: dict, aliases: list[str]) -> Optional[int]:
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None


def detect_import_header(ws, aliases: dict, required_keys: list[str], max_scan: int = 10):
    """Detecta la fila real de encabezados en plantillas con título arriba.
    Devuelve (header_row_number, idx_map). Así no nos tropezamos con archivos bonitos,
    esa rareza humana de poner un título antes de las columnas.
    """
    best_score = -1
    best_row = 1
    best_idx = {k: None for k in aliases}
    last_row = min(ws.max_row or 1, max_scan)
    for row_number in range(1, last_row + 1):
        raw_headers = [as_text(cell.value).lower().strip() for cell in ws[row_number]]
        normalized = {h: idx for idx, h in enumerate(raw_headers) if h}
        idx_map = {k: _idx_from_aliases(normalized, v) for k, v in aliases.items()}
        score = sum(1 for v in idx_map.values() if v is not None)
        if score > best_score:
            best_score = score
            best_row = row_number
            best_idx = idx_map
        if all(idx_map.get(key) is not None for key in required_keys):
            return row_number, idx_map
    return best_row, best_idx


def attendance_row_values(row, index):
    return row[index] if index is not None and index < len(row) else None


@app.post("/importar/asistencias", response_class=HTMLResponse)
async def importar_asistencias_historicas(request: Request, archivo_asistencias: UploadFile = File(...), actualizar_asistencias: Optional[str] = Form(None)):
    admin_user = require_admin_http(request)
    content = await archivo_asistencias.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        return templates.TemplateResponse("importar.html", {"request": request, "result_asistencias": {"ok": False, "errors": [f"No se pudo leer el Excel: {exc}"]}, "result": None})

    ws = wb.active
    aliases = {
        "id_registro": ["id registro", "id_registro", "registro", "id asistencia"],
        "employee_id": ["id empleado", "id_empleado", "empleado", "id", "codigo", "código"],
        "shift_date": ["fecha turno", "fecha_turno", "fecha", "dia", "día"],
        "turno": ["turno"],
        "entrada": ["entrada", "hora entrada", "hora_entrada", "entry_at"],
        "salida": ["salida", "hora salida", "hora_salida", "exit_at"],
        "guardia_entrada": ["guardia entrada", "guardia_entrada"],
        "guardia_salida": ["guardia salida", "guardia_salida"],
        "motivo_retardo": ["motivo retardo", "motivo_retardo", "late_reason"],
        "motivo_salida": ["motivo salida temprana", "motivo_salida_temprana", "early_reason"],
        "vehiculo": ["vehiculo", "vehículo", "vehicle"],
        "observaciones": ["observaciones", "obs", "incidencia"],
    }
    header_row, idx = detect_import_header(ws, aliases, ["employee_id", "shift_date", "entrada"])
    missing = [name for name in ["employee_id", "shift_date", "entrada"] if idx[name] is None]
    if missing:
        return templates.TemplateResponse("importar.html", {"request": request, "result_asistencias": {"ok": False, "errors": ["Faltan columnas requeridas: " + ", ".join(missing) + ". Revisa que no se haya borrado la fila de encabezados."]}, "result": None})

    errors = []
    rows_to_apply = []
    seen_keys = set()

    with engine.begin() as conn:
        allowed_turnos = {r["name"] for r in fetch_all(conn, "SELECT name FROM shift_settings WHERE active = 1")}

    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(row):
            continue
        raw_id_reg = attendance_row_values(row, idx["id_registro"])
        if isinstance(raw_id_reg, (int, float)) and float(raw_id_reg).is_integer():
            id_registro = str(int(raw_id_reg))
        else:
            id_registro = as_text(raw_id_reg)
        employee_id = clean_employee_id(as_text(attendance_row_values(row, idx["employee_id"])))
        shift_day = parse_excel_date_value(attendance_row_values(row, idx["shift_date"]))
        turno = as_text(attendance_row_values(row, idx["turno"])) if idx["turno"] is not None else ""
        if turno == "Dia":
            turno = "Día"
        guardia_entrada = as_text(attendance_row_values(row, idx["guardia_entrada"])) if idx["guardia_entrada"] is not None else ""
        guardia_salida = as_text(attendance_row_values(row, idx["guardia_salida"])) if idx["guardia_salida"] is not None else ""
        guardia_entrada = guardia_entrada or "IMPORTACION_HISTORICA"
        guardia_salida = guardia_salida or "IMPORTACION_HISTORICA"
        motivo_retardo = as_text(attendance_row_values(row, idx["motivo_retardo"])) if idx["motivo_retardo"] is not None else ""
        motivo_salida = as_text(attendance_row_values(row, idx["motivo_salida"])) if idx["motivo_salida"] is not None else ""
        vehiculo = bool_from_excel(attendance_row_values(row, idx["vehiculo"])) if idx["vehiculo"] is not None else 0
        observaciones = as_text(attendance_row_values(row, idx["observaciones"])) if idx["observaciones"] is not None else ""

        if not employee_id:
            errors.append(f"Fila {row_number}: falta ID empleado")
            continue
        if not shift_day:
            errors.append(f"Fila {row_number}: fecha turno inválida")
            continue

        with engine.begin() as conn:
            emp = get_employee(conn, employee_id)
            if not emp:
                errors.append(f"Fila {row_number}: empleado no existe: {employee_id}")
                continue
            if not turno:
                turno = emp.get("turno") or "Día"
            if turno not in allowed_turnos and turno not in {"Día", "Noche"}:
                errors.append(f"Fila {row_number}: turno inválido: {turno}")
                continue
            config = get_shift_config(conn, turno)

        entry_dt = parse_excel_datetime_value(attendance_row_values(row, idx["entrada"]), shift_day, config, is_exit=False)
        exit_dt = parse_excel_datetime_value(attendance_row_values(row, idx["salida"]), shift_day, config, is_exit=True) if idx["salida"] is not None else None
        if not entry_dt:
            errors.append(f"Fila {row_number}: entrada inválida")
            continue
        if exit_dt and exit_dt < entry_dt:
            errors.append(f"Fila {row_number}: la salida es menor que la entrada")
            continue

        eval_entry_data = evaluate_entry(config, shift_day.isoformat(), entry_dt)
        entry_stat = eval_entry_data["status"]
        if entry_stat == "Retardo" and not motivo_retardo:
            motivo_retardo = "NO REGISTRADO - CARGA HISTORICA"
            observaciones = (observaciones + " | " if observaciones else "") + "Retardo histórico sin motivo original."

        eval_exit_data = evaluate_exit(config, shift_day.isoformat(), exit_dt) if exit_dt else {
            "status": "",
            "scheduled_entry_at": eval_entry_data["scheduled_entry_at"],
            "scheduled_exit_at": eval_entry_data["scheduled_exit_at"],
            "entry_limit_at": eval_entry_data["entry_limit_at"],
            "exit_early_limit_at": eval_entry_data["exit_early_limit_at"],
            "extra_limit_at": eval_entry_data["extra_limit_at"],
            "entry_tolerance_minutes": eval_entry_data["entry_tolerance_minutes"],
            "exit_tolerance_minutes": eval_entry_data["exit_tolerance_minutes"],
            "extra_after_minutes": eval_entry_data["extra_after_minutes"],
            "early_minutes": 0,
            "extra_minutes": 0,
        }
        if exit_dt and eval_exit_data["status"] == "Salida temprana" and not motivo_salida:
            motivo_salida = "NO REGISTRADO - CARGA HISTORICA"
            observaciones = (observaciones + " | " if observaciones else "") + "Salida temprana histórica sin motivo original."

        key = id_registro or f"{employee_id}|{shift_day.isoformat()}|{turno}"
        if key in seen_keys:
            errors.append(f"Fila {row_number}: registro duplicado en el Excel: {key}")
            continue
        seen_keys.add(key)

        rows_to_apply.append({
            "row_number": row_number,
            "id_registro": id_registro,
            "employee_id": employee_id,
            "shift_date": shift_day.isoformat(),
            "turno": turno,
            "entry_at": entry_dt.isoformat(),
            "exit_at": exit_dt.isoformat() if exit_dt else "",
            "entry_guard": guardia_entrada or "Importación",
            "exit_guard": guardia_salida if exit_dt else "",
            "entry_status": entry_stat,
            "exit_status": eval_exit_data.get("status") or "",
            "late_reason": motivo_retardo,
            "early_reason": motivo_salida,
            "vehicle_expected": vehiculo,
            "vehicle_entered": vehiculo,
            "incident": observaciones,
            "scheduled_entry_at": eval_entry_data["scheduled_entry_at"],
            "scheduled_exit_at": eval_entry_data["scheduled_exit_at"],
            "entry_limit_at": eval_entry_data["entry_limit_at"],
            "exit_early_limit_at": eval_exit_data["exit_early_limit_at"],
            "extra_limit_at": eval_exit_data["extra_limit_at"],
            "entry_tolerance_minutes": eval_entry_data["entry_tolerance_minutes"],
            "exit_tolerance_minutes": eval_exit_data["exit_tolerance_minutes"],
            "extra_after_minutes": eval_exit_data["extra_after_minutes"],
            "late_minutes": eval_entry_data["late_minutes"],
            "early_minutes": eval_exit_data.get("early_minutes") or 0,
            "extra_minutes": eval_exit_data.get("extra_minutes") or 0,
        })

    if errors:
        return templates.TemplateResponse("importar.html", {"request": request, "result_asistencias": {"ok": False, "errors": errors[:250], "total_errors": len(errors)}, "result": None})

    creados = 0
    actualizados = 0
    ts = now_mx().isoformat()
    with engine.begin() as conn:
        for item in rows_to_apply:
            existing = None
            if item["id_registro"]:
                existing = fetch_one(conn, "SELECT * FROM attendance WHERE id = :id", {"id": item["id_registro"]})
                if not existing:
                    raise HTTPException(status_code=400, detail=f"ID registro no existe: {item['id_registro']}")
            else:
                existing = fetch_one(conn, "SELECT * FROM attendance WHERE employee_id=:employee_id AND shift_date=:shift_date AND turno=:turno AND COALESCE(anulled,0)=0 ORDER BY id DESC LIMIT 1", item)
            if existing and not actualizar_asistencias:
                errors.append(f"Fila {item['row_number']}: ya existe registro para {item['employee_id']} {item['shift_date']} {item['turno']}. Marca actualizar históricos para modificarlo.")
                continue
            params = {**item, "updated_at": ts}
            if existing:
                params["id"] = existing["id"]
                conn.execute(text("""
                    UPDATE attendance
                    SET employee_id=:employee_id, shift_date=:shift_date, turno=:turno, entry_at=:entry_at, exit_at=:exit_at,
                        entry_guard=:entry_guard, exit_guard=:exit_guard, entry_status=:entry_status, exit_status=:exit_status,
                        late_reason=:late_reason, early_reason=:early_reason, vehicle_expected=:vehicle_expected, vehicle_entered=:vehicle_entered,
                        incident=:incident, scheduled_entry_at=:scheduled_entry_at, scheduled_exit_at=:scheduled_exit_at, entry_limit_at=:entry_limit_at,
                        exit_early_limit_at=:exit_early_limit_at, extra_limit_at=:extra_limit_at, entry_tolerance_minutes=:entry_tolerance_minutes,
                        exit_tolerance_minutes=:exit_tolerance_minutes, extra_after_minutes=:extra_after_minutes, late_minutes=:late_minutes,
                        early_minutes=:early_minutes, extra_minutes=:extra_minutes, updated_at=:updated_at
                    WHERE id=:id
                """), params)
                audit(conn, "attendance", str(existing["id"]), "UPDATE", admin_user["username"], reason="Importación histórica de asistencias")
                actualizados += 1
            else:
                params["created_at"] = ts
                result = conn.execute(text("""
                    INSERT INTO attendance (
                        employee_id, shift_date, turno, entry_at, exit_at, entry_guard, exit_guard, entry_status, exit_status,
                        late_reason, early_reason, vehicle_expected, vehicle_entered, vehicle_front_entry, vehicle_trunk_entry, vehicle_front_exit, vehicle_trunk_exit, incident,
                        scheduled_entry_at, scheduled_exit_at, entry_limit_at, exit_early_limit_at, extra_limit_at,
                        entry_tolerance_minutes, exit_tolerance_minutes, extra_after_minutes, late_minutes, early_minutes, extra_minutes,
                        late_justified, early_justified, extra_authorized, provisional_exit, review_required, review_status, auto_closed_at, anulled, created_at, updated_at
                    ) VALUES (
                        :employee_id, :shift_date, :turno, :entry_at, :exit_at, :entry_guard, :exit_guard, :entry_status, :exit_status,
                        :late_reason, :early_reason, :vehicle_expected, :vehicle_entered, '', '', '', '', :incident,
                        :scheduled_entry_at, :scheduled_exit_at, :entry_limit_at, :exit_early_limit_at, :extra_limit_at,
                        :entry_tolerance_minutes, :exit_tolerance_minutes, :extra_after_minutes, :late_minutes, :early_minutes, :extra_minutes,
                        0, 0, 0, 0, 0, '', '', 0, :created_at, :updated_at
                    ) RETURNING id
                """), params)
                new_id = result.scalar_one()
                audit(conn, "attendance", str(new_id), "CREATE", admin_user["username"], reason="Importación histórica de asistencias")
                creados += 1
        if errors:
            return templates.TemplateResponse("importar.html", {"request": request, "result_asistencias": {"ok": False, "errors": errors[:250], "total_errors": len(errors)}, "result": None})
        log_event(conn, "info", "importar", "attendance_import", f"Importación histórica: {creados} creados, {actualizados} actualizados", admin_user["username"])

    return templates.TemplateResponse("importar.html", {"request": request, "result": None, "result_asistencias": {"ok": True, "leidos": len(rows_to_apply), "creados": creados, "actualizados": actualizados}})


@app.get("/correcciones", response_class=HTMLResponse)
def correcciones_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    today = now_mx().date()
    month_start = today.replace(day=1).isoformat()
    with engine.begin() as conn:
        batches = fetch_all(conn, "SELECT * FROM correction_batches ORDER BY created_at DESC LIMIT 20")
    return templates.TemplateResponse(
        "correcciones.html",
        {
            "request": request,
            "fecha_inicio": month_start,
            "fecha_fin": today.isoformat(),
            "result": None,
            "batches": batches,
        },
    )


@app.get("/correcciones/exportar-editable.xlsx")
def correcciones_exportar_editable(request: Request, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    require_admin_http(request)
    with engine.begin() as conn:
        wb = build_editable_corrections_workbook(conn, fecha_inicio, fecha_fin)
    filename = f"correcciones_editables_{fecha_inicio or 'todo'}_{fecha_fin or 'todo'}_{now_mx().strftime('%Y%m%d_%H%M')}.xlsx"
    return workbook_response(wb, filename)


@app.post("/correcciones/analizar", response_class=HTMLResponse)
async def correcciones_analizar(request: Request, archivo: UploadFile = File(...)):
    admin_user = require_admin_http(request)
    content = await archivo.read()
    analysis = analyze_corrections_excel(content, admin_user["username"])
    batch_id = save_correction_preview(analysis, admin_user["username"], archivo.filename or "")
    today = now_mx().date()
    month_start = today.replace(day=1).isoformat()
    with engine.begin() as conn:
        batches = fetch_all(conn, "SELECT * FROM correction_batches ORDER BY created_at DESC LIMIT 20")
    return templates.TemplateResponse(
        "correcciones.html",
        {
            "request": request,
            "fecha_inicio": month_start,
            "fecha_fin": today.isoformat(),
            "result": {"batch_id": batch_id, "analysis": analysis},
            "batches": batches,
        },
    )


@app.post("/correcciones/aplicar", response_class=HTMLResponse)
def correcciones_aplicar(request: Request, batch_id: str = Form(...)):
    admin_user = require_admin_http(request)
    apply_result = apply_correction_batch(batch_id, admin_user["username"])
    today = now_mx().date()
    month_start = today.replace(day=1).isoformat()
    with engine.begin() as conn:
        batches = fetch_all(conn, "SELECT * FROM correction_batches ORDER BY created_at DESC LIMIT 20")
    return templates.TemplateResponse(
        "correcciones.html",
        {
            "request": request,
            "fecha_inicio": month_start,
            "fecha_fin": today.isoformat(),
            "result": {"apply_result": apply_result},
            "batches": batches,
        },
    )


@app.get("/qr", response_class=HTMLResponse)
def qr_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    with engine.begin() as conn:
        employees = fetch_all(conn, "SELECT * FROM employees ORDER BY nombre")
    return templates.TemplateResponse("qr.html", {"request": request, "employees": employees})




@app.get("/exportar", response_class=HTMLResponse)
def exportar_page(request: Request):
    admin_user = require_admin_page(request)
    if not admin_user:
        return admin_login_redirect(request)
    today = now_mx().date()
    month_start = today.replace(day=1).isoformat()
    return templates.TemplateResponse(
        "exportar.html",
        {
            "request": request,
            "fecha_inicio": month_start,
            "fecha_fin": today.isoformat(),
        },
    )


@app.get("/exportar/empleados.xlsx")
def exportar_empleados_excel(request: Request):
    require_admin_http(request)
    wb = Workbook()
    with engine.begin() as conn:
        headers, rows = get_employees_export_rows(conn)
    make_sheet(wb, "Empleados", headers, rows, "Empleados")
    remove_default_sheet(wb)
    filename = f"empleados_{now_mx().strftime('%Y%m%d_%H%M')}.xlsx"
    return workbook_response(wb, filename)


@app.get("/exportar/asistencias.xlsx")
def exportar_asistencias_excel(request: Request, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    require_admin_http(request)
    wb = Workbook()
    title = "Asistencias"
    if fecha_inicio or fecha_fin:
        title += f" ({fecha_inicio or 'inicio'} a {fecha_fin or 'fin'})"
    with engine.begin() as conn:
        headers, rows = get_attendance_export_rows(conn, fecha_inicio, fecha_fin)
    make_sheet(wb, "Asistencias", headers, rows, title)
    remove_default_sheet(wb)
    filename = f"asistencias_{fecha_inicio or 'todo'}_{fecha_fin or 'todo'}_{now_mx().strftime('%Y%m%d_%H%M')}.xlsx"
    return workbook_response(wb, filename)


@app.get("/exportar/incidencias.xlsx")
def exportar_incidencias_excel(request: Request, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    require_admin_http(request)
    wb = Workbook()
    title = "Incidencias"
    if fecha_inicio or fecha_fin:
        title += f" ({fecha_inicio or 'inicio'} a {fecha_fin or 'fin'})"
    with engine.begin() as conn:
        headers, rows = get_incidents_export_rows(conn, fecha_inicio, fecha_fin)
    make_sheet(wb, "Incidencias", headers, rows, title)
    remove_default_sheet(wb)
    filename = f"incidencias_{fecha_inicio or 'todo'}_{fecha_fin or 'todo'}_{now_mx().strftime('%Y%m%d_%H%M')}.xlsx"
    return workbook_response(wb, filename)


@app.get("/exportar/auditoria.xlsx")
def exportar_auditoria_excel(request: Request):
    require_admin_http(request)
    wb = Workbook()
    with engine.begin() as conn:
        headers, rows = get_audit_export_rows(conn)
    make_sheet(wb, "Auditoria", headers, rows, "Auditoría")
    remove_default_sheet(wb)
    filename = f"auditoria_{now_mx().strftime('%Y%m%d_%H%M')}.xlsx"
    return workbook_response(wb, filename)


@app.get("/exportar/general.xlsx")
def exportar_general_excel(request: Request, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    require_admin_http(request)
    wb = Workbook()
    with engine.begin() as conn:
        add_summary_sheet(wb, conn, fecha_inicio, fecha_fin)
        emp_headers, emp_rows = get_employees_export_rows(conn)
        att_headers, att_rows = get_attendance_export_rows(conn, fecha_inicio, fecha_fin)
        inc_headers, inc_rows = get_incidents_export_rows(conn, fecha_inicio, fecha_fin)
        aud_headers, aud_rows = get_audit_export_rows(conn)
    make_sheet(wb, "Empleados", emp_headers, emp_rows, "Empleados")
    make_sheet(wb, "Asistencias", att_headers, att_rows, "Asistencias")
    make_sheet(wb, "Incidencias", inc_headers, inc_rows, "Incidencias")
    make_sheet(wb, "Auditoria", aud_headers, aud_rows, "Auditoría")
    remove_default_sheet(wb)
    filename = f"reporte_general_{fecha_inicio or 'todo'}_{fecha_fin or 'todo'}_{now_mx().strftime('%Y%m%d_%H%M')}.xlsx"
    return workbook_response(wb, filename)


@app.get("/exportar/nomina.xlsx")
def exportar_nomina_excel(request: Request, fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None):
    require_admin_http(request)
    today = now_mx().date()
    start = safe_date(fecha_inicio, today.replace(day=1))
    end = safe_date(fecha_fin, today)
    if end < start:
        start, end = end, start
    with engine.begin() as conn:
        employees = active_employee_rows(conn)
        faltas_by_employee, _ = absence_map(conn, start, end)
        records = fetch_all(conn, """
            SELECT a.*, e.nombre, e.area, e.puesto
            FROM attendance a
            JOIN employees e ON e.id=a.employee_id
            WHERE a.shift_date >= :start AND a.shift_date <= :end AND COALESCE(a.anulled,0)=0
            ORDER BY e.nombre, a.shift_date
        """, {"start": start.isoformat(), "end": end.isoformat()})

    by_emp = {e["id"]: {"id": e["id"], "nombre": e["nombre"], "area": e.get("area") or "", "puesto": e.get("puesto") or "", "min_incidencia": 0, "faltas": faltas_by_employee.get(e["id"], 0), "worked": 0, "retardos": 0, "tempranas": 0, "pendientes_comida": 0} for e in employees}
    detail_rows = []
    for r in records:
        emp_id = r["employee_id"]
        if emp_id not in by_emp:
            continue
        late = int(r.get("late_minutes") or 0)
        early = int(r.get("early_minutes") or 0)
        worked, lunch, note = compute_worked_minutes(r.get("entry_at"), r.get("exit_at"), r.get("scheduled_entry_at"), r.get("scheduled_exit_at"), r.get("lunch_taken") or "", r.get("exit_status") or "")
        by_emp[emp_id]["min_incidencia"] += late + early
        by_emp[emp_id]["worked"] += worked
        by_emp[emp_id]["retardos"] += 1 if r.get("entry_status") in ("Tarde", "Retardo") else 0
        by_emp[emp_id]["tempranas"] += 1 if r.get("exit_status") == "Salida temprana" else 0
        if note:
            by_emp[emp_id]["pendientes_comida"] += 1
        detail_rows.append([r.get("shift_date"), emp_id, r.get("nombre"), r.get("turno"), short_datetime(r.get("entry_at")), short_datetime(r.get("exit_at")), r.get("entry_status") or "", r.get("exit_status") or "", late, early, late+early, lunch, worked, round(worked/60,2), note])

    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina"
    ws.append(["ID", "Nombre completo", "Área", "Puesto", "Min retardo + salida temprana", "Faltas", "Tiempo total trabajado (min)", "Tiempo total trabajado (h)", "Retardos", "Salidas tempranas", "Pendientes comida"] )
    for item in sorted(by_emp.values(), key=lambda x: x["nombre"]):
        ws.append([item["id"], item["nombre"], item["area"], item["puesto"], item["min_incidencia"], item["faltas"], item["worked"], round(item["worked"]/60,2), item["retardos"], item["tempranas"], item["pendientes_comida"]])
    style_worksheet(ws, f"Exportación nómina {start.isoformat()} a {end.isoformat()}")
    ws2 = wb.create_sheet("Detalle")
    ws2.append(["Fecha turno", "ID", "Nombre", "Turno", "Entrada", "Salida", "Estado entrada", "Estado salida", "Min retardo", "Min salida temprana", "Min incidencia", "Comida tomada", "Min trabajados", "Horas trabajadas", "Observación"] )
    for row in detail_rows:
        ws2.append(row)
    style_worksheet(ws2, "Detalle para nómina")
    return workbook_response(wb, f"nomina_{start.isoformat()}_{end.isoformat()}.xlsx")


def make_qr_core(employee_id: str, size: int = 900) -> Image.Image:
    return qrcode.make(employee_id).convert("RGB").resize((size, size))


def generate_simple_qr_image(employee_id: str) -> Image.Image:
    return make_qr_core(employee_id, 900)


def text_fit(draw: ImageDraw.ImageDraw, text_value: str, max_width: int, font_path_size: int, bold: bool = False, min_size: int = 28):
    size = font_path_size
    while size >= min_size:
        font = load_font(size, bold)
        bbox = draw.textbbox((0, 0), text_value, font=font)
        if bbox[2] - bbox[0] <= max_width:
            return font
        size -= 2
    return load_font(min_size, bold)


def generate_cell_card_image(emp: dict) -> Image.Image:
    employee_id = emp["id"]
    W, H = 1080, 1920
    azul = (5, 26, 57)
    azul2 = (11, 42, 85)
    azul_claro = (227, 239, 255)
    gris = (244, 247, 251)
    blanco = (255, 255, 255)
    oscuro = (24, 32, 42)
    muted = (92, 104, 120)

    img = Image.new("RGB", (W, H), gris)
    draw = ImageDraw.Draw(img)

    # Header sobrio en azul. Nada de carnaval naranja, por fin.
    draw.rectangle((0, 0, W, 390), fill=azul)
    draw.rectangle((0, 330, W, 390), fill=azul2)

    font_brand = load_font(54, True)
    font_subtitle = load_font(32, False)
    font_name = text_fit(draw, emp.get("nombre") or employee_id, W - 180, 72, True, 42)
    font_label = load_font(32, False)
    font_id = load_font(50, True)
    font_small = load_font(28, False)
    font_tiny = load_font(24, False)

    center_text(draw, (0, 76, W), "MS", font_brand, blanco)
    center_text(draw, (0, 150, W), "Código personal de acceso", font_subtitle, (220, 235, 248))

    # Tarjeta principal
    draw.rounded_rectangle((62, 260, W - 62, H - 116), radius=62, fill=blanco)
    draw.rounded_rectangle((112, 328, W - 112, 1250), radius=48, fill=(250, 252, 255), outline=azul_claro, width=6)

    qr_img = make_qr_core(employee_id, 760)
    img.paste(qr_img, ((W - 760) // 2, 410))

    # Nombre + ID
    center_text(draw, (96, 1310, W - 192), emp.get("nombre") or "Empleado", font_name, oscuro)
    center_text(draw, (96, 1408, W - 192), employee_id, font_id, azul)

    info_parts = [x for x in [emp.get("area") or "Área no asignada", emp.get("puesto") or "", emp.get("turno") or "Turno no asignado"] if x]
    info = " · ".join(info_parts)
    info_font = text_fit(draw, info, W - 200, 34, False, 24)
    center_text(draw, (100, 1490, W - 200), info, info_font, muted)

    vehicle_text = "CON VEHÍCULO" if emp.get("tiene_vehiculo") else "SIN VEHÍCULO"
    badge_w, badge_h = 430, 78
    bx = (W - badge_w) // 2
    by = 1588
    draw.rounded_rectangle((bx, by, bx + badge_w, by + badge_h), radius=39, fill=azul)
    center_text(draw, (bx, by + 20, badge_w), vehicle_text, font_label, blanco)

    center_text(draw, (100, 1740, W - 200), "Muestra esta imagen a vigilancia", font_small, muted)
    center_text(draw, (100, 1790, W - 200), "El QR contiene únicamente el ID del empleado", font_tiny, (120, 130, 145))

    return img


def png_response(img: Image.Image, filename: str, inline: bool = True):
    output = io.BytesIO()
    img.save(output, format="PNG")
    output.seek(0)
    disposition = "inline" if inline else "attachment"
    return StreamingResponse(output, media_type="image/png", headers={"Content-Disposition": f"{disposition}; filename={filename}"})


def image_png_bytes(img: Image.Image, dpi: Optional[tuple] = None) -> bytes:
    output = io.BytesIO()
    save_kwargs = {"format": "PNG", "optimize": True, "compress_level": 6}
    if dpi:
        save_kwargs["dpi"] = dpi
    img.save(output, **save_kwargs)
    return output.getvalue()


def png_response_dpi(img: Image.Image, filename: str, inline: bool = True, dpi: Optional[tuple] = None):
    output = io.BytesIO(image_png_bytes(img, dpi=dpi))
    disposition = "inline" if inline else "attachment"
    return StreamingResponse(output, media_type="image/png", headers={"Content-Disposition": f"{disposition}; filename={filename}"})


@app.get("/qr/{employee_id}.png")
def qr_png(employee_id: str):
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
    filename = f"QR_{employee_file_base(emp)}.png"
    return png_response(generate_simple_qr_image(employee_id), filename, inline=True)


def load_font(size: int, bold: bool = False):
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def center_text(draw: ImageDraw.ImageDraw, xy, text_value: str, font, fill):
    x, y, width = xy
    bbox = draw.textbbox((0, 0), text_value, font=font)
    tw = bbox[2] - bbox[0]
    draw.text((x + (width - tw) / 2, y), text_value, font=font, fill=fill)


@app.get("/qr/card/{employee_id}.png")
def qr_card_png(employee_id: str):
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
    filename = f"CELULAR_{employee_file_base(emp)}.png"
    return png_response(generate_cell_card_image(emp), filename, inline=True)


def draw_wrapped_center(draw: ImageDraw.ImageDraw, text_value: str, box, font, fill, line_gap: int = 8, max_lines: int = 2):
    x1, y1, x2, y2 = box
    words = str(text_value or "").split()
    if not words:
        return
    lines = []
    current = ""
    for word in words:
        test = (current + " " + word).strip()
        bbox = draw.textbbox((0, 0), test, font=font)
        if bbox[2] - bbox[0] <= (x2 - x1) or not current:
            current = test
        else:
            lines.append(current)
            current = word
            if len(lines) >= max_lines - 1:
                break
    if current and len(lines) < max_lines:
        lines.append(current)
    if len(lines) == max_lines and len(words) > len(" ".join(lines).split()):
        last = lines[-1]
        while last and draw.textbbox((0, 0), last + "...", font=font)[2] > (x2 - x1):
            last = last[:-1].rstrip()
        lines[-1] = last + "..."
    heights = [draw.textbbox((0, 0), ln, font=font)[3] - draw.textbbox((0, 0), ln, font=font)[1] for ln in lines]
    total_h = sum(heights) + line_gap * (len(lines)-1)
    y = y1 + ((y2 - y1) - total_h) / 2
    for ln, h in zip(lines, heights):
        bbox = draw.textbbox((0, 0), ln, font=font)
        tw = bbox[2] - bbox[0]
        draw.text((x1 + ((x2-x1)-tw)/2, y), ln, font=font, fill=fill)
        y += h + line_gap


def generate_badge_card_image(emp: dict) -> Image.Image:
    """Gafete vertical 6 x 10 cm a 300 dpi: 708 x 1181 px."""
    employee_id = emp["id"]
    W, H = 708, 1181
    azul = (5, 26, 57)
    azul2 = (12, 48, 96)
    azul_claro = (229, 240, 255)
    blanco = (255, 255, 255)
    gris = (246, 248, 252)
    oscuro = (24, 32, 42)
    muted = (92, 104, 120)

    img = Image.new("RGB", (W, H), blanco)
    draw = ImageDraw.Draw(img)

    # Fondo y encabezado
    draw.rounded_rectangle((22, 22, W-22, H-22), radius=34, fill=gris, outline=(210, 220, 235), width=3)
    draw.rounded_rectangle((22, 22, W-22, 205), radius=34, fill=azul)
    draw.rectangle((22, 145, W-22, 205), fill=azul)
    center_text(draw, (0, 56, W), "MS", load_font(58, True), blanco)
    center_text(draw, (0, 112, W), "GAFETE DE ACCESO", load_font(20, False), (220, 235, 248))

    # QR
    qr_size = 435
    qr = make_qr_core(employee_id, qr_size)
    qr_x = (W - qr_size) // 2
    qr_y = 250
    draw.rounded_rectangle((qr_x-24, qr_y-24, qr_x+qr_size+24, qr_y+qr_size+24), radius=34, fill=blanco, outline=azul_claro, width=5)
    img.paste(qr, (qr_x, qr_y))

    # Nombre y datos
    name_font = text_fit(draw, emp.get("nombre") or employee_id, W-100, 44, True, 27)
    draw_wrapped_center(draw, emp.get("nombre") or "Empleado", (56, 742, W-56, 835), name_font, oscuro, line_gap=5, max_lines=2)
    center_text(draw, (56, 850, W-112), employee_id, load_font(34, True), azul)

    info = " · ".join([x for x in [emp.get("area") or "Área", emp.get("puesto") or "", emp.get("turno") or "Turno"] if x])
    info_font = text_fit(draw, info, W-90, 22, False, 17)
    center_text(draw, (45, 910, W-90), info, info_font, muted)

    vehicle_text = "CON VEHÍCULO" if emp.get("tiene_vehiculo") else "SIN VEHÍCULO"
    badge_w, badge_h = 300, 52
    bx = (W-badge_w)//2
    by = 970
    draw.rounded_rectangle((bx, by, bx+badge_w, by+badge_h), radius=26, fill=azul2)
    center_text(draw, (bx, by+14, badge_w), vehicle_text, load_font(18, True), blanco)

    center_text(draw, (50, 1062, W-100), "Escanear en vigilancia", load_font(20, False), muted)
    center_text(draw, (50, 1092, W-100), "El QR contiene únicamente el ID", load_font(16, False), (120,130,145))
    return img


@app.get("/qr/gafete/{employee_id}.png")
def qr_badge_png(employee_id: str):
    employee_id = clean_id(employee_id)
    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
    filename = f"GAFETE_6x10cm_{employee_file_base(emp)}.png"
    return png_response_dpi(generate_badge_card_image(emp), filename, inline=True, dpi=(300, 300))


@app.get("/qr/gafetes/todos.zip")
def qr_badges_zip(request: Request):
    require_admin_http(request)
    with engine.begin() as conn:
        employees = fetch_all(conn, "SELECT * FROM employees ORDER BY nombre")
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for emp in employees:
            img = generate_badge_card_image(emp)
            zf.writestr(f"GAFETE_6x10cm_{employee_file_base(emp)}.png", image_png_bytes(img, dpi=(300, 300)))
        zf.writestr("LEEME.txt", "Gafetes verticales de 10 cm de alto por 6 cm de ancho, generados a 300 dpi. Cada archivo contiene nombre completo e ID del empleado.\n")
    output.seek(0)
    filename = f"gafetes_6x10cm_{now_mx().strftime('%Y%m%d_%H%M')}.zip"
    return StreamingResponse(output, media_type="application/zip", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def generate_guard_qr_image(guard: dict) -> Image.Image:
    W, H = 900, 1200
    azul = (5, 26, 57)
    gris = (244, 247, 251)
    blanco = (255, 255, 255)
    img = Image.new("RGB", (W, H), gris)
    draw = ImageDraw.Draw(img)
    draw.rectangle((0, 0, W, 210), fill=azul)
    center_text(draw, (0, 56, W), "VIGILANTE MS", load_font(48, True), blanco)
    center_text(draw, (0, 130, W), "Código de cambio de turno", load_font(28, False), (220, 235, 248))
    qr_img = make_qr_core(f"VIG:{guard['code']}", 650)
    img.paste(qr_img, ((W-650)//2, 270))
    center_text(draw, (60, 960, W-120), guard_display(guard), text_fit(draw, guard_display(guard), W-120, 46, True, 30), azul)
    center_text(draw, (60, 1030, W-120), guard["code"], load_font(34, True), (24,32,42))
    center_text(draw, (60, 1090, W-120), "Escanear al iniciar turno de vigilancia", load_font(26, False), (92,104,120))
    return img


@app.get("/vigilantes/{guard_code}.png")
def guard_qr_png(request: Request, guard_code: str):
    require_admin_http(request)
    guard_code = clean_guard_code(guard_code)
    with engine.begin() as conn:
        guard = get_guard(conn, guard_code)
        if not guard:
            raise HTTPException(status_code=404, detail="Vigilante no encontrado")
    filename = f"VIGILANTE_{filename_slug(guard.get('alias') or guard_code)}_{filename_slug(guard_code)}.png"
    return png_response(generate_guard_qr_image(guard), filename, inline=True)


@app.get("/qr/celular/todos.zip")
def qr_cell_cards_zip(request: Request):
    require_admin_http(request)
    with engine.begin() as conn:
        employees = fetch_all(conn, "SELECT * FROM employees ORDER BY nombre")
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for emp in employees:
            img = generate_cell_card_image(emp)
            buff = io.BytesIO()
            # Se reduce un poco para exportación masiva: mejor descarga estable que museo de PNGs gigantes.
            img.thumbnail((720, 1280))
            img.save(buff, format="PNG", optimize=True, compress_level=6)
            zf.writestr(f"CELULAR_{employee_file_base(emp)}.png", buff.getvalue())
        zf.writestr("LEEME.txt", "Imágenes QR para celular generadas por el sistema. Cada archivo contiene nombre completo e ID del empleado.\n")
    output.seek(0)
    filename = f"imagenes_celular_qr_{now_mx().strftime('%Y%m%d_%H%M')}.zip"
    return StreamingResponse(output, media_type="application/zip", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/qr/simples/todos.zip")
def qr_simple_zip(request: Request):
    require_admin_http(request)
    with engine.begin() as conn:
        employees = fetch_all(conn, "SELECT * FROM employees ORDER BY nombre")
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for emp in employees:
            img = generate_simple_qr_image(emp["id"])
            buff = io.BytesIO()
            img.save(buff, format="PNG", optimize=True, compress_level=6)
            zf.writestr(f"QR_{employee_file_base(emp)}.png", buff.getvalue())
        zf.writestr("LEEME.txt", "QR simples generados por el sistema. Cada archivo contiene nombre completo e ID del empleado.\n")
    output.seek(0)
    filename = f"qr_simples_{now_mx().strftime('%Y%m%d_%H%M')}.zip"
    return StreamingResponse(output, media_type="application/zip", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# -----------------------------
# API para vigilancia
# -----------------------------

@app.get("/api/scan/{raw_code}")
def api_scan(request: Request, raw_code: str):
    user = require_vigilancia_http(request)
    raw_code = as_text(raw_code)
    guard_code = clean_guard_code(raw_code)
    employee_id = clean_employee_id(raw_code)
    dt = now_mx()
    with engine.begin() as conn:
        auto_close_overdue_records(conn, user["username"])
        guard = find_guard(conn, raw_code)
        if guard:
            if not int(guard.get("active") or 0):
                return JSONResponse(status_code=400, content={"ok": False, "message": f"Vigilante inactivo: {guard_display(guard)}"})
            if not int(guard.get("qr_activo") or 0):
                return JSONResponse(status_code=400, content={"ok": False, "message": "QR de vigilante inactivo"})
            active = set_active_guard(conn, guard, user["username"])
            return {"ok": True, "type": "guard", "message": f"Vigilante activo: {active['display']}", "active_guard": active}

        active_guard = get_active_guard(conn)
        if not active_guard:
            return JSONResponse(status_code=409, content={"ok": False, "message": "Primero escanea el QR del vigilante para iniciar captura."})

        emp = get_employee(conn, employee_id)
        if not emp:
            return JSONResponse(status_code=404, content={"ok": False, "message": "Código no encontrado como empleado ni vigilante"})
        open_att = find_open_attendance(conn, employee_id)
        config = get_shift_config(conn, emp.get("turno") or "Día")
        if open_att:
            entry_dt_open = parse_dt(open_att.get("entry_at") or "")
            if entry_dt_open and (dt - entry_dt_open).total_seconds() < MIN_EXIT_AFTER_ENTRY_MINUTES * 60:
                remaining = MIN_EXIT_AFTER_ENTRY_MINUTES - int((dt - entry_dt_open).total_seconds() // 60)
                return JSONResponse(status_code=409, content={"ok": False, "message": f"Entrada recién registrada. La salida solo puede capturarse mínimo {MIN_EXIT_AFTER_ENTRY_MINUTES} minutos después. Espera aprox. {max(1, remaining)} min."})
            exit_config = get_shift_config(conn, open_att.get("turno") or emp.get("turno") or "Día")
            preview = evaluate_exit(exit_config, open_att["shift_date"], dt)
            preview["movement"] = "salida"
        else:
            shift_date = current_shift_date(emp.get("turno") or "Día", dt, config)
            latest_shift_att = find_latest_shift_attendance(conn, employee_id, shift_date, emp.get("turno") or "Día")
            preview = evaluate_entry(config, shift_date, dt)
            if latest_shift_att and not is_open_attendance_row(latest_shift_att):
                preview["status"] = "Reentrada"
                preview["late_minutes"] = 0
            preview["movement"] = "entrada"
            preview["shift_date"] = shift_date
            confirmation = build_entry_confirmation(config, shift_date, dt, latest_shift_att)
            if confirmation:
                preview.update(confirmation)
    return {
        "ok": True,
        "type": "employee",
        "employee": emp,
        "foto_url": "",
        "has_open_attendance": bool(open_att),
        "next_movement": preview.get("movement"),
        "open_attendance": open_att,
        "preview": preview,
        "active_guard": active_guard,
        "now": dt.isoformat(),
    }


@app.get("/api/empleado/{employee_id}")
def api_empleado(request: Request, employee_id: str):
    user = require_vigilancia_http(request)
    employee_id = clean_employee_id(employee_id)
    if not employee_id:
        return JSONResponse(status_code=400, content={"ok": False, "message": "El ID de empleado debe ser numérico"})
    dt = now_mx()
    with engine.begin() as conn:
        auto_close_overdue_records(conn, user["username"])
        active_guard = get_active_guard(conn)
        if not active_guard:
            return JSONResponse(status_code=409, content={"ok": False, "message": "Primero escanea el QR del vigilante para iniciar captura."})
        emp = get_employee(conn, employee_id)
        if not emp:
            return JSONResponse(status_code=404, content={"ok": False, "message": "Empleado no encontrado"})
        open_att = find_open_attendance(conn, employee_id)
        config = get_shift_config(conn, emp.get("turno") or "Día")
        if open_att:
            entry_dt_open = parse_dt(open_att.get("entry_at") or "")
            if entry_dt_open and (dt - entry_dt_open).total_seconds() < MIN_EXIT_AFTER_ENTRY_MINUTES * 60:
                remaining = MIN_EXIT_AFTER_ENTRY_MINUTES - int((dt - entry_dt_open).total_seconds() // 60)
                return JSONResponse(status_code=409, content={"ok": False, "message": f"Entrada recién registrada. La salida solo puede capturarse mínimo {MIN_EXIT_AFTER_ENTRY_MINUTES} minutos después. Espera aprox. {max(1, remaining)} min."})
            exit_config = get_shift_config(conn, open_att.get("turno") or emp.get("turno") or "Día")
            preview = evaluate_exit(exit_config, open_att["shift_date"], dt)
            preview["movement"] = "salida"
        else:
            shift_date = current_shift_date(emp.get("turno") or "Día", dt, config)
            latest_shift_att = find_latest_shift_attendance(conn, employee_id, shift_date, emp.get("turno") or "Día")
            preview = evaluate_entry(config, shift_date, dt)
            if latest_shift_att and not is_open_attendance_row(latest_shift_att):
                preview["status"] = "Reentrada"
                preview["late_minutes"] = 0
            preview["movement"] = "entrada"
            preview["shift_date"] = shift_date
            confirmation = build_entry_confirmation(config, shift_date, dt, latest_shift_att)
            if confirmation:
                preview.update(confirmation)
    return {
        "ok": True,
        "employee": emp,
        "foto_url": "",
        "has_open_attendance": bool(open_att),
        "next_movement": preview.get("movement"),
        "open_attendance": open_att,
        "preview": preview,
        "active_guard": active_guard,
        "now": dt.isoformat(),
    }


@app.post("/api/registro")
async def api_registro(
    request: Request,
    employee_id: str = Form(...),
    movimiento: str = Form(...),
    guardia: str = Form("Vigilancia"),
    vehiculo: str = Form("0"),
    motivo_retardo: str = Form(""),
    motivo_salida_temprana: str = Form(""),
    comida_tomada: str = Form(""),
    confirmacion_operativa: str = Form(""),
    observaciones: str = Form(""),
    foto_frontal: Optional[UploadFile] = File(None),
    foto_cajuela: Optional[UploadFile] = File(None),
):
    user = require_vigilancia_http(request)
    employee_id = clean_employee_id(employee_id)
    if not employee_id:
        return JSONResponse(status_code=400, content={"ok": False, "message": "El ID de empleado debe ser numérico"})
    dt = now_mx()
    ts = dt.isoformat()

    with engine.begin() as conn:
        auto_close_overdue_records(conn, user["username"])
        emp = get_employee(conn, employee_id)
        if not emp:
            return JSONResponse(status_code=404, content={"ok": False, "message": "Empleado no encontrado"})
        if emp["estado"] != "Activo":
            return JSONResponse(status_code=400, content={"ok": False, "message": f"Empleado no activo: {emp['estado']}"})
        if not emp["qr_activo"]:
            return JSONResponse(status_code=400, content={"ok": False, "message": "QR inactivo"})

        active_guard = get_active_guard(conn)
        if not active_guard:
            return JSONResponse(status_code=409, content={"ok": False, "message": "Primero escanea el QR del vigilante para iniciar captura."})
        active_guard_name = active_guard.get("display") or "Vigilancia"

        vehicle_required = bool(int(vehiculo or "0")) or bool(emp["tiene_vehiculo"])

        # El sistema decide el movimiento. Si hay entrada abierta, toca salida.
        # Si no hay entrada abierta, toca entrada. El formulario no manda aquí.
        open_att = find_open_attendance(conn, employee_id)
        movimiento = "salida" if open_att else "entrada"
        if movimiento == "entrada":
            config_for_block = get_shift_config(conn, emp.get("turno") or "Día")
            shift_date_for_block = current_shift_date(emp.get("turno") or "Día", dt, config_for_block)
            latest_shift_att = find_latest_shift_attendance(conn, employee_id, shift_date_for_block, emp.get("turno") or "Día")
            confirmation = build_entry_confirmation(config_for_block, shift_date_for_block, dt, latest_shift_att)
            if confirmation and as_text(confirmacion_operativa).upper() != "CONFIRMAR":
                return entry_confirmation_required_response(config_for_block, shift_date_for_block, dt, latest_shift_att)

        # Versión Render Free: no se capturan ni almacenan imágenes.
        front_path = ""
        trunk_path = ""

        if movimiento == "entrada":
            config = get_shift_config(conn, emp["turno"])
            shift_date = current_shift_date(emp["turno"], dt, config)
            eval_data = evaluate_entry(config, shift_date, dt)
            if 'latest_shift_att' in locals() and latest_shift_att and not is_open_attendance_row(latest_shift_att):
                eval_data["status"] = "Reentrada"
                eval_data["late_minutes"] = 0
            status = eval_data["status"]
            if status == "Retardo" and not motivo_retardo.strip():
                return JSONResponse(status_code=400, content={"ok": False, "message": "Retardo detectado: el motivo es obligatorio"})

            result = conn.execute(
                text(
                    """
                    INSERT INTO attendance (
                        employee_id, shift_date, turno, entry_at, entry_guard, entry_status, late_reason,
                        vehicle_expected, vehicle_entered, vehicle_front_entry, vehicle_trunk_entry, incident,
                        scheduled_entry_at, scheduled_exit_at, entry_limit_at, exit_early_limit_at, extra_limit_at,
                        entry_tolerance_minutes, exit_tolerance_minutes, extra_after_minutes, late_minutes,
                        early_minutes, extra_minutes, late_justified, early_justified, extra_authorized, created_at, updated_at
                    ) VALUES (
                        :employee_id, :shift_date, :turno, :entry_at, :entry_guard, :entry_status, :late_reason,
                        :vehicle_expected, :vehicle_entered, :vehicle_front_entry, :vehicle_trunk_entry, :incident,
                        :scheduled_entry_at, :scheduled_exit_at, :entry_limit_at, :exit_early_limit_at, :extra_limit_at,
                        :entry_tolerance_minutes, :exit_tolerance_minutes, :extra_after_minutes, :late_minutes,
                        0, 0, 0, 0, 0, :created_at, :updated_at
                    ) RETURNING id
                    """
                ),
                {
                    "employee_id": employee_id,
                    "shift_date": shift_date,
                    "turno": emp["turno"],
                    "entry_at": ts,
                    "entry_guard": active_guard_name,
                    "entry_status": status,
                    "late_reason": motivo_retardo.strip(),
                    "vehicle_expected": 1 if emp["tiene_vehiculo"] else 0,
                    "vehicle_entered": 1 if vehicle_required else 0,
                    "vehicle_front_entry": front_path,
                    "vehicle_trunk_entry": trunk_path,
                    "incident": observaciones.strip(),
                    **eval_data,
                    "created_at": ts,
                    "updated_at": ts,
                }
            )
            record_id = str(result.scalar_one())
            audit(conn, "attendance", record_id, "ENTRY", user["username"], reason=f"{status}; retardo {eval_data.get('late_minutes', 0)} min")
            log_event(conn, "info", "vigilancia", "entry", f"Entrada {employee_id}: {status}", user["username"])
            msg = f"Entrada registrada: {status}"
            if status == "Retardo":
                msg += f" ({eval_data.get('late_minutes', 0)} min)"
            return {"ok": True, "message": msg, "record_id": record_id, "status": status, "evaluation": eval_data}

        if movimiento == "salida":
            open_att = find_open_attendance(conn, employee_id)
            if not open_att:
                return JSONResponse(status_code=400, content={"ok": False, "message": "No hay entrada abierta para registrar salida"})

            config = get_shift_config(conn, open_att["turno"])
            entry_dt_open = parse_dt(open_att.get("entry_at") or "")
            if entry_dt_open and (dt - entry_dt_open).total_seconds() < MIN_EXIT_AFTER_ENTRY_MINUTES * 60:
                remaining = MIN_EXIT_AFTER_ENTRY_MINUTES - int((dt - entry_dt_open).total_seconds() // 60)
                return JSONResponse(status_code=409, content={"ok": False, "message": f"Entrada recién registrada. La salida solo puede capturarse mínimo {MIN_EXIT_AFTER_ENTRY_MINUTES} minutos después. Espera aprox. {max(1, remaining)} min."})
            eval_data = evaluate_exit(config, open_att["shift_date"], dt)
            status = eval_data["status"]
            lunch_value = as_text(comida_tomada).upper()
            if status == "Salida temprana" and not motivo_salida_temprana.strip():
                return JSONResponse(status_code=400, content={"ok": False, "message": "Salida temprana: el motivo es obligatorio"})
            if status == "Salida temprana" and lunch_value not in {"SI", "NO"}:
                return JSONResponse(status_code=400, content={"ok": False, "message": "Salida temprana: debes indicar si tomó la hora de comida"})
            if status != "Salida temprana" and not lunch_value:
                lunch_value = "SI"
            worked_minutes, lunch_value, payroll_note = compute_worked_minutes(open_att.get("entry_at"), ts, eval_data.get("scheduled_entry_at"), eval_data.get("scheduled_exit_at"), lunch_value, status)

            conn.execute(
                text(
                    """
                    UPDATE attendance
                    SET exit_at=:exit_at,
                        exit_guard=:exit_guard,
                        exit_status=:exit_status,
                        early_reason=:early_reason,
                        vehicle_front_exit=:vehicle_front_exit,
                        vehicle_trunk_exit=:vehicle_trunk_exit,
                        scheduled_entry_at=COALESCE(NULLIF(scheduled_entry_at, ''), :scheduled_entry_at),
                        scheduled_exit_at=COALESCE(NULLIF(scheduled_exit_at, ''), :scheduled_exit_at),
                        entry_limit_at=COALESCE(NULLIF(entry_limit_at, ''), :entry_limit_at),
                        exit_early_limit_at=:exit_early_limit_at,
                        extra_limit_at=:extra_limit_at,
                        entry_tolerance_minutes=:entry_tolerance_minutes,
                        exit_tolerance_minutes=:exit_tolerance_minutes,
                        extra_after_minutes=:extra_after_minutes,
                        early_minutes=:early_minutes,
                        extra_minutes=:extra_minutes,
                        lunch_taken=:lunch_taken,
                        worked_minutes=:worked_minutes,
                        updated_at=:updated_at,
                        incident=:incident
                    WHERE id=:id
                    """
                ),
                {
                    "exit_at": ts,
                    "exit_guard": active_guard_name,
                    "exit_status": status,
                    "early_reason": motivo_salida_temprana.strip(),
                    "vehicle_front_exit": front_path,
                    "vehicle_trunk_exit": trunk_path,
                    "updated_at": ts,
                    "incident": observaciones.strip() or open_att["incident"],
                    "lunch_taken": lunch_value,
                    "worked_minutes": worked_minutes,
                    "id": open_att["id"],
                    **eval_data,
                }
            )
            audit(conn, "attendance", str(open_att["id"]), "EXIT", user["username"], reason=f"{status}; temprano {eval_data.get('early_minutes', 0)} min; extra {eval_data.get('extra_minutes', 0)} min")
            log_event(conn, "info", "vigilancia", "exit", f"Salida {employee_id}: {status}", user["username"])
            msg = f"Salida registrada: {status}"
            if status == "Extra":
                msg += f" ({eval_data.get('extra_minutes', 0)} min después de salida programada)"
            if status == "Salida temprana":
                msg += f" ({eval_data.get('early_minutes', 0)} min; comida: {lunch_value})"
            return {"ok": True, "message": msg, "record_id": open_att["id"], "status": status, "evaluation": eval_data, "worked_minutes": worked_minutes, "lunch_taken": lunch_value}

    return JSONResponse(status_code=400, content={"ok": False, "message": "Movimiento inválido"})



@app.post("/manual/entrada", response_class=HTMLResponse)
def manual_entrada(
    request: Request,
    employee_id: str = Form(...),
    fecha: str = Form(""),
    hora_entrada: str = Form("08:00"),
    motivo: str = Form(""),
    vehiculo: str = Form("0"),
):
    admin_user = require_admin_http(request)
    employee_id = clean_employee_id(employee_id)
    today = now_mx().date()
    shift_day = safe_date(fecha, today)
    motivo = as_text(motivo).strip()
    if not employee_id:
        raise HTTPException(status_code=400, detail="ID de empleado inválido")
    if not motivo:
        raise HTTPException(status_code=400, detail="El motivo de entrada manual es obligatorio")
    entrada_t = parse_time_value(hora_entrada, time(8, 0))
    entry_dt = datetime.combine(shift_day, entrada_t, tzinfo=TZ)
    ts = now_mx().isoformat()

    with engine.begin() as conn:
        emp = get_employee(conn, employee_id)
        if not emp:
            raise HTTPException(status_code=404, detail="Empleado no encontrado")
        if emp.get("estado") != "Activo":
            raise HTTPException(status_code=400, detail=f"Empleado no activo: {emp.get('estado')}")

        open_att = find_open_attendance(conn, employee_id) if "find_open_attendance" in globals() else fetch_one(
            conn,
            "SELECT * FROM attendance WHERE employee_id=:employee_id AND (exit_at IS NULL OR TRIM(COALESCE(exit_at,''))='') AND COALESCE(anulled,0)=0 ORDER BY id DESC LIMIT 1",
            {"employee_id": employee_id},
        )
        if open_att:
            raise HTTPException(status_code=409, detail="Este empleado ya tiene una entrada abierta. No se creó otra entrada manual.")

        config = get_shift_config(conn, emp.get("turno") or "Día")
        shift_date = current_shift_date(emp.get("turno") or "Día", entry_dt, config)
        existing_same_day = fetch_one(
            conn,
            """
            SELECT * FROM attendance
            WHERE employee_id=:employee_id
              AND shift_date=:shift_date
              AND turno=:turno
              AND COALESCE(anulled,0)=0
            ORDER BY id DESC
            LIMIT 1
            """,
            {"employee_id": employee_id, "shift_date": shift_date, "turno": emp.get("turno") or "Día"},
        )
        if existing_same_day:
            raise HTTPException(status_code=409, detail="Este empleado ya tiene registro para ese turno/día. Usa reentrada o correcciones para evitar duplicados.")

        eval_data = evaluate_entry(config, shift_date, entry_dt)
        incident = f"ENTRADA MANUAL ADMIN: {motivo}"
        result = conn.execute(
            text(
                """
                INSERT INTO attendance (
                    employee_id, shift_date, turno, entry_at, entry_guard, entry_status, late_reason,
                    vehicle_expected, vehicle_entered, vehicle_front_entry, vehicle_trunk_entry, incident,
                    scheduled_entry_at, scheduled_exit_at, entry_limit_at, exit_early_limit_at, extra_limit_at,
                    entry_tolerance_minutes, exit_tolerance_minutes, extra_after_minutes, late_minutes,
                    early_minutes, extra_minutes, late_justified, early_justified, extra_authorized,
                    review_required, review_status, created_at, updated_at
                ) VALUES (
                    :employee_id, :shift_date, :turno, :entry_at, :entry_guard, :entry_status, :late_reason,
                    :vehicle_expected, :vehicle_entered, '', '', :incident,
                    :scheduled_entry_at, :scheduled_exit_at, :entry_limit_at, :exit_early_limit_at, :extra_limit_at,
                    :entry_tolerance_minutes, :exit_tolerance_minutes, :extra_after_minutes, :late_minutes,
                    0, 0, 0, 0, 0,
                    1, 'Entrada manual activa en planta', :created_at, :updated_at
                ) RETURNING id
                """
            ),
            {
                "employee_id": employee_id,
                "shift_date": shift_date,
                "turno": emp.get("turno") or "Día",
                "entry_at": entry_dt.isoformat(),
                "entry_guard": f"MANUAL / {admin_user['username']}",
                "entry_status": eval_data["status"],
                "late_reason": motivo if eval_data["status"] == "Retardo" else "",
                "vehicle_expected": 1 if emp.get("tiene_vehiculo") else 0,
                "vehicle_entered": 1 if vehiculo == "1" else 0,
                "incident": incident,
                **eval_data,
                "created_at": ts,
                "updated_at": ts,
            },
        )
        record_id = str(result.scalar_one())
        audit(conn, "attendance", record_id, "MANUAL_ENTRY", admin_user["username"], reason=incident)
        log_event(conn, "info", "manual", "manual_entry", f"Entrada manual {employee_id} {entry_dt.isoformat()}", admin_user["username"])

    return RedirectResponse(url="/monitor?manual=entrada_ok", status_code=303)


@app.get("/healthz")
def healthz():
    try:
        with engine.begin() as conn:
            conn.execute(text("SELECT 1")).scalar()
            db_ok = True
    except Exception:
        db_ok = False
    return {
        "ok": db_ok,
        "app": APP_NAME,
        "db": engine.dialect.name,
        "storage": str(DATA_DIR),
        "server_time": now_mx().isoformat(),
    }
